#!/usr/bin/env python3
"""
Vaka Formu v2.xlsx - TÜM HÜCRE ANALİZİ
Bu script Excel'deki tüm dolu hücreleri ve yapıyı çıkarır
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json

wb = load_workbook('templates/VAKA_FORMU_TEMPLATE.xlsx')
ws = wb.active

print("=" * 80)
print("VAKA FORMU v2.xlsx - TAM YAPI ANALİZİ")
print(f"Toplam Satır: {ws.max_row}, Toplam Sütun: {ws.max_column}")
print("=" * 80)

# Birleşik hücre haritası
merged_ranges = {}
for mr in ws.merged_cells.ranges:
    top_left = f"{get_column_letter(mr.min_col)}{mr.min_row}"
    merged_ranges[top_left] = {
        "range": str(mr),
        "rows": mr.max_row - mr.min_row + 1,
        "cols": mr.max_col - mr.min_col + 1
    }

# Tüm dolu hücreleri satır satır listele
all_cells = {}
current_section = "Başlık"

for row in range(1, ws.max_row + 1):
    row_cells = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        coord = f"{get_column_letter(col)}{row}"
        
        if cell.value is not None and str(cell.value).strip():
            value = str(cell.value).strip()
            merged = merged_ranges.get(coord, None)
            
            row_cells.append({
                "cell": coord,
                "value": value[:50],  # İlk 50 karakter
                "merged": merged
            })
    
    if row_cells:
        all_cells[row] = row_cells

# Satır satır yazdır
print("\n" + "=" * 80)
print("SATIR SATIR TÜM DOLU HÜCRELER")
print("=" * 80)

for row_num in sorted(all_cells.keys()):
    print(f"\n--- ROW {row_num} ---")
    for cell_info in all_cells[row_num]:
        merged_info = ""
        if cell_info["merged"]:
            merged_info = f" [BİRLEŞİK: {cell_info['merged']['range']}]"
        print(f"  {cell_info['cell']}: {cell_info['value']}{merged_info}")

# Bölüm bölüm analiz
print("\n" + "=" * 80)
print("BÖLÜM ANALİZİ (Label ve Değer Hücreleri)")
print("=" * 80)

sections = {
    "ÜST BÖLÜM (Row 1-2)": range(1, 3),
    "İSTASYON / SAATLER / HASTA BİLGİLERİ (Row 3-9)": range(3, 10),
    "ÇAĞRI TİPİ / NEDENİ / OLAY YERİ (Row 10-14)": range(10, 15),
    "İLK MUAYENE BULGULARI (Row 15-22)": range(15, 23),
    "ÖN TANI / AÇIKLAMALAR (Row 23)": range(23, 24),
    "SONUÇ / NAKİL (Row 24-29)": range(24, 30),
    "İŞLEMLER / İLAÇLAR (Row 30+)": range(30, ws.max_row + 1),
}

for section_name, row_range in sections.items():
    print(f"\n### {section_name} ###")
    for row_num in row_range:
        if row_num in all_cells:
            for cell_info in all_cells[row_num]:
                merged_info = ""
                if cell_info["merged"]:
                    merged_info = f" [{cell_info['merged']['range']}]"
                print(f"  {cell_info['cell']}: {cell_info['value']}{merged_info}")

# JSON olarak kaydet
output = {
    "max_row": ws.max_row,
    "max_column": ws.max_column,
    "cells": all_cells,
    "merged_ranges": merged_ranges
}

with open('vaka_formu_structure.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print("\n" + "=" * 80)
print("Detaylı yapı 'vaka_formu_structure.json' dosyasına kaydedildi")
print("=" * 80)

