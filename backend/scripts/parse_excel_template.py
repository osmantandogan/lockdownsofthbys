"""
Excel VAKA FORMU şablonunu parse eder ve yapısını çıkarır
"""
import openpyxl
from openpyxl.utils import get_column_letter
import json
import os

def parse_excel_template(xlsx_path):
    """Excel dosyasını parse edip yapısını döndürür"""
    
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    
    result = {
        "sheet_name": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merged_cells": [],
        "cells": [],
        "row_heights": {},
        "column_widths": {}
    }
    
    # Birleşik hücreler
    for merged in ws.merged_cells.ranges:
        result["merged_cells"].append({
            "range": str(merged),
            "min_row": merged.min_row,
            "max_row": merged.max_row,
            "min_col": merged.min_col,
            "max_col": merged.max_col
        })
    
    # Satır yükseklikleri
    for row_idx in range(1, ws.max_row + 1):
        if ws.row_dimensions[row_idx].height:
            result["row_heights"][row_idx] = ws.row_dimensions[row_idx].height
    
    # Sütun genişlikleri
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if ws.column_dimensions[col_letter].width:
            result["column_widths"][col_letter] = ws.column_dimensions[col_letter].width
    
    # Hücreler
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            
            # Sadece değer veya stil içeren hücreleri kaydet
            if cell.value is not None or cell.fill.fgColor.rgb != "00000000":
                cell_data = {
                    "row": row,
                    "col": col,
                    "col_letter": get_column_letter(col),
                    "address": f"{get_column_letter(col)}{row}",
                    "value": str(cell.value) if cell.value else "",
                    "font": {
                        "name": cell.font.name if cell.font else None,
                        "size": cell.font.size if cell.font else None,
                        "bold": cell.font.bold if cell.font else False,
                        "italic": cell.font.italic if cell.font else False,
                        "color": cell.font.color.rgb if cell.font and cell.font.color and cell.font.color.rgb else None
                    },
                    "fill": {
                        "color": cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
                    },
                    "alignment": {
                        "horizontal": cell.alignment.horizontal if cell.alignment else None,
                        "vertical": cell.alignment.vertical if cell.alignment else None,
                        "wrap_text": cell.alignment.wrap_text if cell.alignment else False
                    },
                    "border": {
                        "left": cell.border.left.style if cell.border and cell.border.left else None,
                        "right": cell.border.right.style if cell.border and cell.border.right else None,
                        "top": cell.border.top.style if cell.border and cell.border.top else None,
                        "bottom": cell.border.bottom.style if cell.border and cell.border.bottom else None
                    }
                }
                result["cells"].append(cell_data)
    
    return result

if __name__ == "__main__":
    xlsx_path = os.path.join(os.path.dirname(__file__), "..", "..", "VAKA FORMU.xlsx")
    
    print("Excel dosyasi parse ediliyor...")
    data = parse_excel_template(xlsx_path)
    
    print(f"\n=== GENEL BILGILER ===")
    print(f"Sheet: {data['sheet_name']}")
    print(f"Satirlar: {data['max_row']}")
    print(f"Sutunlar: {data['max_column']}")
    print(f"Birlesik hucre sayisi: {len(data['merged_cells'])}")
    print(f"Dolu hucre sayisi: {len(data['cells'])}")
    
    print(f"\n=== BIRLESIK HUCRELER (ilk 20) ===")
    for m in data['merged_cells'][:20]:
        print(f"  {m['range']}")
    
    print(f"\n=== DEGERLI HUCRELER (ilk 30) ===")
    value_cells = [c for c in data['cells'] if c['value']]
    for c in value_cells[:30]:
        val = c['value'][:40] if len(c['value']) > 40 else c['value']
        print(f"  {c['address']}: {val}")
    
    # JSON olarak kaydet
    output_path = os.path.join(os.path.dirname(__file__), "vaka_formu_structure.json")
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"\nYapi JSON olarak kaydedildi: {output_path}")

