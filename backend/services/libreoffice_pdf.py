#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
LibreOffice-based Excel to PDF Generator
Uses LibreOffice headless to convert Excel files to PDF with perfect formatting
"""

import subprocess
import os
import shutil
import tempfile
import logging
from datetime import datetime
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def populate_excel_with_case_data(excel_path: str, output_path: str, case_data: dict, form_data: dict = None) -> str:
    """
    Excel template'i vaka verileriyle doldurur
    
    Args:
        excel_path: Orijinal Excel template yolu
        output_path: Doldurulmuş Excel'in kaydedileceği yol
        case_data: Backend'den gelen vaka verisi
        form_data: Frontend'den gelen form verisi
    
    Returns:
        str: Doldurulmuş Excel dosyasının yolu
    """
    form_data = form_data or {}
    case_data = case_data or {}
    
    # Excel'i yükle
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Helper fonksiyonlar
    def safe_get(d, *keys, default=''):
        for key in keys:
            if isinstance(d, dict):
                d = d.get(key, {})
            else:
                return default
        return d if d else default
    
    # Tarih formatla
    date_str = form_data.get('date', '')
    if not date_str and case_data.get('created_at'):
        try:
            created_at = case_data['created_at']
            if isinstance(created_at, str):
                date_obj = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            else:
                date_obj = created_at
            date_str = date_obj.strftime('%d.%m.%Y')
        except:
            date_str = ''
    
    # Saat formatla
    call_time = form_data.get('callTime', '')
    if not call_time and case_data.get('created_at'):
        try:
            created_at = case_data['created_at']
            if isinstance(created_at, str):
                date_obj = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            else:
                date_obj = created_at
            call_time = date_obj.strftime('%H:%M')
        except:
            call_time = ''
    
    # Hasta adı
    patient_name = form_data.get('patientName', '')
    if not patient_name:
        first = safe_get(case_data, 'patient', 'name')
        last = safe_get(case_data, 'patient', 'surname')
        patient_name = f"{first} {last}".strip()
    
    # Araç plakası
    vehicle_plate = form_data.get('vehicleType', '') or form_data.get('vehiclePlate', '')
    if not vehicle_plate:
        assigned_team = case_data.get('assigned_team', {})
        if isinstance(assigned_team, dict):
            vehicle_plate = assigned_team.get('vehicle_plate', '')
    
    # Adres
    address = form_data.get('address', '')
    if not address:
        location = case_data.get('location', {})
        if isinstance(location, dict):
            address = location.get('address', '')
    
    # Telefon
    phone = form_data.get('phone', '')
    if not phone:
        phone = safe_get(case_data, 'patient', 'phone') or safe_get(case_data, 'caller', 'phone')
    
    # Şikayet
    complaint = form_data.get('complaint', '')
    if not complaint:
        complaint = safe_get(case_data, 'patient', 'complaint') or case_data.get('description', '')
    
    # Yaş
    age = form_data.get('age', '')
    if not age:
        patient_age = safe_get(case_data, 'patient', 'age')
        if patient_age:
            age = str(patient_age)
    
    # Cinsiyet
    gender = form_data.get('gender', '')
    if not gender:
        gender = safe_get(case_data, 'patient', 'gender')
    
    # TC Kimlik No
    tc_no = form_data.get('tcNo', '')
    if not tc_no:
        tc_no = safe_get(case_data, 'patient', 'tc_no')
    
    # Extended form verileri
    extended_form = form_data.get('extended_form', {}) or {}
    vehicle_info = form_data.get('vehicle_info', {}) or {}
    
    # Kronik hastalıklar
    chronic_diseases = extended_form.get('chronicDiseases', '') or form_data.get('chronicDiseases', '')
    
    # Triyaj kodu
    triage_code = extended_form.get('triageCode', '') or form_data.get('triageCode', '')
    
    # İstasyon kodu
    station_code = extended_form.get('stationCode', '') or form_data.get('stationCode', '')
    
    # Nakledilen hastane
    hospital_name = extended_form.get('hospitalName', '') or form_data.get('hospitalName', '')
    
    # Vakayı veren kurum
    referral_source = extended_form.get('referralSource', '') or form_data.get('callerOrganization', '')
    
    # Nakil mesafe
    transfer_type = extended_form.get('transferType', '') or form_data.get('transferType', '')
    
    # KM bilgileri
    start_km = vehicle_info.get('startKm', '') or form_data.get('startKm', '')
    end_km = vehicle_info.get('endKm', '') or form_data.get('endKm', '')
    
    # Kaza araç plakaları
    accident_vehicles = extended_form.get('accidentVehicles', []) or form_data.get('accidentVehicles', [])
    if isinstance(accident_vehicles, list):
        accident_plate_1 = accident_vehicles[0] if len(accident_vehicles) > 0 else ''
        accident_plate_2 = accident_vehicles[1] if len(accident_vehicles) > 1 else ''
        accident_plate_3 = accident_vehicles[2] if len(accident_vehicles) > 2 else ''
        accident_plate_4 = accident_vehicles[3] if len(accident_vehicles) > 3 else ''
    else:
        accident_plate_1 = accident_plate_2 = accident_plate_3 = accident_plate_4 = ''
    
    # Alerjiler
    allergies = form_data.get('allergies', '')
    
    # Vital bulgular
    vitals = form_data.get('vitals', {}) or {}
    vital_signs = form_data.get('vital_signs', []) or []
    clinical_obs = form_data.get('clinical_obs', {}) or {}
    
    # İlk ölçüm
    if vital_signs and len(vital_signs) > 0:
        first_vital = vital_signs[0] if isinstance(vital_signs[0], dict) else {}
        blood_pressure = first_vital.get('bp', '') or vitals.get('bloodPressure', '') or form_data.get('bloodPressure', '')
        pulse = first_vital.get('pulse', '') or vitals.get('pulse', '') or form_data.get('pulse', '')
        spo2 = first_vital.get('spo2', '') or vitals.get('spo2', '') or form_data.get('spo2', '')
        temperature = first_vital.get('temp', '') or vitals.get('temperature', '') or form_data.get('temperature', '')
        respiration = first_vital.get('respiration', '') or vitals.get('respiration', '') or form_data.get('respiration', '')
    else:
        blood_pressure = vitals.get('bloodPressure', '') or form_data.get('bloodPressure', '')
        pulse = vitals.get('pulse', '') or form_data.get('pulse', '')
        spo2 = vitals.get('spo2', '') or form_data.get('spo2', '')
        temperature = vitals.get('temperature', '') or form_data.get('temperature', '')
        respiration = vitals.get('respiration', '') or form_data.get('respiration', '')
    
    # GCS değerleri (clinical_obs'dan)
    gcs_motor = clinical_obs.get('motorResponse', '') or form_data.get('gcsMotor', '')
    gcs_verbal = clinical_obs.get('verbalResponse', '') or form_data.get('gcsVerbal', '')
    gcs_eye = clinical_obs.get('eyeOpening', '') or form_data.get('gcsEye', '')
    
    # Kan şekeri
    blood_sugar = extended_form.get('bloodSugar', '') or form_data.get('bloodSugar', '')
    
    gcs = vitals.get('gcs', '') or form_data.get('gcs', '')
    
    # Hücre eşlemesi - Vaka formu v2.xlsx yapısına göre
    # NOT: Değerler label'ın yanındaki DEĞER hücrelerine yazılmalı, label'a değil!
    # Birleşik hücrelerin sol-üst köşesine yazılır
    cell_mapping = {
        # ÜST BÖLÜM - ATN NO, KM bilgileri (Row 2 - değer satırı)
        'U2': case_data.get('case_number', ''),  # ATN NO değeri (U1 label, U2 değer)
        'W2': start_km,  # BAŞLANGIÇ KM değeri
        'Y2': end_km,  # BİTİŞ KM değeri
        
        # İSTASYON BÖLÜMÜ - Değer hücreleri C sütununda
        'C4': form_data.get('healmedyProtocol') or case_data.get('case_number', ''),  # PROTOKOL NO değeri
        'C5': date_str,  # TARİH değeri (A5 label, C5 değer - eğer ayrıysa)
        'C6': station_code,  # KODU değeri
        'C7': vehicle_plate,  # PLAKA değeri
        'C8': address,  # HASTANIN ALINDIĞI ADRES değeri
        'C9': referral_source,  # VAKAYI VEREN KURUM değeri
        
        # SAATLER BÖLÜMÜ - Değer hücreleri I sütununda
        'I4': call_time,  # ÇAĞRI SAATİ değeri
        'I5': form_data.get('arrivalTime', ''),  # OLAY YERİNE VARIŞ değeri
        'I6': form_data.get('patientArrivalTime', ''),  # HASTAYA VARIŞ değeri
        'I7': form_data.get('departureTime', ''),  # OLAY YERİNDEN AYRILIŞ değeri
        'I8': form_data.get('hospitalArrivalTime', ''),  # HASTANEYE VARIŞ değeri
        'I9': form_data.get('returnTime', ''),  # İSTASYONA DÖNÜŞ değeri
        
        # HASTA BİLGİLERİ BÖLÜMÜ - Değer hücreleri M sütununda
        'M4': patient_name,  # ADI SOYADI değeri (K4 label, M4 değer)
        'M5': address,  # ADRESİ değeri (aslında K5:L7 birleşik, M5 değer)
        'M8': tc_no,  # T.C. KİMLİK NO değeri
        'M9': phone,  # TELEFON değeri
        
        # CİNSİYET / YAŞ / DURUMU - T sütunu değerler için
        'T9': age,  # YAŞ değeri (S9 label, T9 değer)
        'T8': form_data.get('birthDate', ''),  # Doğum Tarihi değeri
        
        # KRONİK HASTALIKLAR - X4 değer hücresi
        'X4': chronic_diseases,  # KRONİK HASTALIKLAR değeri
        
        # HASTANIN ŞİKAYETİ - X7 değer hücresi
        'X7': complaint,  # HASTANIN ŞİKAYETİ değeri
        
        # VİTAL BULGULAR (Row 17-22)
        # Tansiyon değerleri - H17 birleşik hücrede sistol/diastol
        'H17': blood_pressure,  # Tansiyon (ör: 120/80)
        'K17': pulse,  # NABIZ değeri
        'M17': respiration,  # SOLUNUM değeri
        'I19': spo2,  # SPO2 (%) değeri
        'Y19': temperature,  # ATEŞ (°C) değeri
        
        # GKS (Glasgow Koma Skalası)
        'O17': gcs_motor,  # Motor değeri
        'R17': gcs_verbal,  # Verbal değeri
        'U17': gcs_eye,  # Göz açma değeri
        
        # KAN ŞEKERİ
        'Z17': blood_sugar,  # Kan şekeri Mg/dL
        
        # ÖN TANI (Row 23)
        'B23': form_data.get('diagnosis', ''),  # ÖN TANI değeri
        
        # AÇIKLAMALAR
        'I23': form_data.get('notes', ''),  # AÇIKLAMALAR değeri
        
        # NAKLEDİLEN HASTANE
        'L24': hospital_name,  # NAKLEDİLEN HASTANE değeri
        
        # KAZAYA KARIŞAN ARAÇ PLAKA
        'P25': accident_plate_1,
        'P26': accident_plate_2,
        'P27': accident_plate_3,
        'P28': accident_plate_4,
        
        # CPR bilgileri
        'U25': form_data.get('cprStartTime', ''),  # CPR BAŞLAMA ZAMANI
        'U26': form_data.get('cprEndTime', ''),  # CPR BIRAKMA ZAMANI
        'U27': form_data.get('cprStopReason', ''),  # CPR BIRAKMA NEDENİ
    }
    
    # Birleşik hücre haritası oluştur
    merged_cells_map = {}
    for mr in ws.merged_cells.ranges:
        for row in range(mr.min_row, mr.max_row + 1):
            for col in range(mr.min_col, mr.max_col + 1):
                cell_coord = ws.cell(row=row, column=col).coordinate
                top_left = ws.cell(row=mr.min_row, column=mr.min_col).coordinate
                merged_cells_map[cell_coord] = top_left
    
    # Hücrelere değer yaz
    for cell_ref, value in cell_mapping.items():
        if value:
            try:
                target_cell = merged_cells_map.get(cell_ref, cell_ref)
                ws[target_cell] = str(value)
            except Exception as e:
                logger.warning(f"Hücreye yazılamadı {cell_ref}: {e}")
    
    # Checkboxları işle - Vaka formu v2.xlsx yapısına göre
    
    # CİNSİYET - T4 ve T6 checkbox hücreleri
    if gender:
        if gender.lower() in ['erkek', 'male', 'e', 'm']:
            try:
                ws['T4'] = 'X'  # ERKEK checkbox
            except:
                pass
        elif gender.lower() in ['kadın', 'kadin', 'female', 'k', 'f']:
            try:
                ws['T6'] = 'X'  # KADIN checkbox
            except:
                pass
    
    # DURUMU / TRİYAJ KODU - V4-V8 checkbox hücreleri
    # triage_code zaten yukarıda tanımlandı (extended_form'dan)
    if triage_code:
        triage_cells = {
            'kirmizi': 'V4',  # KIRMIZI KOD
            'sari': 'V5',  # SARI KOD
            'yesil': 'V6',  # YEŞİL KOD
            'siyah': 'V7',  # SİYAH KOD
            'sosyal': 'V8',  # SOSYAL ENDİKASYON
        }
        target_cell = triage_cells.get(triage_code.lower().replace(' ', '_').replace('ı', 'i'))
        if target_cell:
            try:
                ws[target_cell] = 'X'
            except:
                pass
    
    # ÇAĞRI TİPİ (A11-A13)
    call_type = form_data.get('callType', '')
    if call_type:
        call_type_cells = {
            'telsiz': 'B11',
            'telefon': 'B12',
            'diger': 'B13',
        }
        target_cell = call_type_cells.get(call_type.lower())
        if target_cell:
            try:
                ws[target_cell] = 'X'
            except:
                pass
    
    # ÇAĞRI NEDENİ - Ana kategori
    call_reason = extended_form.get('callReason', '') or form_data.get('callReason', '') or case_data.get('case_type', '')
    if call_reason:
        reason_cells = {
            'medikal': 'F11',
            'trafik_kazasi': 'F12',
            'trafik_kaza': 'F12',
            'diger_kaza': 'F13',
            'is_kazasi': 'F14',
        }
        target_cell = reason_cells.get(call_reason.lower().replace(' ', '_').replace('ı', 'i'))
        if target_cell:
            try:
                ws[target_cell] = 'X'
            except:
                pass
    
    # ÇAĞRI NEDENİ DETAY - Çoklu seçim (extended_form'dan)
    call_reason_detail = extended_form.get('callReasonDetail', []) or form_data.get('callReasonDetail', [])
    if call_reason_detail and isinstance(call_reason_detail, list):
        detail_cells = {
            'yangin': 'I11',
            'intihar': 'I12',
            'kimyasal': 'I13',
            'allerji': 'I14',
            'elektrik_carp': 'K11',
            'elektrik_carpmasi': 'K11',
            'atesli_silah': 'K12',
            'bogulma': 'K13',
            'kesici_delici': 'K14',
            'dusme': 'M11',
            'alkol_ilac': 'M12',
            'kunt_trav': 'M13',
            'kunt_travma': 'M13',
            'yanik': 'M14',
            'lpg': 'O11',
            'tedbir': 'O12',
            'protokol': 'O13',
        }
        for detail in call_reason_detail:
            target_cell = detail_cells.get(detail.lower().replace(' ', '_').replace('ı', 'i'))
            if target_cell:
                try:
                    ws[target_cell] = 'X'
                except:
                    pass
    
    # OLAY YERİ - Tek seçim (sceneType)
    scene_type = extended_form.get('sceneType', '') or form_data.get('sceneLocation', '')
    location_cells = {
        'ev': 'Q11',
        'aracta': 'S11',
        'stadyum': 'U11',
        'saglik_kurumu': 'W11',
        'yaya': 'Q12',
        'buro': 'S12',
        'huzurevi': 'U12',
        'resmi_daire': 'W12',
        'suda': 'Q13',
        'fabrika': 'S13',
        'cami': 'U13',
        'egitim_kurumu': 'W13',
        'arazi': 'Q14',
        'sokak': 'S14',
        'yurt': 'U14',
        'spor_salonu': 'W14',
    }
    if scene_type:
        target_cell = location_cells.get(scene_type.lower().replace(' ', '_').replace('ı', 'i'))
        if target_cell:
            try:
                ws[target_cell] = 'X'
            except:
                pass
    
    # OLAY YERİ DETAY - Çoklu seçim (incidentLocation)
    incident_locations = extended_form.get('incidentLocation', []) or form_data.get('incidentLocation', [])
    if incident_locations and isinstance(incident_locations, list):
        for loc in incident_locations:
            target_cell = location_cells.get(loc.lower().replace(' ', '_').replace('ı', 'i'))
            if target_cell:
                try:
                    ws[target_cell] = 'X'
                except:
                    pass
    
    # PUPİLLER (B17-B22) - clinical_obs'dan al
    pupils = clinical_obs.get('pupils', '') or form_data.get('pupils', '')
    if pupils:
        pupil_cells = {
            'normal': 'C17',
            'miyotik': 'C18',
            'midriatik': 'C19',
            'anizokorik': 'C20',
            'reaksiyon_yok': 'C21',
            'reak_yok': 'C21',
            'fiks_dilate': 'C22',
        }
        target_cell = pupil_cells.get(pupils.lower().replace(' ', '_'))
        if target_cell:
            try:
                ws[target_cell] = 'X'
            except:
                pass
    
    # DERİ (D17-D22) - clinical_obs'dan al
    skin = clinical_obs.get('skin', '') or form_data.get('skin', '')
    if skin:
        skin_cells = {
            'normal': 'E17',
            'soluk': 'E18',
            'siyanotik': 'E19',
            'siyatonik': 'E19',  # typo desteği
            'hiperemik': 'E20',
            'ikterik': 'E21',
            'terli': 'E22',
        }
        target_cell = skin_cells.get(skin.lower())
        if target_cell:
            try:
                ws[target_cell] = 'X'
            except:
                pass
    
    # NABIZ TİPİ - clinical_obs'dan al
    pulse_type = clinical_obs.get('pulseType', '') or form_data.get('pulseType', '')
    if pulse_type:
        pulse_type_cells = {
            'duzenli': 'K19',
            'düzenli': 'K19',
            'ritmik': 'K20',
            'filiform': 'K21',
            'alinmiyor': 'K22',
            'alınmıyor': 'K22',
        }
        target_cell = pulse_type_cells.get(pulse_type.lower())
        if target_cell:
            try:
                ws[target_cell] = 'X'
            except:
                pass
    
    # SOLUNUM TİPİ - clinical_obs'dan al
    respiration_type = clinical_obs.get('respirationType', '') or form_data.get('respirationType', '')
    if respiration_type:
        respiration_cells = {
            'duzenli': 'M19',
            'düzenli': 'M19',
            'duzensiz': 'M20',
            'düzensiz': 'M20',
            'dispne': 'M21',
            'yok': 'M22',
        }
        target_cell = respiration_cells.get(respiration_type.lower())
        if target_cell:
            try:
                ws[target_cell] = 'X'
            except:
                pass
    
    # SONUÇ (B25-D29)
    result = form_data.get('result', '')
    if result:
        result_cells = {
            'yerinde_mudahale': 'C25',
            'hastaneye_nakil': 'C26',
            'hastaneler_arasi_nakil': 'C27',
            'tibbi_tetkik_icin_nakil': 'C28',
            'eve_nakil': 'C29',
            'ex_terinde_birakildi': 'E25',
            'ex_morga_nakil': 'E26',
            'nakil_reddi': 'E27',
            'diger_ulasilan': 'E28',
            'gorev_iptali': 'E29',
            'baska_aracla_nakil': 'G25',
            'tlfla_bsk_aracla_nakil': 'G26',
            'asilsiz_ihbar': 'G27',
            'yaralanan_yok': 'G28',
            'olay_yerinde_bekleme': 'G29',
        }
        target_cell = result_cells.get(result.lower().replace(' ', '_').replace('ı', 'i'))
        if target_cell:
            try:
                ws[target_cell] = 'X'
            except:
                pass
    
    # ADLİ VAKA (Q29: EVET, S29: HAYIR)
    is_judicial = form_data.get('isJudicialCase', '')
    if is_judicial:
        if str(is_judicial).lower() in ['true', 'evet', '1', 'yes']:
            try:
                ws['R29'] = 'X'  # EVET
            except:
                pass
        elif str(is_judicial).lower() in ['false', 'hayir', '0', 'no']:
            try:
                ws['T29'] = 'X'  # HAYIR
            except:
                pass
    
    # NAKIL TÜRÜ - İLÇE İÇİ/DIŞI/İL DIŞI (K27-K29)
    # transfer_type zaten yukarıda tanımlandı (extended_form'dan)
    if transfer_type:
        transfer_cells = {
            'ilce_ici': 'L27',
            'ilce_disi': 'L28',
            'il_disi': 'L29',
        }
        target_cell = transfer_cells.get(transfer_type.lower().replace(' ', '_'))
        if target_cell:
            try:
                ws[target_cell] = 'X'
            except:
                pass
    
    # İŞLEMLER - Checkbox ve Adet (procedures)
    procedures_data = form_data.get('procedures', {})
    if procedures_data and isinstance(procedures_data, dict):
        # İşlem-hücre eşlemesi (checkbox, adet)
        procedure_cells = {
            'Muayene (Acil)': ('A31', 'G31'),
            'Enjeksiyon IM': ('A34', 'G34'),
            'Enjeksiyon IV': ('A35', 'G35'),
            'Damar yolu açılması': ('A38', 'G38'),
            'Pansuman (küçük)': ('A42', 'G42'),
            'Servikal collar uygulama': ('A52', 'G52'),
            'Sırt tahtası uygulaması': ('A55', 'G55'),
            'CPR uygulaması': ('A57', 'G57'),
            'EKG Uygulaması': ('A58', 'G58'),
            'Defibrilasyon': ('A59', 'G59'),
            'Monitörizasyon': ('A61', 'G61'),
            'Kanama kontrolü': ('A62', 'G62'),
            'Balon valf maske uygulaması': ('H32', 'M32'),
            'Aspirasyon uygulaması': ('H33', 'M33'),
            'Entübasyon uygulaması': ('H35', 'M35'),
            'Mekanik ventilasyon': ('H36', 'M36'),
            'Oksijen inhalasyon tedavisi 1 Saat': ('H37', 'M37'),
            'Nebulizatör ile ilaç uygulama': ('H38', 'M38'),
            'Kan şekeri ölçümü': ('H40', 'M40'),
        }
        
        for proc_name, proc_value in procedures_data.items():
            if proc_name in procedure_cells:
                checkbox_cell, adet_cell = procedure_cells[proc_name]
                
                # Yeni format: {checked: bool, adet: int}
                if isinstance(proc_value, dict):
                    if proc_value.get('checked'):
                        try:
                            ws[checkbox_cell] = 'X'
                            adet = proc_value.get('adet', 1)
                            if adet > 1:
                                ws[adet_cell] = str(adet)
                        except:
                            pass
                # Eski format: boolean
                elif proc_value:
                    try:
                        ws[checkbox_cell] = 'X'
                    except:
                        pass
    
    # MALZEMELER - Checkbox ve Adet (materials)
    materials_data = form_data.get('materials', {})
    if materials_data and isinstance(materials_data, dict):
        # Malzeme-hücre eşlemesi (checkbox, adet)
        material_cells = {
            'Enjektör 1-2 cc': ('U32', 'Z32'),
            'Enjektör 5 cc': ('U33', 'Z33'),
            'Enjektör 10-20 cc': ('U34', 'Z34'),
            'Monitör pedi (EKG elektrodu)': ('U35', 'Z35'),
            'I.V. katater (No: 14-22)': ('U36', 'Z36'),
            'I.V. katater (No: 24)': ('U37', 'Z37'),
            'Serum seti': ('U38', 'Z38'),
            'Steril eldiven': ('U39', 'Z39'),
            'Cerrahi eldiven': ('U40', 'Z40'),
            'Sponç': ('U41', 'Z41'),
            'Sargı bezi': ('U42', 'Z42'),
            'İdrar torbası': ('U43', 'Z43'),
            'Bistüri ucu': ('U44', 'Z44'),
            'Entübasyon tüpü (Balonlu)': ('U45', 'Z45'),
            'Entübasyon tüpü (Balonsuz)': ('U46', 'Z46'),
            'Airway': ('U47', 'Z47'),
            'Foley sonda': ('U48', 'Z48'),
            'Nazo gastrik sonda': ('U49', 'Z49'),
            'Atravmatik ipek (3/0)': ('U50', 'Z50'),
            'Atravmatik kat-küt (3/0)': ('U51', 'Z51'),
            'Doğum seti': ('U52', 'Z52'),
            'Yanık battaniyesi': ('U53', 'Z53'),
            'O2 Maskesi hazneli erişkin': ('U54', 'Z54'),
            'O2 Maskesi hazneli pediatrik': ('U55', 'Z55'),
            'O2 Kanülü nazal erişkin': ('U56', 'Z56'),
            'O2 Kanülü nazal pediatrik': ('U57', 'Z57'),
            'Flaster': ('U58', 'Z58'),
            'Servikal collar': ('U59', 'Z59'),
            'Elastik bandaj': ('U60', 'Z60'),
            'Etil Chloride Sprey': ('U61', 'Z61'),
            'O2 Maskesi haznesiz erişkin': ('U62', 'Z62'),
            'O2 Maskesi haznesiz pediatrik': ('U63', 'Z63'),
        }
        
        for mat_name, mat_value in materials_data.items():
            if mat_name in material_cells:
                checkbox_cell, adet_cell = material_cells[mat_name]
                
                # Yeni format: {checked: bool, adet: int}
                if isinstance(mat_value, dict):
                    if mat_value.get('checked'):
                        try:
                            ws[checkbox_cell] = 'X'
                            adet = mat_value.get('adet', 1)
                            if adet > 1:
                                ws[adet_cell] = str(adet)
                        except:
                            pass
                # Eski format: boolean
                elif mat_value:
                    try:
                        ws[checkbox_cell] = 'X'
                    except:
                        pass
    
    # Sayfa düzenini portrait (dikey) ve fit-to-page yap
    try:
        from openpyxl.worksheet.page import PageMargins, PrintPageSetup
        
        # Portrait (dikey) moduna ayarla
        ws.page_setup.orientation = 'portrait'
        
        # Sayfaya sığdır (1 sayfa genişlik, 1 sayfa yükseklik)
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        
        # Kenar boşluklarını küçült
        ws.page_margins = PageMargins(
            left=0.15,
            right=0.15,
            top=0.15,
            bottom=0.15,
            header=0.1,
            footer=0.1
        )
        
        # Yazdırma alanını ayarla (tüm kullanılan alan)
        ws.print_area = f"A1:Z{ws.max_row}"
        
        # Ölçeklendirme ayarı
        ws.page_setup.scale = 50  # %50 ölçek - tek sayfaya sığdırmak için
        
        logger.info("Sayfa düzeni portrait ve fit-to-page olarak ayarlandı")
    except Exception as e:
        logger.warning(f"Sayfa düzeni ayarlanamadı: {e}")
    
    # Kaydet
    wb.save(output_path)
    logger.info(f"Excel dolduruldu: {output_path}")
    
    return output_path


def convert_excel_to_pdf_libreoffice(excel_path: str, output_dir: str) -> str:
    """
    LibreOffice Headless kullanarak Excel'i PDF'e çevirir
    
    Args:
        excel_path: Excel dosyasının yolu
        output_dir: PDF'in kaydedileceği dizin
    
    Returns:
        str: Oluşturulan PDF dosyasının yolu
    """
    try:
        # LibreOffice komutu
        cmd = [
            'soffice',
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', output_dir,
            excel_path
        ]
        
        logger.info(f"LibreOffice çalıştırılıyor: {' '.join(cmd)}")
        
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=60  # 60 saniye timeout
        )
        
        if result.returncode != 0:
            logger.error(f"LibreOffice hatası: {result.stderr}")
            raise Exception(f"LibreOffice dönüşüm hatası: {result.stderr}")
        
        # PDF dosya adını hesapla
        excel_filename = os.path.basename(excel_path)
        pdf_filename = os.path.splitext(excel_filename)[0] + '.pdf'
        pdf_path = os.path.join(output_dir, pdf_filename)
        
        if not os.path.exists(pdf_path):
            raise Exception(f"PDF oluşturulamadı: {pdf_path}")
        
        logger.info(f"PDF oluşturuldu: {pdf_path}")
        return pdf_path
        
    except subprocess.TimeoutExpired:
        logger.error("LibreOffice zaman aşımı")
        raise Exception("PDF dönüşümü zaman aşımına uğradı")
    except FileNotFoundError:
        logger.error("LibreOffice bulunamadı")
        raise Exception("LibreOffice kurulu değil. Dockerfile'da 'libreoffice-calc' yüklendiğinden emin olun.")


def generate_case_pdf_with_libreoffice(
    template_path: str,
    case_data: dict,
    form_data: dict = None,
    output_filename: str = None
) -> str:
    """
    Vaka verilerini kullanarak Excel template'ten PDF oluşturur
    
    Args:
        template_path: Excel template dosyasının yolu
        case_data: Vaka verileri
        form_data: Form verileri (opsiyonel)
        output_filename: Çıktı dosya adı (opsiyonel)
    
    Returns:
        str: Oluşturulan PDF dosyasının yolu
    """
    # Temp dizin oluştur
    backend_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    temp_dir = os.path.join(backend_dir, "temp")
    os.makedirs(temp_dir, exist_ok=True)
    
    case_id = case_data.get('_id', case_data.get('id', 'unknown'))
    case_number = case_data.get('case_number', case_id)
    
    # Geçici Excel dosyası oluştur
    temp_excel = os.path.join(temp_dir, f"temp_case_{case_id}.xlsx")
    
    try:
        # Excel'i verilerle doldur
        populate_excel_with_case_data(template_path, temp_excel, case_data, form_data)
        
        # LibreOffice ile PDF'e çevir
        pdf_path = convert_excel_to_pdf_libreoffice(temp_excel, temp_dir)
        
        # İstenen dosya adına yeniden adlandır
        if output_filename:
            final_pdf_path = os.path.join(temp_dir, output_filename)
            if os.path.exists(final_pdf_path):
                os.remove(final_pdf_path)
            shutil.move(pdf_path, final_pdf_path)
            pdf_path = final_pdf_path
        
        return pdf_path
        
    finally:
        # Geçici Excel dosyasını temizle
        if os.path.exists(temp_excel):
            try:
                os.remove(temp_excel)
            except:
                pass


def check_libreoffice_installed() -> bool:
    """LibreOffice'in kurulu olup olmadığını kontrol eder"""
    try:
        result = subprocess.run(
            ['soffice', '--version'],
            capture_output=True,
            text=True,
            timeout=10
        )
        return result.returncode == 0
    except:
        return False

