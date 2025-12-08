"""
Excel Export Service - Orijinal VAKA FORMU şablonunu kullanarak doldurulmuş Excel oluşturur
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
from io import BytesIO
import os

# Template dosyasının yolu
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', 'templates', 'VAKA_FORMU_TEMPLATE.xlsx')


def format_date(date_str):
    """Tarih formatla"""
    if not date_str:
        return ""
    try:
        if isinstance(date_str, datetime):
            return date_str.strftime("%d.%m.%Y")
        dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
        return dt.strftime("%d.%m.%Y")
    except:
        return str(date_str)[:10] if date_str else ""


def format_time(time_str):
    """Saat formatla"""
    if not time_str:
        return ""
    try:
        if isinstance(time_str, datetime):
            return time_str.strftime("%H:%M")
        # HH:MM formatında ise direkt döndür
        if len(str(time_str)) == 5 and ':' in str(time_str):
            return str(time_str)
        dt = datetime.fromisoformat(time_str.replace('Z', '+00:00'))
        return dt.strftime("%H:%M")
    except:
        return str(time_str)[:5] if time_str else ""


def set_checkbox(ws, cell_ref, is_checked):
    """Checkbox işaretle (☑ veya ☐)"""
    ws[cell_ref] = "☑" if is_checked else "☐"


def export_case_to_excel(case_data: dict, 
                          medical_form: dict = None,
                          vital_signs: list = None,
                          clinical_obs: dict = None,
                          cpr_data: dict = None,
                          procedures: list = None,
                          medications: list = None,
                          materials: list = None,
                          signatures: dict = None) -> BytesIO:
    """
    Vaka verilerini Excel şablonuna doldurur ve BytesIO döndürür
    """
    
    # Şablonu yükle
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active
    
    # Tüm merged cell'leri unmerge et (değer yazabilmek için)
    # Önce tüm merged range'leri kopyala (iteration sırasında değiştiremeyiz)
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        try:
            ws.unmerge_cells(str(merged_range))
        except:
            pass
    
    # Varsayılan değerler
    medical_form = medical_form or {}
    vital_signs = vital_signs or []
    clinical_obs = clinical_obs or {}
    cpr_data = cpr_data or {}
    procedures = procedures or []
    medications = medications or []
    materials = materials or []
    signatures = signatures or {}
    
    # =====================
    # TEMEL BİLGİLER (Artık tüm cell'ler unmerged)
    # =====================
    
    # Protokol No
    ws['E9'] = case_data.get('case_number', '')
    
    # Tarih
    ws['E11'] = format_date(case_data.get('created_at', ''))
    
    # Plaka
    vehicle = case_data.get('assigned_team', {}).get('vehicle', '')
    ws['E13'] = vehicle
    
    # Hastanın Alındığı Adres
    location = case_data.get('location', {})
    address = location.get('address', '')
    ws['G15'] = address
    
    # Başlangıç KM
    vehicle_info = case_data.get('vehicle_info', {})
    ws['X3'] = vehicle_info.get('start_km', '')
    
    # Bitiş KM
    ws['AA3'] = vehicle_info.get('end_km', '')
    
    # =====================
    # SAATLER
    # =====================
    time_info = case_data.get('time_info', {})
    
    ws['J9'] = format_time(time_info.get('call_time', ''))
    ws['J10'] = format_time(time_info.get('arrival_time', ''))
    ws['J11'] = format_time(time_info.get('patient_arrival_time', ''))
    ws['J12'] = format_time(time_info.get('departure_time', ''))
    ws['J13'] = format_time(time_info.get('hospital_arrival_time', ''))
    ws['J14'] = format_time(time_info.get('return_time', ''))
    
    # =====================
    # ÇAĞRI TİPİ
    # =====================
    call_type = case_data.get('call_type', '')
    
    ws['F17'] = "☑" if call_type == 'telsiz' else "☐"
    ws['F18'] = "☑" if call_type == 'telefon' else "☐"
    ws['F19'] = "☑" if call_type == 'diger' else "☐"
    
    # =====================
    # ÇAĞRI NEDENİ (Checkboxlar)
    # =====================
    call_reason = case_data.get('call_reason', '')
    
    # Medikal (H17)
    ws['H17'] = "☑" if call_reason == 'medikal' else "☐"
    
    # Trafik Kaz (H18)
    ws['H18'] = "☑" if call_reason == 'trafik' else "☐"
    
    # Diğer Kaza (H19)
    ws['H19'] = "☑" if call_reason == 'kaza' else "☐"
    
    # İş Kazası (H20)
    ws['H20'] = "☑" if call_reason == 'is_kazasi' else "☐"
    
    # Yangın (I17)
    ws['I17'] = "☑" if call_reason == 'yangin' else "☐"
    
    # İntihar (I18)
    ws['I18'] = "☑" if call_reason == 'intihar' else "☐"
    
    # =====================
    # HASTA BİLGİLERİ
    # =====================
    patient = case_data.get('patient', {})
    
    # Ad Soyad
    name = patient.get('name', '')
    surname = patient.get('surname', '')
    full_name = f"{name} {surname}".strip()
    ws['N9'] = full_name
    
    # Adres
    patient_address = patient.get('address', address)
    ws['N10'] = patient_address
    
    # Telefon
    caller = case_data.get('caller', {})
    phone = caller.get('phone', patient.get('phone', ''))
    ws['N14'] = phone
    
    # T.C. Kimlik
    tc_no = patient.get('tc_no', '')
    ws['V15'] = tc_no
    
    # Yaş
    age = patient.get('age', '')
    ws['U13'] = age
    
    # Cinsiyet
    gender = patient.get('gender', '').lower()
    ws['U9'] = "☑" if gender == 'erkek' or gender == 'male' else "☐"
    ws['U11'] = "☑" if gender == 'kadın' or gender == 'kadin' or gender == 'female' else "☐"
    
    # =====================
    # DURUM/TRİYAJ
    # =====================
    priority = case_data.get('priority', '').lower()
    
    ws['X9'] = "☑" if priority == 'critical' or priority == 'kirmizi' else "☐"
    ws['X10'] = "☑" if priority == 'high' or priority == 'sari' else "☐"
    ws['X11'] = "☑" if priority == 'medium' or priority == 'yesil' else "☐"
    ws['X12'] = "☑" if priority == 'low' or priority == 'siyah' else "☐"
    
    # =====================
    # KRONİK HASTALIKLAR
    # =====================
    chronic = case_data.get('chronic_diseases', '') or medical_form.get('chronic_diseases', '')
    ws['Y8'] = chronic
    
    # HASTANIN ŞİKAYETİ
    complaint = patient.get('complaint', '') or case_data.get('complaint', '')
    ws['Y11'] = complaint
    
    # =====================
    # VİTAL BULGULAR
    # =====================
    if vital_signs:
        for i, vs in enumerate(vital_signs[:3]):
            row = 23 + i
            
            # Saat
            ws[f'H{row}'] = format_time(vs.get('time', ''))
            
            # Tansiyon
            bp = vs.get('blood_pressure', '')
            if bp:
                parts = str(bp).split('/')
                if len(parts) == 2:
                    ws[f'I{row}'] = parts[0]
                    ws[f'K{row}'] = parts[1]
            
            # Nabız
            ws[f'N{row}'] = vs.get('pulse', '')
            
            # Solunum
            ws[f'P{row}'] = vs.get('respiration', '')
            
            # SpO2
            ws[f'H{row + 2}'] = vs.get('spo2', '')
    
    # =====================
    # GLASGOW SKALASI
    # =====================
    gks = clinical_obs.get('gks', {})
    
    # Motor değeri checkbox
    motor = gks.get('motor', 0)
    for m in range(1, 7):
        row = 28 - m + 1
        ws[f'R{row}'] = "☑" if motor == m else ""
    
    # Verbal değeri checkbox
    verbal = gks.get('verbal', 0)
    for v in range(1, 6):
        row = 27 - v + 1
        ws[f'U{row}'] = "☑" if verbal == v else ""
    
    # Göz açma değeri checkbox
    eye = gks.get('eye', 0)
    for e in range(1, 5):
        row = 26 - e + 1
        ws[f'X{row}'] = "☑" if eye == e else ""
    
    # GKS Puanı
    gks_total = gks.get('total', motor + verbal + eye)
    ws['W28'] = gks_total
    
    # =====================
    # KAN ŞEKERİ ve VÜCUT SICAKLIĞI
    # =====================
    ws['AC22'] = case_data.get('blood_sugar', '')
    ws['AC26'] = case_data.get('body_temperature', '')
    
    # =====================
    # PUPİLLER
    # =====================
    pupils = clinical_obs.get('pupils', '')
    pupil_map = {
        'normal': 'E23', 'miyotik': 'E24', 'midriatik': 'E25',
        'anizokorik': 'E26', 'reaktif_yok': 'E27', 'fiks_dilate': 'E28'
    }
    for key, cell in pupil_map.items():
        ws[cell] = "☑" if pupils.lower() == key else "☐"
    
    # =====================
    # DERİ
    # =====================
    skin = clinical_obs.get('skin', '')
    skin_map = {
        'normal': 'G23', 'soluk': 'G24', 'siyanotik': 'G25',
        'hiperemik': 'G26', 'ikterik': 'G27', 'terli': 'G28'
    }
    for key, cell in skin_map.items():
        ws[cell] = "☑" if skin.lower() == key else "☐"
    
    # =====================
    # ÖN TANI
    # =====================
    diagnoses = medical_form.get('diagnoses', [])
    diagnosis_text = ', '.join([d.get('description', d.get('code', '')) for d in diagnoses]) if diagnoses else ''
    ws['F29'] = diagnosis_text
    
    # =====================
    # VAKAYI VEREN KURUM
    # =====================
    ws['D31'] = case_data.get('company', '') or case_data.get('referring_institution', '')
    
    # =====================
    # SONUÇ (Checkboxlar)
    # =====================
    result = case_data.get('case_result', case_data.get('status', ''))
    
    result_map = {
        ('yerinde_mudahale', 'completed'): 'E34',
        ('hastaneye_nakil', 'transferred'): 'E35',
        ('hastaneler_arasi', 'inter_hospital'): 'E36',
        ('tibbi_tetkik', 'medical_exam'): 'E37',
        ('eve_nakil', 'home_transfer'): 'E38',
        ('ex_terinde', 'deceased_scene'): 'H34',
        ('ex_morga', 'deceased_morgue'): 'H35',
        ('nakil_reddi', 'refused'): 'H36',
        ('diger_ulasilan',): 'H37',
        ('gorev_iptali', 'cancelled'): 'H38',
        ('baska_aracla',): 'J34',
        ('telefon_baska',): 'J35',
        ('asilsiz',): 'J36',
        ('yaralanan_yok',): 'J37',
        ('olay_yerinde_bekleme',): 'J38',
    }
    
    for keys, cell in result_map.items():
        ws[cell] = "☑" if result.lower() in keys else "☐"
    
    # =====================
    # NAKLEDİLEN HASTANE
    # =====================
    ws['N32'] = case_data.get('transfer_hospital', '')
    
    # İlçe içi/dışı/il dışı
    transfer_type = case_data.get('transfer_type', '')
    ws['L36'] = "☑" if transfer_type == 'ilce_ici' else "☐"
    ws['L37'] = "☑" if transfer_type == 'ilce_disi' else "☐"
    ws['L38'] = "☑" if transfer_type == 'il_disi' else "☐"
    
    # =====================
    # ADLİ VAKA
    # =====================
    is_forensic = case_data.get('is_forensic', False)
    ws['R38'] = "☑" if is_forensic else "☐"  # Evet
    ws['U38'] = "☑" if not is_forensic else "☐"  # Hayır
    
    # =====================
    # CPR BİLGİLERİ
    # =====================
    if cpr_data:
        # Başlama zamanı (V34)
        ws['V34'] = format_time(cpr_data.get('start_time', ''))
        
        # Bırakma zamanı (V35)
        ws['V35'] = format_time(cpr_data.get('end_time', ''))
        
        # Bırakma nedeni (V36)
        ws['V36'] = cpr_data.get('reason', '')
    
    # =====================
    # İŞLEMLER (Checkboxlar)
    # =====================
    procedure_cells = {
        'muayene_acil': ('E40', 'H40'),
        'ambulans_ucreti': ('E41', 'H41'),
        'enjeksiyon_im': ('E43', 'H43'),
        'enjeksiyon_iv': ('E44', 'H44'),
        'enjeksiyon_sc': ('E45', 'H45'),
        'iv_ilac': ('E46', 'H46'),
        'damar_yolu': ('E47', 'H47'),
        'sutur_kucuk': ('E48', 'H48'),
        'mesane_sondasi': ('E49', 'H49'),
        'mide_yikama': ('E50', 'H50'),
        'pansuman_kucuk': ('E51', 'H51'),
        'apse': ('E52', 'H52'),
        'yabanci_cisim': ('E53', 'H53'),
        'cpr': ('E66', 'H66'),
        'ekg': ('E67', 'H67'),
        'defibrilasyon': ('E68', 'H68'),
        'kardiyoversiyon': ('E69', 'H69'),
        'monitorizasyon': ('E70', 'H70'),
        'kanama_kontrolu': ('E71', 'H71'),
        'balon_valf': ('J41', 'H41'),
        'aspirasyon': ('J42', 'H42'),
        'orofaringeal': ('J43', 'H43'),
        'entübasyon': ('J44', 'H44'),
        'mekanik_ventilasyon': ('J45', 'H45'),
        'oksijen': ('J46', 'H46'),
    }
    
    for proc in procedures:
        proc_code = proc.get('code', '').lower()
        proc_name = proc.get('name', '').lower()
        
        for key, (checkbox_cell, count_cell) in procedure_cells.items():
            if key in proc_code or key in proc_name:
                ws[checkbox_cell] = "☑"
                ws[count_cell] = proc.get('count', 1)
                break
    
    # =====================
    # KULLANILAN İLAÇLAR
    # =====================
    # İlaçlar için satır başlangıç noktası
    med_start_row = 41
    med_col_name = 'P'
    med_col_qty = 'T'
    
    for i, med in enumerate(medications[:15]):  # Max 15 ilaç
        row = med_start_row + i
        ws[f'{med_col_name}{row}'] = med.get('name', '')
        ws[f'{med_col_qty}{row}'] = med.get('quantity', 1)
    
    # =====================
    # KULLANILAN MALZEMELER
    # =====================
    mat_start_row = 41
    mat_col_name = 'W'
    mat_col_qty = 'AB'
    
    for i, mat in enumerate(materials[:15]):  # Max 15 malzeme
        row = mat_start_row + i
        ws[f'{mat_col_name}{row}'] = mat.get('name', '')
        ws[f'{mat_col_qty}{row}'] = mat.get('quantity', 1)
    
    # =====================
    # İMZALAR
    # =====================
    # Hastayi Teslim Alanin Adi Soyadi (E80)
    ws['E80'] = signatures.get('receiver_name', '')
    
    # Ünvanı (E81)
    ws['E81'] = signatures.get('receiver_title', '')
    
    # İmza alanları doldurulabilir değil - manuel atılacak
    
    # Ambulans Personeli
    ws['O80'] = signatures.get('doctor_name', '')  # Hekim/PRM
    ws['O81'] = signatures.get('paramedic_name', '')  # Sağlık Per./ATT
    ws['O82'] = signatures.get('driver_name', '')  # Sür./Tekn.
    
    # =====================
    # Çıktıyı BytesIO olarak döndür
    # =====================
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def get_template_info():
    """Template bilgilerini döndür"""
    return {
        "path": TEMPLATE_PATH,
        "exists": os.path.exists(TEMPLATE_PATH)
    }

