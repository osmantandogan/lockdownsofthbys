"""
Dinamik Excel Export Service
Excel şablonundaki data_mappings kullanarak vaka verilerini doldurur
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
from io import BytesIO
import logging
import re

logger = logging.getLogger(__name__)


def format_date(date_val):
    """Tarih formatla"""
    if not date_val:
        return ""
    try:
        if isinstance(date_val, datetime):
            return date_val.strftime("%d.%m.%Y")
        if isinstance(date_val, str):
            dt = datetime.fromisoformat(date_val.replace('Z', '+00:00'))
            return dt.strftime("%d.%m.%Y")
    except:
        pass
    return str(date_val)[:10] if date_val else ""


def format_time(time_val):
    """Saat formatla"""
    if not time_val:
        return ""
    try:
        if isinstance(time_val, datetime):
            return time_val.strftime("%H:%M")
        if isinstance(time_val, str):
            if len(time_val) == 5 and ':' in time_val:
                return time_val
            dt = datetime.fromisoformat(time_val.replace('Z', '+00:00'))
            return dt.strftime("%H:%M")
    except:
        pass
    return str(time_val)[:5] if time_val else ""


def get_nested_value(data: dict, key_path: str):
    """
    Noktalı notasyon ile nested değer al
    Örnek: 'patient.name' -> data['patient']['name']
    """
    keys = key_path.split('.')
    value = data
    
    for key in keys:
        if isinstance(value, dict):
            value = value.get(key)
        else:
            return None
        if value is None:
            return None
    
    return value


def get_case_field_value(case_data: dict, field_key: str) -> str:
    """
    field_key'e göre vaka verisinden değer çek
    """
    
    # Helper: Ön tanı (ICD array'den veya string'den)
    def get_on_tani(d):
        # Direkt string
        if d.get('on_tani'):
            return d.get('on_tani')
        
        # preliminary_diagnosis (ICD array) - Frontend bunu kullanıyor
        prelim = d.get('preliminary_diagnosis', []) or d.get('medical_form', {}).get('preliminary_diagnosis', [])
        if isinstance(prelim, list) and prelim:
            # ICD kodlarını birleştir: "J06.9 - Akut üst solunum yolu enfeksiyonu"
            return ', '.join([f"{p.get('code', '')} - {p.get('name', '')}" for p in prelim if p.get('code')])
        
        # extended_form'dan
        return d.get('extended_form', {}).get('onTani', '') or d.get('extended_form', {}).get('preliminaryDiagnosis', '')
    
    # Helper: Açıklamalar - medical_form.notes en doğru kaynak
    def get_aciklamalar(d):
        mf = d.get('medical_form', {}) or {}
        mf_extended = mf.get('extended_form', {}) or {}
        extended = d.get('extended_form', {}) or {}
        
        return (mf.get('notes', '') or  # EN DOĞRU KAYNAK - medical_form.notes
                d.get('aciklamalar', '') or 
                d.get('notes', '') or 
                mf_extended.get('aciklamalar', '') or
                mf_extended.get('generalNotes', '') or
                extended.get('aciklamalar', '') or
                extended.get('generalNotes', '') or
                '')
    
    # Helper: Nakil Hastanesi (object veya string olabilir)
    def get_transfer_hospital(d):
        medical_form = d.get('medical_form', {}) or {}
        extended_form = d.get('extended_form', {}) or medical_form.get('extended_form', {}) or {}
        
        th = (d.get('transfer_hospital', '') or 
              extended_form.get('transferHospital', '') or 
              extended_form.get('nakledilenHastane', '') or
              medical_form.get('transfer_hospital', ''))
        
        if isinstance(th, dict):
            # Frontend object olarak kaydediyor: {name, type, province}
            name = th.get('name', '')
            province = th.get('province', '')
            return f"{name}{' (' + province + ')' if province else ''}"
        return str(th) if th else ''
    
    # Helper: KM değerleri - medical_form.vehicle_info içinden de kontrol et
    def get_start_km(d):
        vehicle_info = d.get('vehicle_info', {}) or {}
        medical_form = d.get('medical_form', {}) or {}
        mf_vehicle_info = medical_form.get('vehicle_info', {}) or {}
        extended_form = d.get('extended_form', {}) or medical_form.get('extended_form', {}) or {}
        
        return str(
            d.get('start_km', '') or 
            mf_vehicle_info.get('startKm', '') or  # medical_form.vehicle_info.startKm
            mf_vehicle_info.get('start_km', '') or
            vehicle_info.get('start_km', '') or 
            vehicle_info.get('startKm', '') or 
            vehicle_info.get('baslangic_km', '') or
            extended_form.get('startKm', '') or
            ''
        )
    
    def get_end_km(d):
        vehicle_info = d.get('vehicle_info', {}) or {}
        medical_form = d.get('medical_form', {}) or {}
        mf_vehicle_info = medical_form.get('vehicle_info', {}) or {}
        extended_form = d.get('extended_form', {}) or medical_form.get('extended_form', {}) or {}
        
        return str(
            d.get('end_km', '') or 
            mf_vehicle_info.get('endKm', '') or  # medical_form.vehicle_info.endKm
            mf_vehicle_info.get('end_km', '') or
            vehicle_info.get('end_km', '') or 
            vehicle_info.get('endKm', '') or 
            vehicle_info.get('bitis_km', '') or
            extended_form.get('endKm', '') or
            ''
        )
    
    def get_total_km(d):
        start = get_start_km(d)
        end = get_end_km(d)
        try:
            if start and end:
                return str(int(end) - int(start))
        except:
            pass
        return str(d.get('total_km', '') or d.get('vehicle_info', {}).get('total_km', '') or '')
    
    # Helper: Plaka - birden fazla kaynak (assigned_team.vehicle_plate en güvenilir)
    def get_vehicle_plate(d):
        assigned_team = d.get('assigned_team', {}) or {}
        vehicle_info = d.get('vehicle_info', {}) or {}
        medical_form = d.get('medical_form', {}) or {}
        mf_vehicle_info = medical_form.get('vehicle_info', {}) or {}
        extended_form = d.get('extended_form', {}) or medical_form.get('extended_form', {}) or {}
        
        # Debug: assigned_team'den plaka al - DOĞRU ALAN: vehicle_plate
        plate = (assigned_team.get('vehicle_plate', '') or  # EN DOĞRU KAYNAK - vehicle_plate!
                 assigned_team.get('vehicle', '') or  # Eski format
                 d.get('vehicle_plate', '') or 
                 mf_vehicle_info.get('plate', '') or
                 mf_vehicle_info.get('plaka', '') or
                 vehicle_info.get('plate', '') or
                 vehicle_info.get('plaka', '') or
                 extended_form.get('vehiclePlate', ''))
        return str(plate) if plate else ''
    
    # Helper: Vakayı Veren Kurum - boş değilse göster
    def get_referring_inst(d):
        inst = (d.get('referring_institution', '') or 
                d.get('company', '') or
                d.get('extended_form', {}).get('referralSource', '') or
                d.get('extended_form', {}).get('referringInstitution', '') or
                d.get('extended_form', {}).get('vakayiVerenKurum', ''))
        # Eğer sadece _ veya - ise boş döndür
        if inst and str(inst).strip() in ['_', '-', '__', '--']:
            return ''
        return str(inst) if inst else ''
    
    # Temel alanlar (case_data'dan doğrudan erişim)
    basic_mappings = {
        'caseNumber': lambda d: d.get('case_number', ''),
        'caseDate': lambda d: format_date(d.get('created_at')),
        'caseCode': lambda d: d.get('case_code', ''),
        'atn_no': lambda d: d.get('atn_no', ''),
        'vehiclePlate': get_vehicle_plate,
        'plaka': get_vehicle_plate,  # Alternatif isim
        'stationName': lambda d: d.get('station_name', ''),
        'pickupAddress': lambda d: d.get('location', {}).get('address', ''),
        'startKm': get_start_km,
        'endKm': get_end_km,
        'totalKm': get_total_km,
        'baslangicKm': get_start_km,  # Alternatif isim
        'bitisKm': get_end_km,        # Alternatif isim
        'referringInstitution': get_referring_inst,
        'vakayiVerenKurum': get_referring_inst,  # Alternatif isim
        # Ön tanı ve açıklama - helper fonksiyonlar kullan
        'on_tani': get_on_tani,
        'onTani': get_on_tani,  # Alternatif isim
        'aciklamalar': get_aciklamalar,
        'notes': get_aciklamalar,  # Alternatif isim
        # Nakledilen hastane - helper fonksiyon kullan
        'transferHospital': get_transfer_hospital,
        'nakledilenHastane': get_transfer_hospital,  # Alternatif isim
        # Ş.İ. Ambulans Ücreti
        'si_ambulans_ucreti': lambda d: '☑' if d.get('extended_form', {}).get('siAmbulansUcreti') else '☐',
    }
    
    # Helper: time_info'yu al (medical_form içinden veya doğrudan)
    def get_time_info(d):
        ti = d.get('time_info', {}) or {}
        mf = d.get('medical_form', {}) or {}
        mf_ti = mf.get('time_info', {}) or {}  # EN DOĞRU KAYNAK
        return {**ti, **mf_ti}  # mf_ti öncelikli
    
    # Saat alanları (V3 formatı) - medical_form.time_info içinden al
    time_mappings = {
        'callTime': lambda d: format_time(get_time_info(d).get('callTime') or get_time_info(d).get('call_time')),
        'arrivalSceneTime': lambda d: format_time(get_time_info(d).get('arrivalTime') or get_time_info(d).get('arrival_time')),
        'arrivalPatientTime': lambda d: format_time(get_time_info(d).get('patientArrivalTime') or get_time_info(d).get('patient_arrival_time')),
        'departureSceneTime': lambda d: format_time(get_time_info(d).get('departureTime') or get_time_info(d).get('departure_time')),
        'arrivalHospitalTime': lambda d: format_time(get_time_info(d).get('hospitalArrivalTime') or get_time_info(d).get('hospital_arrival_time')),
        'returnStationTime': lambda d: format_time(get_time_info(d).get('stationReturnTime') or get_time_info(d).get('return_time')),
        # Eski isimler için uyumluluk
        'departureTime': lambda d: format_time(get_time_info(d).get('departureTime') or get_time_info(d).get('departure_time')),
        'arrivalTime': lambda d: format_time(get_time_info(d).get('arrivalTime') or get_time_info(d).get('arrival_time')),
        'patientTime': lambda d: format_time(get_time_info(d).get('patientArrivalTime') or get_time_info(d).get('patient_arrival_time')),
        'hospitalTime': lambda d: format_time(get_time_info(d).get('hospitalArrivalTime') or get_time_info(d).get('hospital_arrival_time')),
        'returnTime': lambda d: format_time(get_time_info(d).get('stationReturnTime') or get_time_info(d).get('return_time')),
    }
    
    # Helper: TC Kimlik No - boş değil ve geçerli formatta olmalı
    def get_patient_tc(d):
        patient = d.get('patient', {})
        tc = (patient.get('tc_no') or 
              patient.get('tcNo') or 
              patient.get('tc') or 
              patient.get('tc_kimlik') or
              d.get('caller', {}).get('tc_no') or
              d.get('extended_form', {}).get('patientTcNo') or
              '')
        # TC numarasını string olarak döndür, boş değilse
        tc_str = str(tc).strip() if tc else ''
        # TC numarası 11 haneli olmalı
        if tc_str and len(tc_str) >= 10:
            return tc_str
        return tc_str
    
    # Helper: Yaş hesaplama (doğum tarihinden)
    def get_patient_age(d):
        patient = d.get('patient', {})
        age = patient.get('age')
        if age:
            return str(age)
        
        # Doğum tarihinden hesapla
        birth_date = patient.get('birth_date') or patient.get('birthDate')
        if birth_date:
            try:
                if isinstance(birth_date, str):
                    from datetime import datetime
                    bd = datetime.fromisoformat(birth_date.replace('Z', '+00:00'))
                    today = datetime.now()
                    age = today.year - bd.year - ((today.month, today.day) < (bd.month, bd.day))
                    return str(age)
            except:
                pass
        return ''
    
    # Hasta bilgileri (V3 formatı)
    patient_mappings = {
        'patientName': lambda d: f"{d.get('patient', {}).get('name', '')} {d.get('patient', {}).get('surname', '')}".strip(),
        'patientTcNo': get_patient_tc,
        'patientAge': get_patient_age,
        'patientGender': lambda d: d.get('patient', {}).get('gender', ''),
        'patientPhone': lambda d: str(d.get('patient', {}).get('phone', '') or d.get('caller', {}).get('phone', '') or ''),
        'patientAddress': lambda d: d.get('patient', {}).get('address', '') or d.get('location', {}).get('address', ''),
        'patientHomeAddress': lambda d: d.get('location', {}).get('address', '') or d.get('patient', {}).get('address', ''),
        'patientPickupAddress': lambda d: d.get('location', {}).get('pickup_location', '') or d.get('location', {}).get('address_description', '') or d.get('location', {}).get('address', ''),
        'patientComplaint': lambda d: d.get('patient', {}).get('complaint', '') or d.get('complaint', ''),
        'chronicDiseases': lambda d: d.get('chronic_diseases', '') or d.get('extended_form', {}).get('chronicDiseases', '') or '',
    }
    
    # Helper: CPR verilerini al (tüm kaynaklardan)
    # Frontend: cprBy, cprStart, cprEnd, cprReason -> medical_form.cpr_data
    def get_cpr_data(d, field):
        cpr = d.get('cpr_data', {}) or {}
        extended = d.get('extended_form', {}) or {}
        medical_form = d.get('medical_form', {}) or {}
        mf_cpr = medical_form.get('cpr_data', {}) or {}
        mf_extended = medical_form.get('extended_form', {}) or {}
        
        # Tüm kaynaklardan birleştir (mf_cpr en doğru kaynak)
        all_cpr = {**cpr, **mf_cpr}  # mf_cpr öncelikli
        all_extended = {**extended, **mf_extended}
        
        field_map = {
            # Frontend: cprStart -> start_time
            'start_time': (all_cpr.get('cprStart') or all_cpr.get('start_time') or 
                          all_cpr.get('startTime') or all_cpr.get('baslama_zamani') or 
                          all_extended.get('cprStartTime') or all_extended.get('cprStart') or ''),
            # Frontend: cprEnd -> stop_time
            'stop_time': (all_cpr.get('cprEnd') or all_cpr.get('stop_time') or 
                         all_cpr.get('stopTime') or all_cpr.get('end_time') or 
                         all_cpr.get('bitis_zamani') or all_extended.get('cprStopTime') or 
                         all_extended.get('cprEnd') or ''),
            # Frontend: cprReason -> stop_reason
            'stop_reason': (all_cpr.get('cprReason') or all_cpr.get('stop_reason') or 
                           all_cpr.get('stopReason') or all_cpr.get('reason') or 
                           all_cpr.get('neden') or all_cpr.get('birakma_nedeni') or 
                           all_extended.get('cprStopReason') or all_extended.get('cprReason') or ''),
            # Frontend: cprBy -> performer (değer varsa CPR yapıldı)
            'performed': bool(all_cpr.get('cprBy') or all_cpr.get('cprStart') or 
                         all_cpr.get('performed') or all_cpr.get('yapildi') or 
                         all_extended.get('cprYapildi') or False),
            # Frontend: cprBy -> performer
            'performer': (all_cpr.get('cprBy') or all_cpr.get('performer') or 
                         all_cpr.get('uygulayan') or all_extended.get('cprUygulayan') or 
                         all_extended.get('cprBy') or ''),
        }
        return field_map.get(field, '')
    
    # Sonuç/Nakil bilgileri
    result_mappings = {
        'transferHospital': get_transfer_hospital,  # Yukarıda tanımlandı
        'transferType': lambda d: d.get('transfer_type', '') or d.get('extended_form', {}).get('transferType', ''),
        'caseResult': lambda d: d.get('case_result', '') or d.get('extended_form', {}).get('outcome', '') or d.get('status', ''),
        'isForensic': lambda d: 'Evet' if d.get('is_forensic') or d.get('extended_form', {}).get('isForensic') else 'Hayır',
        'priority': lambda d: d.get('priority', '') or d.get('triage_code', ''),
        # Kazaya karışan araç plakaları - accidentVehicles veya crashVehicles
        'crashVehicle1': lambda d: (d.get('extended_form', {}).get('accidentVehicles', ['']) or d.get('extended_form', {}).get('crashVehicles', ['']))[0] if (d.get('extended_form', {}).get('accidentVehicles') or d.get('extended_form', {}).get('crashVehicles')) else '',
        'crashVehicle2': lambda d: (d.get('extended_form', {}).get('accidentVehicles', ['', '']) or d.get('extended_form', {}).get('crashVehicles', ['', '']))[1] if len(d.get('extended_form', {}).get('accidentVehicles', []) or d.get('extended_form', {}).get('crashVehicles', [])) > 1 else '',
        'crashVehicle3': lambda d: (d.get('extended_form', {}).get('accidentVehicles', ['', '', '']) or d.get('extended_form', {}).get('crashVehicles', ['', '', '']))[2] if len(d.get('extended_form', {}).get('accidentVehicles', []) or d.get('extended_form', {}).get('crashVehicles', [])) > 2 else '',
        'crashVehicle4': lambda d: (d.get('extended_form', {}).get('accidentVehicles', ['', '', '', '']) or d.get('extended_form', {}).get('crashVehicles', ['', '', '', '']))[3] if len(d.get('extended_form', {}).get('accidentVehicles', []) or d.get('extended_form', {}).get('crashVehicles', [])) > 3 else '',
        # CPR bilgileri
        'cprStartTime': lambda d: format_time(get_cpr_data(d, 'start_time')),
        'cprStopTime': lambda d: format_time(get_cpr_data(d, 'stop_time')),
        'cprStopReason': lambda d: get_cpr_data(d, 'stop_reason'),
        'cprPerformer': lambda d: get_cpr_data(d, 'performer'),
    }
    
    # CPR checkbox
    if field_key == 'cpr.yapildi':
        return '☑' if get_cpr_data(case_data, 'performed') else '☐'
    
    # İmza bilgileri
    signature_mappings = {
        # Hekim/PRM
        'hekimAdi': lambda d: d.get('signatures', {}).get('hekim', {}).get('name', '') or d.get('extended_form', {}).get('hekimAdi', ''),
        'hekimImza': lambda d: '✓' if d.get('signatures', {}).get('hekim', {}).get('signed') else '',
        # Sağlık Personeli/ATT
        'saglikPerAdi': lambda d: d.get('signatures', {}).get('saglikPer', {}).get('name', '') or d.get('extended_form', {}).get('saglikPerAdi', ''),
        'saglikPerImza': lambda d: '✓' if d.get('signatures', {}).get('saglikPer', {}).get('signed') else '',
        # Şoför/Tekn.
        'soforAdi': lambda d: d.get('signatures', {}).get('sofor', {}).get('name', '') or d.get('extended_form', {}).get('soforAdi', ''),
        'soforImza': lambda d: '✓' if d.get('signatures', {}).get('sofor', {}).get('signed') else '',
        # Hasta/Hasta Yakını
        'hastaYakiniAdi': lambda d: d.get('signatures', {}).get('hastaYakini', {}).get('name', '') or d.get('extended_form', {}).get('hastaYakiniAdi', ''),
        'hastaYakiniImza': lambda d: '✓' if d.get('signatures', {}).get('hastaYakini', {}).get('signed') else '',
        # Hastayı Teslim Alan
        'teslimAlanAdi': lambda d: d.get('signatures', {}).get('teslimAlan', {}).get('name', '') or d.get('extended_form', {}).get('teslimAlanAdi', ''),
        'teslimAlanUnvan': lambda d: d.get('signatures', {}).get('teslimAlan', {}).get('unvan', '') or d.get('extended_form', {}).get('teslimAlanUnvan', ''),
        'teslimAlanImza': lambda d: '✓' if d.get('signatures', {}).get('teslimAlan', {}).get('signed') else '',
        # Hizmet Reddi
        'hizmetReddiAdi': lambda d: d.get('signatures', {}).get('hizmetReddi', {}).get('name', '') or d.get('extended_form', {}).get('hizmetReddiAdi', ''),
        'hizmetReddiImza': lambda d: '✓' if d.get('signatures', {}).get('hizmetReddi', {}).get('signed') else '',
    }
    
    # Kan şekeri ve ateş - birden fazla kaynak
    def get_kan_sekeri(d):
        clinical = d.get('clinical_observations', {}) or {}
        extended = d.get('extended_form', {}) or {}
        medical_form = d.get('medical_form', {}) or {}
        mf_extended = medical_form.get('extended_form', {}) or {}  # EN DOĞRU KAYNAK
        vital_signs = d.get('vital_signs', []) or []
        
        # Öncelik sırası - medical_form.extended_form en güvenilir
        val = (mf_extended.get('bloodSugar', '') or  # Frontend buraya kaydediyor
               extended.get('bloodSugar', '') or
               clinical.get('blood_sugar', '') or 
               clinical.get('bloodSugar', '') or
               extended.get('kanSekeri', '') or
               extended.get('blood_sugar', ''))
        
        # Vital signs'dan da dene
        if not val and vital_signs and len(vital_signs) > 0:
            val = vital_signs[0].get('blood_sugar', '') or vital_signs[0].get('bloodSugar', '')
        
        return str(val) if val else ''
    
    def get_ates(d):
        clinical = d.get('clinical_observations', {})
        extended = d.get('extended_form', {})
        vital_signs = d.get('vital_signs', [])
        
        # Öncelik sırası
        val = (clinical.get('temperature', '') or 
               clinical.get('temp', '') or
               clinical.get('bodyTemp', '') or
               extended.get('bodyTemp', '') or
               extended.get('ates', '') or
               extended.get('temperature', ''))
        
        # Vital signs'dan da dene
        if not val and vital_signs and len(vital_signs) > 0:
            val = vital_signs[0].get('temp', '') or vital_signs[0].get('temperature', '') or vital_signs[0].get('ates', '')
        
        return str(val) if val else ''
    
    clinical_extra = {
        'kan_sekeri': get_kan_sekeri,
        'ates': get_ates,
        'bloodSugar': get_kan_sekeri,  # Alternatif isim
        'bodyTemp': get_ates,          # Alternatif isim
    }
    
    # Vital bulgular (dinamik) - Format: vitalTime1, vitalBP1, vitalPulse1, etc.
    vital_pattern = re.match(r'vital(Time|BP|Pulse|SpO2|Resp|Temp)(\d+)', field_key)
    if vital_pattern:
        vital_type = vital_pattern.group(1)
        vital_index = int(vital_pattern.group(2)) - 1
        vital_signs = case_data.get('vital_signs', [])
        
        if vital_index < len(vital_signs):
            vs = vital_signs[vital_index]
            # Frontend'de: time, bp, pulse, spo2, respiration, temp
            type_map = {
                'Time': lambda v: format_time(v.get('time', '')),
                'BP': lambda v: str(v.get('bp', '') or v.get('blood_pressure', '')),
                'Pulse': lambda v: str(v.get('pulse', '')),
                'SpO2': lambda v: str(v.get('spo2', '')),
                'Resp': lambda v: str(v.get('respiration', '')),
                'Temp': lambda v: str(v.get('temp', '') or v.get('temperature', '')),
            }
            return str(type_map.get(vital_type, lambda v: '')(vs))
        return ''
    
    # Vital bulgular - Alternatif format: vital1.saat, vital1.nabiz, vital1.tansiyon, vital1.solunum, vital1.spo2
    vital_alt_pattern = re.match(r'vital(\d+)\.(saat|nabiz|tansiyon|solunum|spo2|ates)', field_key)
    if vital_alt_pattern:
        vital_index = int(vital_alt_pattern.group(1)) - 1
        vital_field = vital_alt_pattern.group(2)
        vital_signs = case_data.get('vital_signs', [])
        
        if vital_index < len(vital_signs):
            vs = vital_signs[vital_index]
            # Frontend'de: time, bp, pulse, spo2, respiration, temp
            field_map = {
                'saat': lambda v: format_time(v.get('time', '')),
                'nabiz': lambda v: str(v.get('pulse', '') or v.get('nabiz', '')),
                'tansiyon': lambda v: str(v.get('bp', '') or v.get('blood_pressure', '') or v.get('tansiyon', '')),
                'solunum': lambda v: str(v.get('respiration', '') or v.get('solunum', '')),
                'spo2': lambda v: str(v.get('spo2', '')),
                'ates': lambda v: str(v.get('temp', '') or v.get('temperature', '') or v.get('ates', '')),
            }
            return field_map.get(vital_field, lambda v: '')(vs)
        return ''
    
    # Nabız tipi checkbox (pulse.duzenli, pulse.ritmik, etc.)
    if field_key.startswith('pulse.'):
        pulse_type = case_data.get('clinical_observations', {}).get('pulse_type', '').lower()
        option = field_key.split('.')[1].lower()
        return '☑' if pulse_type == option or option in pulse_type else '☐'
    
    # Solunum tipi checkbox (resp.duzenli, resp.duzensiz, etc.)
    if field_key.startswith('resp.'):
        resp_type = case_data.get('clinical_observations', {}).get('resp_type', '').lower()
        option = field_key.split('.')[1].lower()
        return '☑' if resp_type == option or option in resp_type else '☐'
    
    # GKS
    gks_mappings = {
        'gcsTotal': lambda d: d.get('clinical_observations', {}).get('gks', {}).get('total', ''),
    }
    
    # Klinik gözlemler
    clinical_mappings = {
        'bloodSugar': lambda d: d.get('blood_sugar', '') or d.get('clinical_observations', {}).get('blood_sugar', ''),
        'bodyTemp': lambda d: d.get('body_temperature', '') or d.get('clinical_observations', {}).get('body_temp', ''),
    }
    
    # Checkbox alanları - değer varsa ☑, yoksa ☐
    extended_form = case_data.get('extended_form', {})
    clinical_obs = case_data.get('clinical_observations', {})
    
    # Çağrı tipi - extended_form'dan veya doğrudan
    if field_key.startswith('callType.'):
        call_type = (extended_form.get('callType', '') or case_data.get('call_type', '')).lower()
        option = field_key.split('.')[1].lower()
        return '☑' if option in call_type or call_type == option else '☐'
    
    # Cinsiyet
    if field_key.startswith('gender.'):
        gender = case_data.get('patient', {}).get('gender', '').lower()
        option = field_key.split('.')[1].lower()
        gender_map = {'erkek': 'erkek', 'male': 'erkek', 'kadın': 'kadin', 'kadin': 'kadin', 'female': 'kadin'}
        return '☑' if gender_map.get(gender) == option else '☐'
    
    # Öncelik/Triyaj - extended_form.triageCode veya priority
    if field_key.startswith('priority.'):
        option = field_key.split('.')[1]  # kirmizi_kod, sari_kod, yesil_kod, siyah_kod, sosyal_endikasyon
        
        # triageCode - medical_form.extended_form içinden de al
        medical_form = case_data.get('medical_form', {}) or {}
        mf_extended = medical_form.get('extended_form', {}) or {}
        triage_code = (extended_form.get('triageCode', '') or 
                      mf_extended.get('triageCode', '') or
                      case_data.get('triage_code', '') or 
                      case_data.get('priority', '') or '')
        triage_code = triage_code.lower() if triage_code else ''
        
        # Object format kontrolü
        priority_obj = extended_form.get('priority', {}) or mf_extended.get('priority', {})
        if isinstance(priority_obj, dict):
            return '☑' if priority_obj.get(option) else '☐'
        
        # String format - triageCode ile karşılaştır
        option_clean = option.lower().replace('_kod', '').replace('_', '')  # kirmizi_kod -> kirmizi
        
        # Eşleştirme tablosu: "sarı (ciddi)" -> "sari"
        triage_map = {
            'kirmizi': ['kırmızı', 'kirmizi', 'red', 'acil'],
            'sari': ['sarı', 'sari', 'yellow', 'ciddi', 'sarı (ciddi)'],
            'yesil': ['yeşil', 'yesil', 'green', 'hafif'],
            'siyah': ['siyah', 'black', 'ex'],
            'sosyalendikasyon': ['sosyal', 'social', 'endikasyon', 'sosyal endikasyon']
        }
        
        # Eşleştirme
        for key, values in triage_map.items():
            if option_clean == key or option.replace('_kod', '').replace('_', '') == key:
                for val in values:
                    if val in triage_code:
                        return '☑'
        
        return '☐'
    
    # Çağrı nedeni - extended_form.callReasons object, callReason string, veya callReasonDetail array
    if field_key.startswith('callReason.'):
        option = field_key.split('.')[1]  # kesici_delici, trafik_kaz, medikal, etc.
        
        # 1. Object format: {kesici_delici: true, medikal: true}
        call_reasons = extended_form.get('callReasons', {})
        if isinstance(call_reasons, dict) and call_reasons.get(option):
            return '☑'
        
        # 2. callReasonDetail array: ["medikal", "trafik_kazasi"]
        detail_array = extended_form.get('callReasonDetail', []) or extended_form.get('incidentLocation', [])
        if isinstance(detail_array, list) and option in detail_array:
            return '☑'
        
        # 3. String format: callReason = "medikal" veya "trafik_kazasi"
        reason = (extended_form.get('callReason', '') or case_data.get('call_reason', '')).lower()
        option_lower = option.lower().replace('_', ' ')
        option_clean = option.lower().replace('_', '')
        
        # Trafik kazası özel eşleşmesi
        if option in ['trafik_kaz', 'trafik_kazasi'] and ('trafik' in reason or 'kaza' in reason):
            return '☑'
        
        # Medikal özel eşleşmesi
        if option == 'medikal' and ('medikal' in reason or 'tibbi' in reason or reason == 'acil'):
            return '☑'
        
        # Genel eşleşme
        if option_lower in reason or option_clean in reason.replace(' ', '').replace('_', ''):
            return '☑'
        
        return '☐'
    
    # Pupil - clinical_obs veya extended_form (çoklu kaynak kontrol)
    if field_key.startswith('pupil.'):
        option = field_key.split('.')[1].lower()
        
        # Object format kontrol - tüm kaynaklardan
        pupils_obj = (clinical_obs.get('pupils', {}) or 
                     extended_form.get('pupils', {}) or
                     case_data.get('clinical_observations', {}).get('pupils', {}))
        
        if isinstance(pupils_obj, dict):
            # normal, miyotik, midriatik, anizokorik, reaksiyonyok, fiksdilate
            option_map = {
                'normal': ['normal', 'N'],
                'miyotik': ['miyotik', 'M', 'constricted'],
                'midriatik': ['midriatik', 'D', 'dilated'],
                'anizokorik': ['anizokorik', 'A', 'anisocoric', 'unequal'],
                'reaksiyonyok': ['reaksiyonyok', 'reaksiyon_yok', 'noreaction', 'noReaction'],
                'fiksdilate': ['fiksdilate', 'fiks_dilate', 'fixed', 'fixedDilated'],
            }
            possible_keys = option_map.get(option, [option])
            for key in possible_keys:
                if pupils_obj.get(key) or pupils_obj.get(key.lower()):
                    return '☑'
            return '☐'
        
        # String format
        pupils = str(pupils_obj).lower()
        return '☑' if option in pupils else '☐'
    
    # Deri - clinical_obs veya extended_form (çoklu kaynak kontrol)
    if field_key.startswith('skin.'):
        option = field_key.split('.')[1].lower()
        
        # Object format kontrol - tüm kaynaklardan
        skin_obj = (clinical_obs.get('skin', {}) or 
                   extended_form.get('skin', {}) or
                   case_data.get('clinical_observations', {}).get('skin', {}))
        
        if isinstance(skin_obj, dict):
            # normal, soluk, siyanotik, hiperemik, ikterik, terli
            option_map = {
                'normal': ['normal', 'N'],
                'soluk': ['soluk', 'S', 'pale'],
                'siyanotik': ['siyanotik', 'C', 'cyanotic'],
                'hiperemik': ['hiperemik', 'H', 'hyperemic'],
                'ikterik': ['ikterik', 'I', 'icteric', 'jaundice'],
                'terli': ['terli', 'T', 'sweaty', 'diaphoretic'],
            }
            possible_keys = option_map.get(option, [option])
            for key in possible_keys:
                if skin_obj.get(key) or skin_obj.get(key.lower()):
                    return '☑'
            return '☐'
        
        # String format
        skin = str(skin_obj).lower()
        return '☑' if option in skin else '☐'
    
    # Nabız tipi
    if field_key.startswith('pulseType.') or field_key.startswith('pulse.'):
        option = field_key.split('.')[1]
        
        # Object format
        pulse_obj = clinical_obs.get('pulseType', {}) or extended_form.get('pulseType', {})
        if isinstance(pulse_obj, dict):
            return '☑' if pulse_obj.get(option) else '☐'
        
        # String format
        pulse_type = str(pulse_obj).lower()
        return '☑' if option.lower() in pulse_type else '☐'
    
    # Solunum tipi
    if field_key.startswith('respType.') or field_key.startswith('resp.'):
        option = field_key.split('.')[1]
        
        # Object format
        resp_obj = clinical_obs.get('respType', {}) or extended_form.get('respType', {})
        if isinstance(resp_obj, dict):
            return '☑' if resp_obj.get(option) else '☐'
        
        # String format
        resp_type = str(resp_obj).lower()
        return '☑' if option.lower() in resp_type else '☐'
    
    # GKS Motor/Verbal/Eye - Birden fazla format desteği
    if field_key.startswith('gcsMotor.'):
        clinical_obs = case_data.get('clinical_observations', {})
        extended_form = case_data.get('extended_form', {})
        
        # Birden fazla kaynaktan al
        gks = clinical_obs.get('gks', {})
        motor = gks.get('motor', 0) or gks.get('m', 0)
        
        # Alternatif format: motorResponse (frontend'de bu isimle kaydediliyor)
        if not motor:
            motor = clinical_obs.get('motorResponse', 0) or extended_form.get('motorResponse', 0)
        
        try:
            motor = int(motor) if motor else 0
            target = int(field_key.split('.')[1])
            return '☑' if motor == target else '☐'
        except:
            return '☐'
    
    if field_key.startswith('gcsVerbal.'):
        clinical_obs = case_data.get('clinical_observations', {})
        extended_form = case_data.get('extended_form', {})
        
        gks = clinical_obs.get('gks', {})
        verbal = gks.get('verbal', 0) or gks.get('v', 0)
        
        # Alternatif format: verbalResponse
        if not verbal:
            verbal = clinical_obs.get('verbalResponse', 0) or extended_form.get('verbalResponse', 0)
        
        try:
            verbal = int(verbal) if verbal else 0
            target = int(field_key.split('.')[1])
            return '☑' if verbal == target else '☐'
        except:
            return '☐'
    
    if field_key.startswith('gcsEye.'):
        clinical_obs = case_data.get('clinical_observations', {})
        extended_form = case_data.get('extended_form', {})
        
        gks = clinical_obs.get('gks', {})
        eye = gks.get('eye', 0) or gks.get('e', 0)
        
        # Alternatif format: eyeOpening
        if not eye:
            eye = clinical_obs.get('eyeOpening', 0) or extended_form.get('eyeOpening', 0)
        
        try:
            eye = int(eye) if eye else 0
            target = int(field_key.split('.')[1])
            return '☑' if eye == target else '☐'
        except:
            return '☐'
    
    # GKS Toplam
    if field_key == 'gcsTotal':
        clinical_obs = case_data.get('clinical_observations', {})
        extended_form = case_data.get('extended_form', {})
        
        gks = clinical_obs.get('gks', {})
        motor = int(gks.get('motor', 0) or clinical_obs.get('motorResponse', 0) or extended_form.get('motorResponse', 0) or 0)
        verbal = int(gks.get('verbal', 0) or clinical_obs.get('verbalResponse', 0) or extended_form.get('verbalResponse', 0) or 0)
        eye = int(gks.get('eye', 0) or clinical_obs.get('eyeOpening', 0) or extended_form.get('eyeOpening', 0) or 0)
        
        total = motor + verbal + eye
        return str(total) if total > 0 else ''
    
    # Sonuç checkboxları - extended_form.outcome'dan al
    if field_key.startswith('outcome.'):
        extended_form = case_data.get('extended_form', {})
        result = extended_form.get('outcome', '') or case_data.get('case_result', '') or case_data.get('status', '')
        result = result.lower() if result else ''
        option = field_key.split('.')[1].lower()
        return '☑' if option in result or result == option else '☐'
    
    # Transfer tipi - extended_form.transferType'dan al
    if field_key.startswith('transferType.') or field_key.startswith('distance.'):
        extended_form = case_data.get('extended_form', {})
        transfer = extended_form.get('transferType', '') or case_data.get('transfer_type', '')
        transfer = transfer.lower() if transfer else ''
        option = field_key.split('.')[1].lower()
        return '☑' if option in transfer or transfer == option else '☐'
    
    # Olay yeri checkboxları - extended_form.sceneType veya scene object
    if field_key.startswith('scene.'):
        option = field_key.split('.')[1]
        
        # Object format: {ev: true, sokak: true}
        scene_obj = extended_form.get('scene', {}) or extended_form.get('sceneType', {})
        if isinstance(scene_obj, dict):
            return '☑' if scene_obj.get(option) else '☐'
        
        # String format
        scene = (str(scene_obj) or case_data.get('scene_type', '')).lower()
        return '☑' if option.lower() in scene else '☐'
    
    # Adli vaka
    if field_key.startswith('forensic.'):
        is_forensic = case_data.get('is_forensic', False)
        option = field_key.split('.')[1].lower()
        if option == 'evet':
            return '☑' if is_forensic else '☐'
        else:
            return '☑' if not is_forensic else '☐'
    
    # ============ PROCEDURE MAPPING ============
    # Frontend'de procedures dictionary olarak kaydediliyor:
    # {"Muayene (Acil)": {checked: true, adet: 2}}
    # Mapping key: proc.muayene_acil.cb veya proc.muayene_acil.adet
    
    # İşlem adı -> Mapping key eşleştirmesi
    PROCEDURE_NAME_MAPPING = {
        'muayene_acil': ['Muayene (Acil)', 'Muayene Acil'],
        'enjeksiyon_im': ['Enjeksiyon IM'],
        'enjeksiyon_iv': ['Enjeksiyon IV'],
        'enjeksiyon_sc': ['Enjeksiyon SC'],
        'iv_ilac': ['I.V. İlaç uygulaması', 'IV İlaç'],
        'damar_yolu': ['Damar yolu açılması', 'Damar yolu'],
        'sutur': ['Sütür (küçük)', 'Sütür'],
        'mesane_sondasi': ['Mesane sondası takılması', 'Mesane sondası'],
        'mide_yikama': ['Mide yıkanması', 'Mide yıkama'],
        'pansuman_kucuk': ['Pansuman (küçük)', 'Pansuman'],
        'apse': ['Apse açmak', 'Apse'],
        'yabanci_cisim': ['Yabancı cisim çıkartılması', 'Yabancı cisim'],
        'yanik_pansuman_kucuk': ['Yanık pansumanı (küçük)', 'Yanık pansuman küçük'],
        'yanik_pansuman_orta': ['Yanık pansumanı (orta)', 'Yanık pansuman orta'],
        'ng_sonda': ['NG sonda takma', 'NG sonda'],
        'kulak_buson': ['Kulaktan buşon temizliği', 'Kulak buşon'],
        'kol_atel': ['Kol atel (kısa)', 'Kol atel'],
        'bacak_atel': ['Bacak atel (kısa)', 'Bacak atel'],
        'cilt_traksiyon': ['Cilt traksiyonu uygulaması', 'Cilt traksiyon'],
        'servikal_collar': ['Servikal collar uygulaması', 'Servikal collar'],
        'travma_yelegi': ['Travma yeleği', 'Travma yelek'],
        'vakum_sedye': ['Vakum sedye uygulaması', 'Vakum sedye'],
        'sirt_tahtasi': ['Sırt tahtası uygulaması', 'Sırt tahtası'],
        # Dolaşım desteği
        'cpr': ['CPR (Resüsitasyon)', 'CPR'],
        'ekg': ['EKG Uygulaması', 'EKG'],
        'defibrilasyon': ['Defibrilasyon (CPR)', 'Defibrilasyon'],
        'kardiyoversiyon': ['Kardiyoversiyon'],
        'monitorizasyon': ['Monitörizasyon'],
        'kanama_kontrolu': ['Kanama kontrolü', 'Kanama kontrol'],
        'cut_down': ['Cut down'],
        # Hava yolu
        'balon_valf': ['Balon Valf Maske', 'Balon valf'],
        'aspirasyon': ['Aspirasyon uygulaması', 'Aspirasyon'],
        'orofaringeal': ['Orofaringeal tüp uygulaması', 'Orofaringeal tüp'],
        'entubasyon': ['Endotrakeal entübasyon', 'Entübasyon'],
        'mekanik_vent': ['Mekanik ventilasyon (CPAP–BIPAP dahil)', 'Mekanik ventilasyon'],
        'oksijen': ['Oksijen inhalasyon', 'Oksijen'],
        # Diğer işlemler
        'normal_dogum': ['Normal doğum (Suda dahil)', 'Normal doğum'],
        'kan_sekeri': ['Kan şekeri ölçümü', 'Kan şekeri'],
        'lokal_anestezi': ['Lokal anestezi', 'Lokal'],
        'tirnak_avulsiyon': ['Tırnak avülsiyonu', 'Tırnak'],
        'transkutan_pao2': ['Transkutan PaO2-CO2', 'Transkutan'],
        'debritman': ['Debridman', 'Debridman'],
        'sutur_alinmasi': ['Sütür alınması', 'Sütür alma'],
        # Yenidoğan
        'transport_kuvoz': ['Transport küvöz', 'Küvöz'],
        'canlandirma': ['Yenidoğan canlandırma', 'Canlandırma'],
        'im_enjeksiyon': ['IM enjeksiyon', 'IM'],
        'iv_enjeksiyon': ['IV enjeksiyon', 'IV'],
        'iv_mayi': ['IV mayi', 'Mayi'],
        # Sıvı tedavisi
        'nacl_250': ['%0.9 NaCl 250 cc', 'NaCl 250'],
        'nacl_500': ['%0.9 NaCl 500 cc', 'NaCl 500'],
        'nacl_100': ['%0.9 NaCl 100 cc', 'NaCl 100'],
        'dextroz_500': ['%5 Dextroz 500 cc', 'Dextroz 500'],
        'mannitol_500': ['%20 Mannitol 500 cc', 'Mannitol 500'],
        'isolyte_p': ['İsolyte P 500 cc', 'Isolyte P'],
        'isolyte_s': ['İsolyte S 500 cc', 'Isolyte S'],
        'dengeleyici': ['Dengeleyici solüsyon', 'Dengeleyici'],
        'ringer_laktat': ['Laktatlı Ringer 500 cc', 'Ringer laktat'],
    }
    
    def find_procedure_value(procedures_dict, proc_key, field_type):
        """Dictionary formatındaki procedures içinden değer bul"""
        if not isinstance(procedures_dict, dict):
            return None
        
        # Mapping'den olası isimler al
        possible_names = PROCEDURE_NAME_MAPPING.get(proc_key, [proc_key])
        
        for proc_name, proc_value in procedures_dict.items():
            # İsim eşleşmesi kontrol et
            proc_name_lower = proc_name.lower()
            matched = False
            
            for possible in possible_names:
                if possible.lower() in proc_name_lower or proc_name_lower in possible.lower():
                    matched = True
                    break
            
            # Genel eşleşme de dene
            if not matched and proc_key.replace('_', ' ') in proc_name_lower:
                matched = True
            if not matched and proc_key.replace('_', '') in proc_name_lower.replace(' ', ''):
                matched = True
            
            if matched:
                if isinstance(proc_value, dict):
                    if proc_value.get('checked'):
                        if field_type == 'adet':
                            return str(proc_value.get('adet', 1))
                        return '☑'
                elif proc_value:  # Boolean true
                    if field_type == 'adet':
                        return '1'
                    return '☑'
        
        return '☐' if field_type == 'cb' else ''
    
    # İşlemler - proc., airway., circ., other., newborn., fluid. prefix'leri
    procedure_prefixes = ['proc.', 'airway.', 'circ.', 'other.', 'newborn.', 'fluid.']
    for prefix in procedure_prefixes:
        if field_key.startswith(prefix):
            procedures = case_data.get('procedures', {})
            parts = field_key.split('.')
            proc_key = parts[1].lower() if len(parts) > 1 else ''
            field_type = parts[2] if len(parts) > 2 else 'cb'
            
            result = find_procedure_value(procedures, proc_key, field_type)
            if result:
                return result
            return '☐' if field_type == 'cb' else ''
    
    # ============ MEDICATION MAPPING ============
    MEDICATION_NAME_MAPPING = {
        'arveles': ['Arveles amp.', 'Arveles'],
        'dikloron': ['Dikloron amp.', 'Dikloron'],
        'spazmolitik': ['Spazmolitik amp.', 'Spazmolitik'],
        'adrenalin_05': ['Adrenalin 0,5 mg amp.', 'Adrenalin 0.5'],
        'adrenalin_1': ['Adrenalin 1 mg amp.', 'Adrenalin 1'],
        'atropin': ['Atropin 0,5 mg amp.', 'Atropin'],
        'flumazenil': ['Flumazenil amp.', 'Flumazenil'],
        'dopamin': ['Dopamin amp.', 'Dopamin'],
        'citanest': ['Citanest flk.', 'Citanest', 'Priloc'],
        'nahco3': ['NaHCO3 amp.', 'NaHCO3', 'Bikarbonat'],
        'dizem': ['Dizem amp.', 'Dizem'],
        'aminocordial': ['Aminocordial amp.', 'Aminocordial'],
        'furosemid': ['Furosemid amp.', 'Furosemid', 'Lasix'],
        'ca_glukonat': ['Ca Glukonat amp.', 'Kalsiyum glukonat'],
        'diltizem': ['Diltizem amp.', 'Diltizem'],
        'avil': ['Avil amp.', 'Avil'],
        'dekort': ['Dekort amp.', 'Dekort'],
        'antiepileptik': ['Antiepileptik amp.', 'Antiepileptik'],
        'prednol': ['Prednol amp.', 'Prednol'],
        'aktif_komur': ['Aktif kömür', 'Aktif komur'],
        'beloc': ['Beloc amp.', 'Beloc'],
        'salbutamol': ['Salbutamol amp.', 'Salbutamol', 'Ventolin'],
        'aritmal': ['Aritmal amp.', 'Aritmal'],
        'isoptin': ['Isoptin amp.', 'Isoptin'],
        'kapril': ['Kapril amp.', 'Kapril'],
        'magnezyum': ['Magnezyum amp.', 'Magnezyum'],
        'isorid': ['Isorid amp.', 'Isorid'],
        'coraspin': ['Coraspin tab.', 'Coraspin', 'Aspirin'],
        'paracetamol': ['Paracetamol', 'Perfalgan'],
        'midazolam': ['Midazolam amp.', 'Midazolam', 'Dormicum'],
        'dramamine': ['Dramamine amp.', 'Dramamine'],
        'rotapamid': ['Rotapamid amp.', 'Rotapamid'],
    }
    
    if field_key.startswith('med.'):
        parts = field_key.split('.')
        med_key = parts[1].lower() if len(parts) > 1 else ''
        field_type = parts[2] if len(parts) > 2 else 'cb'
        
        # PDF medications öncelikli - direkt kod ile eşleşir
        # Format: {"arveles": {"checked": true, "adet": 2}}
        pdf_medications = case_data.get('pdf_medications', {})
        if isinstance(pdf_medications, dict) and med_key in pdf_medications:
            med_data = pdf_medications[med_key]
            if isinstance(med_data, dict) and med_data.get('checked'):
                if field_type == 'adet':
                    return str(med_data.get('adet', 1))
                if field_type == 'tur':
                    return med_data.get('tur', '') or med_data.get('route', '') or ''
                return '☑'
        
        # Eski format - medications dictionary veya list
        medications = case_data.get('medications', []) or case_data.get('extended_form', {}).get('medications', {})
        possible_names = MEDICATION_NAME_MAPPING.get(med_key, [med_key])
        
        # medications dict formatında: {"Arveles amp.": {checked: true, adet: 2, tur: "IV"}}
        if isinstance(medications, dict):
            for med_name, med_data in medications.items():
                med_name_lower = med_name.lower()
                matched = False
                
                for possible in possible_names:
                    if possible.lower() in med_name_lower or med_key.replace('_', ' ') in med_name_lower:
                        matched = True
                        break
                
                if matched and isinstance(med_data, dict) and med_data.get('checked'):
                    if field_type == 'adet':
                        return str(med_data.get('adet', 1))
                    if field_type == 'tur':
                        return med_data.get('tur', '') or med_data.get('route', '')
                    return '☑'
        
        # medications liste formatında
        elif isinstance(medications, list):
            for med in medications:
                med_name = (med.get('name', '') or '').lower()
                med_code = (med.get('code', '') or '').lower()
                
                for possible in possible_names:
                    if possible.lower() in med_name or possible.lower() in med_code or med_key in med_name:
                        if field_type == 'adet':
                            return str(med.get('quantity', med.get('count', med.get('adet', 1))))
                        if field_type == 'tur':
                            return med.get('route', med.get('type', med.get('tur', '')))
                        return '☑'
        
        return '☐' if field_type == 'cb' else ''
    
    # ============ MATERIAL MAPPING ============
    MATERIAL_NAME_MAPPING = {
        'enjektor_1_2': ['Enjektör 1-2 cc', 'Enjektör 1-2'],
        'enjektor_5': ['Enjektör 5 cc', 'Enjektör 5'],
        'enjektor_10_20': ['Enjektör 10-20 cc', 'Enjektör 10-20'],
        'monitor_pedi': ['Monitör pedi', 'Monitor pedi'],
        'iv_kateter_14_22': ['I.V. katater 14-22', 'IV kateter 14-22'],
        'iv_kateter_24': ['I.V. katater 24', 'IV kateter 24'],
        'serum_seti': ['Serum seti'],
        'steril_eldiven': ['Steril eldiven'],
        'cerrahi_eldiven': ['Cerrahi eldiven'],
        'sponc': ['Sponç'],
        'sargi_bezi': ['Sargı bezi'],
        'idrar_torbasi': ['İdrar torbası', 'Idrar torbası'],
        'bisturi_ucu': ['Bistüri ucu', 'Bisturi'],
        'entubasyon_balonlu': ['Entübasyon tüpü (balonlu)', 'Entübasyon balonlu'],
        'entubasyon_balonsuz': ['Entübasyon tüpü (balonsuz)', 'Entübasyon balonsuz'],
        'airway': ['Airway'],
        'foley_sonda': ['Foley sonda'],
        'ng_sonda': ['NG sonda'],
        'atravmatik_ipek': ['Atravmatik ipek sütür', 'Atravmatik ipek'],
        'atravmatik_katkut': ['Atravmatik katkut sütür', 'Atravmatik katkut'],
        'dogum_seti': ['Doğum seti'],
        'yanik_battaniyesi': ['Yanık battaniyesi'],
        'o2_maskesi_hazneli_eriskin': ['O2 maskesi hazneli (erişkin)', 'O2 maskesi hazneli erişkin'],
        'o2_maskesi_hazneli_pediatrik': ['O2 maskesi hazneli (pediatrik)', 'O2 maskesi hazneli pediatrik'],
        'o2_kanulu_eriskin': ['O2 kanülü (erişkin)', 'O2 kanülü erişkin'],
        'o2_kanulu_pediatrik': ['O2 kanülü (pediatrik)', 'O2 kanülü pediatrik'],
        'flaster': ['Flaster'],
        'servikal_collar': ['Servikal collar'],
        'elastik_bandaj': ['Elastik bandaj'],
        'etil_chloride': ['Etil chloride', 'Etil klorid'],
        'o2_maskesi_haznesiz_eriskin': ['O2 maskesi haznesiz (erişkin)', 'O2 maskesi haznesiz erişkin'],
        'o2_maskesi_haznesiz_pediatrik': ['O2 maskesi haznesiz (pediatrik)', 'O2 maskesi haznesiz pediatrik'],
    }
    
    if field_key.startswith('mat.'):
        materials = case_data.get('materials', {})
        parts = field_key.split('.')
        mat_key = parts[1].lower() if len(parts) > 1 else ''
        field_type = parts[2] if len(parts) > 2 else 'cb'
        
        possible_names = MATERIAL_NAME_MAPPING.get(mat_key, [mat_key])
        
        # materials dictionary formatında
        if isinstance(materials, dict):
            for mat_name, mat_value in materials.items():
                mat_name_lower = mat_name.lower()
                matched = False
                
                for possible in possible_names:
                    if possible.lower() in mat_name_lower or mat_name_lower in possible.lower():
                        matched = True
                        break
                
                if not matched and mat_key.replace('_', ' ') in mat_name_lower:
                    matched = True
                
                if matched:
                    if isinstance(mat_value, dict):
                        if mat_value.get('checked'):
                            if field_type == 'adet':
                                return str(mat_value.get('adet', 1))
                            return '☑'
                    elif mat_value:
                        if field_type == 'adet':
                            return '1'
                        return '☑'
        
        return '☐' if field_type == 'cb' else ''
    
    # İmzalar (Frontend key'leriyle uyumlu) - inline_consents, signatures ve assigned_team'den al
    def get_inline_consent(d, key):
        return d.get('inline_consents', {}).get(key, '')
    
    def get_team_member_name(d, role):
        """assigned_team'den personel adını al"""
        assigned_team = d.get('assigned_team', {})
        team_members = d.get('team_members', {})  # Bazı vakalar bunda tutuyor
        
        # Role mapping
        role_map = {
            'doctor': ['doctor', 'hekim', 'prm'],
            'paramedic': ['paramedic', 'saglik_per', 'att', 'hemsire'],
            'driver': ['driver', 'sofor', 'pilot']
        }
        
        # assigned_team'de isim varsa direkt al
        for r in role_map.get(role, [role]):
            if assigned_team.get(f'{r}_name'):
                return assigned_team.get(f'{r}_name')
            if team_members.get(f'{r}_name'):
                return team_members.get(f'{r}_name')
        
        return ''
    
    sig_mappings = {
        # Eski format
        'sig.hekim_prm_name': lambda d: d.get('signatures', {}).get('doctor_name', '') or get_inline_consent(d, 'doctor_paramedic_name') or get_team_member_name(d, 'doctor'),
        'sig.saglik_per_name': lambda d: d.get('signatures', {}).get('paramedic_name', '') or get_inline_consent(d, 'health_personnel_name') or get_team_member_name(d, 'paramedic'),
        'sig.sofor_name': lambda d: d.get('signatures', {}).get('driver_name', '') or get_inline_consent(d, 'driver_pilot_name') or get_team_member_name(d, 'driver'),
        'sig.teslim_adi': lambda d: d.get('signatures', {}).get('receiver_name', '') or get_inline_consent(d, 'receiver_title_name'),
        'sig.teslim_unvan': lambda d: d.get('signatures', {}).get('receiver_title', ''),
        'sig.hasta_adi': lambda d: d.get('signatures', {}).get('patient_name', '') or get_inline_consent(d, 'patient_info_consent_name'),
        
        # Yeni format (Frontend'deki key'ler) - inline_consents'dan al
        'sig.teslim_alan_adi': lambda d: get_inline_consent(d, 'receiver_title_name') or d.get('signatures', {}).get('receiver_name', ''),
        'sig.teslim_alan_unvan': lambda d: d.get('signatures', {}).get('receiver_title', ''),
        'sig.teslim_alan_imza': lambda d: '✓' if get_inline_consent(d, 'receiver_signature') else '',
        
        'sig.hekim_prm_adi': lambda d: get_inline_consent(d, 'doctor_paramedic_name') or d.get('signatures', {}).get('doctor_name', '') or get_team_member_name(d, 'doctor'),
        'sig.hekim_prm_imza': lambda d: '✓' if get_inline_consent(d, 'doctor_paramedic_signature') else '',
        
        'sig.saglik_per_adi': lambda d: get_inline_consent(d, 'health_personnel_name') or d.get('signatures', {}).get('paramedic_name', '') or get_team_member_name(d, 'paramedic'),
        'sig.saglik_per_imza': lambda d: '✓' if get_inline_consent(d, 'health_personnel_signature') else '',
        
        'sig.sofor_teknisyen_adi': lambda d: get_inline_consent(d, 'driver_pilot_name') or d.get('signatures', {}).get('driver_name', '') or get_team_member_name(d, 'driver'),
        'sig.sofor_teknisyen_imza': lambda d: '✓' if get_inline_consent(d, 'driver_pilot_signature') else '',
        
        'sig.hasta_yakin_adi': lambda d: get_inline_consent(d, 'patient_info_consent_name') or d.get('signatures', {}).get('patient_name', ''),
        'sig.hasta_yakin_imza': lambda d: '✓' if get_inline_consent(d, 'patient_info_consent_signature') else '',
        
        # Hasta reddi
        'sig.hasta_reddi_adi': lambda d: get_inline_consent(d, 'patient_rejection_name'),
        'sig.hasta_reddi_imza': lambda d: '✓' if get_inline_consent(d, 'patient_rejection_signature') else '',
        
        # Hastane reddi
        'sig.hastane_reddi_doktor': lambda d: get_inline_consent(d, 'hospital_rejection_doctor_name'),
        'sig.hastane_reddi_imza': lambda d: '✓' if get_inline_consent(d, 'hospital_rejection_doctor_signature') else '',
        'sig.hastane_reddi_neden': lambda d: get_inline_consent(d, 'hospital_rejection_reason'),
        'sig.hastane_reddi_kurum': lambda d: get_inline_consent(d, 'hospital_rejection_institution'),
        
        # Ambulans personeli isimleri (PDF'de ayrı hücrelerde)
        'ambulans_hekim': lambda d: get_inline_consent(d, 'doctor_paramedic_name') or get_team_member_name(d, 'doctor'),
        'ambulans_saglik_per': lambda d: get_inline_consent(d, 'health_personnel_name') or get_team_member_name(d, 'paramedic'),
        'ambulans_sofor': lambda d: get_inline_consent(d, 'driver_pilot_name') or get_team_member_name(d, 'driver'),
    }
    
    # Red formları
    reject_mappings = {
        'reject.hastane_neden': lambda d: d.get('hospital_rejection', {}).get('reason', ''),
        'reject.hastane_doktor': lambda d: d.get('hospital_rejection', {}).get('doctor_name', ''),
        'reject.hasta_aciklama': lambda d: d.get('patient_rejection', {}).get('reason', ''),
        'reject.hasta_adi': lambda d: d.get('patient_rejection', {}).get('patient_name', ''),
    }
    
    # Refakatçi
    escort_mappings = {
        'escort.adi': lambda d: d.get('escort', {}).get('name', ''),
    }
    
    # Genel notlar
    if field_key == 'generalNotes':
        return case_data.get('notes', '') or case_data.get('medical_form', {}).get('notes', '')
    
    # Tüm mapping'leri kontrol et
    all_mappings = {
        **basic_mappings,
        **time_mappings,
        **patient_mappings,
        **result_mappings,
        **gks_mappings,
        **clinical_mappings,
        **clinical_extra,
        **signature_mappings,
        **sig_mappings,
        **reject_mappings,
        **escort_mappings,
    }
    
    if field_key in all_mappings:
        result = all_mappings[field_key](case_data)
        return str(result) if result is not None else ''
    
    # Nested path dene
    nested_value = get_nested_value(case_data, field_key)
    if nested_value is not None:
        return str(nested_value)
    
    return ''


def parse_cell_address(address: str):
    """
    A1 -> (1, 1), B2 -> (2, 2), AA10 -> (10, 27)
    """
    match = re.match(r'^([A-Z]+)(\d+)$', address.upper())
    if not match:
        return None, None
    
    col_str = match.group(1)
    row = int(match.group(2))
    
    col = 0
    for char in col_str:
        col = col * 26 + (ord(char) - ord('A') + 1)
    
    return row, col


def export_case_with_template(template: dict, case_data: dict) -> BytesIO:
    """
    Excel şablonu ve data_mappings kullanarak vaka formunu doldur
    
    Args:
        template: Excel şablonu (cells, merged_cells, data_mappings, vb.)
        case_data: Vaka verileri
    
    Returns:
        BytesIO: Doldurulmuş Excel dosyası
    """
    
    wb = Workbook()
    ws = wb.active
    ws.title = template.get('name', 'Vaka Formu')[:31]
    
    max_row = template.get('max_row', 100)
    max_col = template.get('max_column', 30)
    
    # Satır yükseklikleri
    row_heights = template.get('row_heights', {})
    for row_str, height in row_heights.items():
        try:
            ws.row_dimensions[int(row_str)].height = height
        except:
            pass
    
    # Sütun genişlikleri
    column_widths = template.get('column_widths', {})
    for col_letter, width in column_widths.items():
        try:
            ws.column_dimensions[col_letter].width = width
        except:
            pass
    
    # Data mappings
    data_mappings = template.get('data_mappings', {})
    
    # Önce tüm hücreleri şablondan doldur
    cells_data = template.get('cells', [])
    for cell_info in cells_data:
        row = cell_info.get('row', 1)
        col = cell_info.get('col', 1)
        cell = ws.cell(row=row, column=col)
        
        address = cell_info.get('address', f"{get_column_letter(col)}{row}")
        
        # Eğer bu hücre bir mapping içeriyorsa, vaka verisini kullan
        if address in data_mappings:
            field_key = data_mappings[address]
            value = get_case_field_value(case_data, field_key)
            cell.value = value
        else:
            # Mapping yoksa şablondaki değeri kullan
            cell.value = cell_info.get('value', '')
        
        # Stilleri uygula
        font_data = cell_info.get('font', {})
        if font_data:
            try:
                font_color = None
                if font_data.get('color') and len(str(font_data['color'])) >= 6:
                    font_color = str(font_data['color'])[-6:]
                
                cell.font = Font(
                    name=font_data.get('name') or 'Calibri',
                    size=font_data.get('size') or 11,
                    bold=font_data.get('bold', False),
                    italic=font_data.get('italic', False),
                    color=font_color
                )
            except Exception as e:
                logger.warning(f"Font hatası: {e}")
        
        fill_data = cell_info.get('fill', {})
        if fill_data and fill_data.get('color') and fill_data['color'] != '00000000':
            try:
                fill_color = str(fill_data['color'])[-6:]
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            except:
                pass
        
        align_data = cell_info.get('alignment', {})
        if align_data:
            try:
                cell.alignment = Alignment(
                    horizontal=align_data.get('horizontal') or 'left',
                    vertical=align_data.get('vertical') or 'center',
                    wrap_text=align_data.get('wrap_text', False)
                )
            except:
                pass
        
        border_data = cell_info.get('border', {})
        if border_data and any([border_data.get(s) for s in ['left', 'right', 'top', 'bottom']]):
            try:
                cell.border = Border(
                    left=Side(style=border_data.get('left')) if border_data.get('left') else None,
                    right=Side(style=border_data.get('right')) if border_data.get('right') else None,
                    top=Side(style=border_data.get('top')) if border_data.get('top') else None,
                    bottom=Side(style=border_data.get('bottom')) if border_data.get('bottom') else None
                )
            except:
                pass
    
    # Mapping'de olup hücrelerde olmayan değerleri de yaz
    for address, field_key in data_mappings.items():
        row, col = parse_cell_address(address)
        if row and col:
            cell = ws.cell(row=row, column=col)
            if not cell.value:  # Eğer hala boşsa
                value = get_case_field_value(case_data, field_key)
                cell.value = value
    
    # Birleşik hücreler
    merged_cells = template.get('merged_cells', [])
    for merge_info in merged_cells:
        try:
            merge_range = merge_info.get('range')
            if merge_range:
                ws.merge_cells(merge_range)
        except Exception as e:
            logger.warning(f"Merge hatası: {merge_range} - {e}")
    
    # BytesIO'ya kaydet
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

