"""
Excel Template API Routes
Excel formatında form şablonları yönetimi
"""
from fastapi import APIRouter, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from datetime import datetime
from typing import Optional, List, Dict, Any
import uuid
import logging
import json
import os
from io import BytesIO

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side

from database import db
from auth_utils import get_current_user

router = APIRouter(tags=["excel-templates"])
logger = logging.getLogger(__name__)

excel_templates_collection = db["excel_templates"]


def parse_excel_file(file_content: bytes) -> dict:
    """Excel dosyasını parse edip yapısını döndürür"""
    from io import BytesIO
    
    wb = openpyxl.load_workbook(BytesIO(file_content))
    ws = wb.active
    
    result = {
        "sheet_name": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merged_cells": [],
        "cells": [],
        "row_heights": {},
        "column_widths": {}
    }
    
    # Birleşik hücreler
    for merged in ws.merged_cells.ranges:
        result["merged_cells"].append({
            "range": str(merged),
            "min_row": merged.min_row,
            "max_row": merged.max_row,
            "min_col": merged.min_col,
            "max_col": merged.max_col
        })
    
    # Satır yükseklikleri
    for row_idx in range(1, ws.max_row + 1):
        if ws.row_dimensions[row_idx].height:
            result["row_heights"][str(row_idx)] = ws.row_dimensions[row_idx].height
    
    # Sütun genişlikleri
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if ws.column_dimensions[col_letter].width:
            result["column_widths"][col_letter] = ws.column_dimensions[col_letter].width
    
    # Hücreler
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            
            # Sadece değer veya stil içeren hücreleri kaydet
            has_value = cell.value is not None
            has_border = any([
                cell.border.left.style if cell.border and cell.border.left else None,
                cell.border.right.style if cell.border and cell.border.right else None,
                cell.border.top.style if cell.border and cell.border.top else None,
                cell.border.bottom.style if cell.border and cell.border.bottom else None
            ])
            has_fill = cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != "00000000"
            
            if has_value or has_border or has_fill:
                # Font rengi güvenli şekilde al
                font_color = None
                try:
                    if cell.font and cell.font.color:
                        if hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                            font_color = str(cell.font.color.rgb)
                        elif hasattr(cell.font.color, 'theme') and cell.font.color.theme:
                            # Theme color - skip for now
                            font_color = None
                except (AttributeError, TypeError) as e:
                    font_color = None
                
                # Fill rengi güvenli şekilde al
                fill_color = None
                try:
                    if cell.fill and cell.fill.fgColor:
                        if hasattr(cell.fill.fgColor, 'rgb') and cell.fill.fgColor.rgb:
                            fill_color = str(cell.fill.fgColor.rgb)
                        elif hasattr(cell.fill.fgColor, 'theme') and cell.fill.fgColor.theme:
                            # Theme color - skip for now
                            fill_color = None
                except (AttributeError, TypeError) as e:
                    fill_color = None
                
                cell_data = {
                    "row": row,
                    "col": col,
                    "col_letter": get_column_letter(col),
                    "address": f"{get_column_letter(col)}{row}",
                    "value": str(cell.value) if cell.value else "",
                    "font": {
                        "name": cell.font.name if cell.font else None,
                        "size": cell.font.size if cell.font else None,
                        "bold": cell.font.bold if cell.font else False,
                        "italic": cell.font.italic if cell.font else False,
                        "color": font_color
                    },
                    "fill": {
                        "color": fill_color
                    },
                    "alignment": {
                        "horizontal": cell.alignment.horizontal if cell.alignment else None,
                        "vertical": cell.alignment.vertical if cell.alignment else None,
                        "wrap_text": cell.alignment.wrap_text if cell.alignment else False
                    },
                    "border": {
                        "left": cell.border.left.style if cell.border and cell.border.left else None,
                        "right": cell.border.right.style if cell.border and cell.border.right else None,
                        "top": cell.border.top.style if cell.border and cell.border.top else None,
                        "bottom": cell.border.bottom.style if cell.border and cell.border.bottom else None
                    }
                }
                result["cells"].append(cell_data)
    
    return result


@router.get("/test")
async def test_endpoint(request: Request):
    """Test endpoint - router'ın çalışıp çalışmadığını kontrol eder"""
    return {"message": "Excel templates router çalışıyor!", "status": "ok"}


@router.get("")
async def get_all_excel_templates(request: Request):
    """Tüm Excel şablonlarını listele"""
    await get_current_user(request)
    
    templates = await excel_templates_collection.find({}).to_list(100)
    
    # ObjectId ve datetime dönüşümü
    for t in templates:
        t["id"] = t.pop("_id")
        if "created_at" in t and t["created_at"]:
            t["created_at"] = t["created_at"].isoformat()
        if "updated_at" in t and t["updated_at"]:
            t["updated_at"] = t["updated_at"].isoformat()
    
    return templates


@router.get("/{template_id}")
async def get_excel_template(template_id: str, request: Request):
    """Belirli bir Excel şablonunu getir"""
    await get_current_user(request)
    
    template = await excel_templates_collection.find_one({"_id": template_id})
    if not template:
        raise HTTPException(status_code=404, detail="Şablon bulunamadı")
    
    template["id"] = template.pop("_id")
    if "created_at" in template and template["created_at"]:
        template["created_at"] = template["created_at"].isoformat()
    if "updated_at" in template and template["updated_at"]:
        template["updated_at"] = template["updated_at"].isoformat()
    
    return template


@router.post("")
async def create_excel_template(request: Request):
    """Yeni Excel şablonu oluştur"""
    user = await get_current_user(request)
    data = await request.json()
    
    template_id = str(uuid.uuid4())
    
    template = {
        "_id": template_id,
        "name": data.get("name", "Yeni Excel Şablonu"),
        "description": data.get("description", ""),
        "template_type": "excel",
        "usage_types": data.get("usage_types", ["vaka_formu"]),
        "is_default": data.get("is_default", False),
        "max_row": data.get("max_row", 100),
        "max_column": data.get("max_column", 30),
        "cells": data.get("cells", []),
        "merged_cells": data.get("merged_cells", []),
        "row_heights": data.get("row_heights", {}),
        "column_widths": data.get("column_widths", {}),
        "data_mappings": data.get("data_mappings", {}),  # Hangi hücre hangi veri alanına bağlı
        "created_by": user.id,
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow()
    }
    
    await excel_templates_collection.insert_one(template)
    
    template["id"] = template.pop("_id")
    template["created_at"] = template["created_at"].isoformat()
    template["updated_at"] = template["updated_at"].isoformat()
    
    return template


@router.post("/upload")
async def upload_excel_template(
    request: Request,
    file: UploadFile = File(...)
):
    """Excel dosyası yükle ve şablon olarak kaydet"""
    user = await get_current_user(request)
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Sadece Excel dosyaları (.xlsx, .xls) yüklenebilir")
    
    content = await file.read()
    
    try:
        # Excel'i parse et
        structure = parse_excel_file(content)
        
        template_id = str(uuid.uuid4())
        
        template = {
            "_id": template_id,
            "name": file.filename.replace('.xlsx', '').replace('.xls', ''),
            "description": f"Yüklenen Excel dosyası: {file.filename}",
            "template_type": "excel",
            "usage_types": ["vaka_formu"],
            "is_default": False,
            "max_row": structure["max_row"],
            "max_column": structure["max_column"],
            "cells": structure["cells"],
            "merged_cells": structure["merged_cells"],
            "row_heights": structure["row_heights"],
            "column_widths": structure["column_widths"],
            "data_mappings": {},
            "original_filename": file.filename,
            "created_by": user.id,
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow()
        }
        
        await excel_templates_collection.insert_one(template)
        
        template["id"] = template.pop("_id")
        template["created_at"] = template["created_at"].isoformat()
        template["updated_at"] = template["updated_at"].isoformat()
        
        return {
            "message": "Excel şablonu başarıyla yüklendi",
            "template": template
        }
        
    except Exception as e:
        logger.error(f"Excel parse hatası: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Excel dosyası işlenemedi: {str(e)}")


@router.post("/import-from-file")
async def import_vaka_formu(
    request: Request,
    file: UploadFile = File(...)
):
    """Kullanıcının yüklediği Excel dosyasını içe aktar"""
    try:
        logger.info("import-from-file endpoint çağrıldı")
        user = await get_current_user(request)
        logger.info(f"Kullanıcı doğrulandı: {user.id}")
        
        # Dosya adı kontrolü
        if not file.filename:
            raise HTTPException(status_code=400, detail="Dosya adı bulunamadı")
        
        # Dosya formatı kontrolü
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Sadece Excel dosyaları (.xlsx, .xls) yüklenebilir")
        
        # Yüklenen dosyayı oku
        content = await file.read()
        logger.info(f"Dosya yüklendi: {file.filename}, boyut: {len(content)} bytes")
        
        if len(content) == 0:
            raise HTTPException(status_code=400, detail="Dosya boş")
        
        # Excel'i parse et
        try:
            structure = parse_excel_file(content)
            logger.info(f"Excel parse edildi: {len(structure.get('cells', []))} hücre bulundu")
        except Exception as parse_error:
            logger.error(f"Excel parse hatası: {str(parse_error)}", exc_info=True)
            raise HTTPException(status_code=400, detail=f"Excel dosyası işlenemedi: {str(parse_error)}")
        
        # Dosya adından template adını oluştur
        template_name = file.filename.replace('.xlsx', '').replace('.xls', '').strip()
        if not template_name:
            template_name = "Yüklenen Excel Şablonu"
        
        # Mevcut şablon var mı kontrol et (aynı isimle)
        existing = await excel_templates_collection.find_one({"name": template_name})
        
        if existing:
            # Güncelle
            await excel_templates_collection.update_one(
                {"_id": existing["_id"]},
                {"$set": {
                    "max_row": structure["max_row"],
                    "max_column": structure["max_column"],
                    "cells": structure["cells"],
                    "merged_cells": structure["merged_cells"],
                    "row_heights": structure["row_heights"],
                    "column_widths": structure["column_widths"],
                    "updated_at": datetime.utcnow(),
                    "original_filename": file.filename
                }}
            )
            logger.info(f"Şablon güncellendi: {template_name}")
            return {"message": f"{template_name} şablonu güncellendi", "id": str(existing["_id"])}
        
        # Yeni oluştur
        template_id = str(uuid.uuid4())
        
        template = {
            "_id": template_id,
            "name": template_name,
            "description": f"Yüklenen Excel dosyası: {file.filename}",
            "template_type": "excel",
            "usage_types": ["vaka_formu"] if "vaka" in template_name.lower() or "formu" in template_name.lower() else [],
            "is_default": False,  # Kullanıcı isterse varsayılan yapabilir
            "max_row": structure["max_row"],
            "max_column": structure["max_column"],
            "cells": structure["cells"],
            "merged_cells": structure["merged_cells"],
            "row_heights": structure["row_heights"],
            "column_widths": structure["column_widths"],
            "data_mappings": {},  # Kullanıcı daha sonra düzenleyebilir
            "original_filename": file.filename,
            "created_by": user.id,
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow()
        }
        
        await excel_templates_collection.insert_one(template)
        logger.info(f"Yeni şablon oluşturuldu: {template_name} (ID: {template_id})")
        
        return {"message": f"{template_name} şablonu oluşturuldu", "id": template_id, "template": {"id": template_id, "name": template_name}}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Import hatası: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"İçe aktarma hatası: {str(e)}")


@router.put("/{template_id}")
async def update_excel_template(template_id: str, request: Request):
    """Excel şablonunu güncelle"""
    user = await get_current_user(request)
    data = await request.json()
    
    template = await excel_templates_collection.find_one({"_id": template_id})
    if not template:
        raise HTTPException(status_code=404, detail="Şablon bulunamadı")
    
    update_data = {
        "updated_at": datetime.utcnow()
    }
    
    # Güncellenebilir alanlar
    updatable_fields = [
        "name", "description", "usage_types", "is_default",
        "max_row", "max_column", "cells", "merged_cells",
        "row_heights", "column_widths", "data_mappings"
    ]
    
    for field in updatable_fields:
        if field in data:
            update_data[field] = data[field]
    
    await excel_templates_collection.update_one(
        {"_id": template_id},
        {"$set": update_data}
    )
    
    return {"message": "Şablon güncellendi", "id": template_id}


@router.delete("/{template_id}")
async def delete_excel_template(template_id: str, request: Request):
    """Excel şablonunu sil"""
    await get_current_user(request)
    
    result = await excel_templates_collection.delete_one({"_id": template_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Şablon bulunamadı")
    
    return {"message": "Şablon silindi"}


@router.post("/{template_id}/set-default")
async def set_default_excel_template(template_id: str, request: Request):
    """Şablonu varsayılan olarak ayarla"""
    await get_current_user(request)
    
    # Önce tüm şablonların default'ını kaldır
    await excel_templates_collection.update_many(
        {},
        {"$set": {"is_default": False}}
    )
    
    # Bu şablonu default yap
    result = await excel_templates_collection.update_one(
        {"_id": template_id},
        {"$set": {"is_default": True}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Şablon bulunamadı")
    
    return {"message": "Varsayılan şablon ayarlandı"}

