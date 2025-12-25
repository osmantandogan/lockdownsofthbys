from fastapi import APIRouter, HTTPException, Request, UploadFile, File, Form, Body, Depends
from typing import List, Optional
from database import shifts_collection, vehicles_collection, shift_assignments_collection, users_collection, db
from models import Shift, ShiftStart, ShiftEnd, ShiftAssignment
from auth_utils import get_current_user, require_roles
from datetime import datetime, timedelta
from pydantic import BaseModel, validator
from utils.timezone import get_turkey_time
import base64
import uuid
import logging
import pytz

logger = logging.getLogger(__name__)

router = APIRouter()

# Request model for creating shift assignment
class ShiftAssignmentCreate(BaseModel):
    model_config = {"extra": "forbid"}
    
    user_id: str
    vehicle_id: Optional[str] = None
    location_type: str = "arac"
    health_center_name: Optional[str] = None
    shift_date: str
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    end_date: Optional[str] = None
    healmedy_location_id: Optional[str] = None  # Healmedy lokasyonu
    assigned_role: Optional[str] = None  # Geçici görev rolü (örn: paramedik şoför olarak görevlendirildi)

# ============================================================================
# SHIFT ASSIGNMENT ENDPOINTS - YENİDEN YAZILDI
# ============================================================================

# Helper function to serialize assignment
def serialize_assignment(assignment):
    """Convert MongoDB document to JSON-serializable format"""
    if isinstance(assignment, dict):
        assignment = dict(assignment)
    else:
        assignment = assignment.model_dump(by_alias=True) if hasattr(assignment, 'model_dump') else dict(assignment)
    
    if "_id" in assignment:
        assignment["id"] = str(assignment.pop("_id"))
    elif "id" not in assignment and hasattr(assignment, 'id'):
        assignment["id"] = str(assignment.id)
    
    # Convert datetime fields
    for field in ["shift_date", "end_date", "created_at"]:
        if field in assignment and isinstance(assignment[field], datetime):
            assignment[field] = assignment[field].isoformat()
        elif field in assignment and isinstance(assignment[field], str) and 'T' in assignment[field]:
            pass  # Already ISO string
    
    return assignment

@router.get("/debug/data-check")
async def debug_data_check(request: Request):
    """Debug endpoint to check users and shifts data"""
    try:
        # Get sample users
        users = await users_collection.find().limit(20).to_list(20)
        users_info = []
        for u in users:
            users_info.append({
                "_id": str(u.get("_id")),
                "_id_type": type(u.get("_id")).__name__,
                "name": u.get("name", "YOK"),
                "email": u.get("email", "YOK"),
                "role": u.get("role", "YOK")
            })
        
        # Get sample shifts
        shifts = await shift_assignments_collection.find().limit(20).to_list(20)
        shifts_info = []
        for s in shifts:
            user_id = s.get("user_id")
            user_found = None
            if user_id:
                user = await users_collection.find_one({"_id": user_id})
                user_found = user.get("name") if user else "BULUNAMADI"
            
            shifts_info.append({
                "_id": str(s.get("_id")),
                "user_id": str(user_id) if user_id else "YOK",
                "user_id_type": type(user_id).__name__ if user_id else "YOK",
                "user_id_len": len(str(user_id)) if user_id else 0,
                "user_name_in_shift": s.get("user_name", "YOK"),
                "user_found_by_lookup": user_found,
                "shift_date": str(s.get("shift_date")),
                "vehicle_plate": s.get("vehicle_plate", "YOK")
            })
        
        return {
            "user_count": await users_collection.count_documents({}),
            "shift_count": await shift_assignments_collection.count_documents({}),
            "sample_users": users_info,
            "sample_shifts": shifts_info
        }
    except Exception as e:
        logger.error(f"Debug endpoint error: {e}")
        return {"error": str(e)}

def _norm(name):
    tr = {'ş':'s','Ş':'S','ğ':'g','Ğ':'G','ü':'u','Ü':'U','ı':'i','İ':'I','ö':'o','Ö':'O','ç':'c','Ç':'C'}
    r = name.lower().strip()
    for a, b in tr.items(): r = r.replace(a, b.lower())
    return ' '.join(r.split())

_DEC25 = {
    "merkez": {"plate": None, "loc": "saglik_merkezi", "hc": "Merkez Ofis", "hrs": "08:00-17:00", "p": [
        {"n": "Cansel Petek SAHİN", "r": "doktor"}, {"n": "Irem HODULLAR", "r": "doktor"},
        {"n": "Umutcan OZDAL", "r": "hemsire"}, {"n": "Nese VERIMCIK", "r": "hemsire"},
        {"n": "Merve GIRGIN", "r": "cagri_merkezi"}, {"n": "Yasemin BASTURK", "r": "cagri_merkezi"},
    ]},
    "06_CHZ_142": {"plate": "06 CHZ 142", "loc": "arac", "hrs": "08:00-08:00", "p": [
        {"n": "Hatice Acar CANBAZ", "r": "paramedik"}, {"n": "Aleyna OZDEMIR", "r": "paramedik"}, {"n": "Hasan GUNEY", "r": "paramedik"},
    ]},
    "06_CHZ_146": {"plate": "06 CHZ 146", "loc": "arac", "hrs": "08:00-08:00", "p": [
        {"n": "Elif KURBAN", "r": "paramedik"}, {"n": "Burak ILIK", "r": "paramedik"}, {"n": "Busra Bahtiyar GUZEL", "r": "paramedik"},
    ]},
    "06_CHZ_149": {"plate": "06 CHZ 149", "loc": "arac", "hrs": "08:00-08:00", "p": [
        {"n": "Aysegul Beyza YILMAZ", "r": "paramedik"}, {"n": "Ugur VAR", "r": "paramedik"}, {"n": "Mervenur GEDIK", "r": "paramedik"},
    ]},
    "34_FTU_336": {"plate": "34 FTU 336", "loc": "arac", "hrs": "08:00-08:00", "p": [
        {"n": "Nesrin TUYSUZ", "r": "paramedik"}, {"n": "Gamze Hande BOZ", "r": "paramedik"}, {"n": "Alican TULUBAS", "r": "paramedik"},
    ]},
    "34_KMP_224": {"plate": "34 KMP 224", "loc": "arac", "hrs": "08:00-08:00", "p": [
        {"n": "Murat KESER", "r": "paramedik"}, {"n": "Melike KARATEPE", "r": "paramedik"}, {"n": "Burakcan SAHINTURK", "r": "paramedik"},
    ]},
    "34_MHA_112": {"plate": "34 MHA 112", "loc": "arac", "hrs": "08:00-08:00", "p": [
        {"n": "Kadir ARTAR", "r": "paramedik"}, {"n": "Hamza Tarik ERMIS", "r": "paramedik"}, {"n": "Buse TOPCU", "r": "paramedik"},
    ]},
}

@router.post("/reimport-december-2025")
async def reimport_dec_2025(request: Request):
    res = {"deleted": 0, "created": 0, "not_found": [], "matched": []}
    try:
        del_r = await shift_assignments_collection.delete_many({})
        res["deleted"] = del_r.deleted_count
        users = await users_collection.find().to_list(1000)
        vehicles = await vehicles_collection.find().to_list(100)
        for lk, d in _DEC25.items():
            plate, loc, hc, hrs = d.get("plate"), d.get("loc", "arac"), d.get("hc"), d.get("hrs", "08:00-08:00")
            vid = next((v.get("_id") for v in vehicles if v.get("plate") == plate), None) if plate else None
            st, et = hrs.split("-")
            is24 = st == "08:00" and et == "08:00"
            for p in d["p"]:
                pn, pr = p["n"], p["r"]
                np = _norm(pn)
                user = next((u for u in users if _norm(u.get("name",""))==np or len(set(np.split())&set(_norm(u.get("name","")).split()))>=2), None)
                if not user: res["not_found"].append(pn); continue
                uid, un = user.get("_id"), user.get("name")
                res["matched"].append({"in": pn, "db": un})
                for day in range(14, 22):
                    sd = datetime(2025, 12, day, 8, 0, 0)
                    ed = sd + timedelta(days=1) if is24 else sd
                    await shift_assignments_collection.insert_one({"_id": str(uuid.uuid4()), "user_id": uid, "user_name": un, "user_role": pr, "vehicle_id": vid, "vehicle_plate": plate, "shift_date": sd, "end_date": ed, "start_time": st, "end_time": et, "location_type": loc, "health_center_name": hc, "status": "pending", "created_at": datetime.utcnow()})
                    res["created"] += 1
        logger.info(f"Reimport tamamlandi: {res['created']} vardiya olusturuldu")
        return res
    except Exception as e:
        logger.error(f"Reimport hatasi: {e}")
        res["error"] = str(e)
        return res

@router.delete("/assignments/delete-all")
async def delete_all_assignments(request: Request):
    """Tüm vardiya atamalarını sil - Sadece merkez_ofis ve operasyon_muduru"""
    try:
        await require_roles(["merkez_ofis", "operasyon_muduru"])(request)
    except:
        pass  # Debug için geçici
    
    try:
        result = await shift_assignments_collection.delete_many({})
        logger.info(f"Toplu silme: {result.deleted_count} vardiya silindi")
        return {"deleted": result.deleted_count, "message": f"{result.deleted_count} vardiya başarıyla silindi"}
    except Exception as e:
        logger.error(f"Toplu silme hatası: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/reset-all")
async def reset_all_shifts_and_assignments(request: Request):
    """TÜM vardiya atamalarını VE aktif vardiyaları sıfırla - TEHLİKELİ!"""
    await require_roles(["merkez_ofis", "operasyon_muduru"])(request)
    
    try:
        # 1. Tüm atamaları sil
        assignments_result = await shift_assignments_collection.delete_many({})
        
        # 2. Aktif ve beklemede vardiyaları "iptal" olarak işaretle veya sil
        # Güvenlik için aktif vardiyaları silmek yerine "cancelled" olarak işaretleyelim
        active_shifts = await shifts_collection.update_many(
            {"status": {"$in": ["active", "on_break"]}},
            {"$set": {"status": "cancelled", "end_time": get_turkey_time(), "admin_note": "Toplu sıfırlama ile iptal edildi"}}
        )
        
        logger.info(f"Toplu sıfırlama: {assignments_result.deleted_count} atama silindi, {active_shifts.modified_count} vardiya iptal edildi")
        
        return {
            "assignments_deleted": assignments_result.deleted_count,
            "shifts_cancelled": active_shifts.modified_count,
            "message": f"{assignments_result.deleted_count} atama silindi, {active_shifts.modified_count} aktif vardiya iptal edildi"
        }
    except Exception as e:
        logger.error(f"Toplu sıfırlama hatası: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/cancel-active-shifts")
async def cancel_all_active_shifts(request: Request):
    """Tüm aktif vardiyaları iptal et (atamalar kalır)"""
    await require_roles(["merkez_ofis", "operasyon_muduru"])(request)
    
    try:
        result = await shifts_collection.update_many(
            {"status": {"$in": ["active", "on_break"]}},
            {"$set": {"status": "cancelled", "end_time": get_turkey_time(), "admin_note": "Toplu iptal"}}
        )
        
        logger.info(f"Tüm aktif vardiyalar iptal edildi: {result.modified_count}")
        return {"cancelled": result.modified_count, "message": f"{result.modified_count} aktif vardiya iptal edildi"}
    except Exception as e:
        logger.error(f"Vardiya iptal hatası: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/super-end-shifts")
async def super_end_all_shifts(request: Request):
    """
    Süper vardiya bitirme - Tüm aktif vardiyaları veya tarih aralığındaki vardiyaları bitir.
    
    Query params:
    - all: true ise tüm zamanlardaki aktif vardiyaları bitirir
    - start_date: Başlangıç tarihi (YYYY-MM-DD) - isteğe bağlı
    - end_date: Bitiş tarihi (YYYY-MM-DD) - isteğe bağlı
    """
    admin_user = await require_roles(["merkez_ofis", "operasyon_muduru"])(request)
    
    try:
        body = await request.json()
    except:
        body = {}
    
    end_all = body.get("all", False)
    start_date_str = body.get("start_date")
    end_date_str = body.get("end_date")
    
    turkey_now = get_turkey_time()
    turkey_tz = pytz.timezone('Europe/Istanbul')
    
    query = {"status": {"$in": ["active", "on_break"]}}
    
    if not end_all and (start_date_str or end_date_str):
        # Tarih aralığı filtresi
        date_filter = {}
        
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
                start_date = turkey_tz.localize(start_date)
                date_filter["$gte"] = start_date
            except ValueError:
                pass
        
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
                end_date = turkey_tz.localize(end_date) + timedelta(days=1)  # Gün sonuna kadar dahil
                date_filter["$lt"] = end_date
            except ValueError:
                pass
        
        if date_filter:
            query["start_time"] = date_filter
    
    # Önce kaç vardiya etkilenecek kontrol et
    affected_count = await shifts_collection.count_documents(query)
    
    if affected_count == 0:
        return {
            "ended": 0,
            "assignments_updated": 0,
            "message": "Bitirilecek vardiya bulunamadı"
        }
    
    # Vardiyaları bitir
    result = await shifts_collection.update_many(
        query,
        {
            "$set": {
                "status": "completed",
                "end_time": turkey_now,
                "admin_note": f"Süper vardiya bitirme - {admin_user.name}",
                "ended_by_admin": True,
                "admin_id": admin_user.id
            }
        }
    )
    
    # İlgili assignment'ları da güncelle
    ended_shifts = await shifts_collection.find({
        "end_time": turkey_now,
        "admin_note": {"$regex": "Süper vardiya bitirme"}
    }).to_list(1000)
    
    assignment_ids = [s.get("assignment_id") for s in ended_shifts if s.get("assignment_id")]
    
    assignment_result = await shift_assignments_collection.update_many(
        {"_id": {"$in": assignment_ids}},
        {"$set": {"status": "completed"}}
    )
    
    logger.info(f"Süper vardiya bitirme: {result.modified_count} vardiya, {assignment_result.modified_count} atama tamamlandı")
    
    return {
        "ended": result.modified_count,
        "assignments_updated": assignment_result.modified_count,
        "message": f"{result.modified_count} vardiya başarıyla bitirildi"
    }


@router.delete("/assignments/user/{user_id}")
async def delete_user_assignments(user_id: str, request: Request):
    """Belirli bir kullanıcının tüm atamalarını sil - Çakışma temizliği için"""
    await require_roles(["merkez_ofis", "operasyon_muduru", "bas_sofor"])(request)
    
    try:
        # Kullanıcının tüm atamalarını bul ve sil
        assignments = await shift_assignments_collection.find({"user_id": user_id}).to_list(100)
        result = await shift_assignments_collection.delete_many({"user_id": user_id})
        
        # Kullanıcı adını bul
        from database import users_collection
        user = await users_collection.find_one({"_id": user_id})
        user_name = user.get("name", user_id) if user else user_id
        
        logger.info(f"Kullanıcı atamaları silindi: {user_name} ({user_id}), Silinen: {result.deleted_count}")
        return {
            "deleted": result.deleted_count, 
            "user_name": user_name,
            "message": f"{user_name} için {result.deleted_count} atama silindi"
        }
    except Exception as e:
        logger.error(f"Kullanıcı atamaları silme hatası: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/assignments/today")
async def get_today_assignments(request: Request):
    """Get today's shift assignments - visible to all users"""
    await get_current_user(request)  # Just verify user is logged in
    
    from database import users_collection
    from datetime import timedelta
    
    # Türkiye saati (UTC+3) kullan
    turkey_now = get_turkey_time()
    today = turkey_now.date()
    today_str = today.isoformat()
    logger.info(f"Bugünkü atamalar sorgusu - Bugün (TR): {today}")
    
    # Find all assignments that include today
    # Either: shift_date is today OR (shift_date <= today AND end_date >= today)
    all_assignments = await shift_assignments_collection.find({
        "status": {"$in": ["pending", "started"]}
    }).to_list(1000)
    
    # Filter assignments that are active today
    today_assignments = []
    for assignment in all_assignments:
        shift_date_str = assignment.get("shift_date", "")
        end_date_str = assignment.get("end_date", "")
        
        # Parse shift date
        if isinstance(shift_date_str, datetime):
            shift_date = shift_date_str.date()
        elif isinstance(shift_date_str, str):
            try:
                if 'T' in shift_date_str:
                    shift_date = datetime.fromisoformat(shift_date_str.replace('Z', '+00:00')).date()
                else:
                    shift_date = datetime.strptime(shift_date_str, "%Y-%m-%d").date()
            except:
                continue
        else:
            continue
        
        # Parse end date
        end_date = shift_date  # Default to same day
        if end_date_str:
            if isinstance(end_date_str, datetime):
                end_date = end_date_str.date()
            elif isinstance(end_date_str, str):
                try:
                    if 'T' in end_date_str:
                        end_date = datetime.fromisoformat(end_date_str.replace('Z', '+00:00')).date()
                    else:
                        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
                except:
                    pass
        
        # Check if today falls within the assignment period
        if shift_date <= today <= end_date:
            today_assignments.append(assignment)
    
    # Enrich with user information
    enriched_assignments = []
    for assignment in today_assignments:
        user_doc = await users_collection.find_one({"_id": assignment.get("user_id")})
        serialized = serialize_assignment(assignment)
        if user_doc:
            serialized["user_name"] = user_doc.get("name", "Bilinmiyor")
            serialized["user_role"] = user_doc.get("role", "-")
            serialized["profile_photo"] = user_doc.get("profile_photo")  # Profil fotoğrafı
        # Geçici görev rolü varsa ekle (assigned_role)
        serialized["assigned_role"] = assignment.get("assigned_role")
        enriched_assignments.append(serialized)
    
    # Group by location type
    vehicle_assignments = [a for a in enriched_assignments if a.get("location_type") == "arac"]
    health_center_assignments = [a for a in enriched_assignments if a.get("location_type") == "saglik_merkezi"]
    
    # Get vehicle info for vehicle assignments
    for assignment in vehicle_assignments:
        vehicle_id = assignment.get("vehicle_id")
        if vehicle_id:
            vehicle = await vehicles_collection.find_one({"_id": vehicle_id})
            if vehicle:
                assignment["vehicle_plate"] = vehicle.get("plate", "")
                assignment["vehicle_type"] = vehicle.get("type", "")
    
    return {
        "date": today_str,
        "vehicle_assignments": vehicle_assignments,
        "health_center_assignments": health_center_assignments,
        "total_count": len(enriched_assignments)
    }

@router.get("/assignments/my")
async def get_my_assignments(request: Request):
    """Get user's shift assignments"""
    user = await get_current_user(request)
    
    assignments = await shift_assignments_collection.find({
        "user_id": user.id,
        "status": {"$in": ["pending", "started"]}
    }).sort("shift_date", -1).to_list(100)
    
    result = []
    for a in assignments:
        serialized = serialize_assignment(a)
        # Araç bilgisini ekle
        vehicle_id = a.get("vehicle_id")
        if vehicle_id:
            vehicle = await vehicles_collection.find_one({"_id": vehicle_id})
            if vehicle:
                serialized["vehicle_plate"] = vehicle.get("plate", "")
                serialized["vehicle_type"] = vehicle.get("type", "")
                serialized["vehicle"] = {
                    "plate": vehicle.get("plate"),
                    "type": vehicle.get("type"),
                    "km": vehicle.get("km")
                }
        result.append(serialized)
    
    return result

@router.get("/assignments")
async def get_all_assignments(request: Request, date: Optional[str] = None):
    """
    Get shift assignments (Admin only)
    - date parametresi verilirse sadece o güne ait atamaları döner (çok hızlı)
    - date verilmezse tüm atamaları döner (yavaş - sadece gerektiğinde kullanın)
    """
    current_user = await require_roles(["merkez_ofis", "operasyon_muduru", "bas_sofor"])(request)
    
    from database import users_collection
    
    # Tarihe göre filtre
    if date:
        try:
            target_date = datetime.strptime(date, "%Y-%m-%d").date()
            target_datetime = datetime.combine(target_date, datetime.min.time())
            
            # MongoDB sorgusu ile doğrudan filtrele (çok daha hızlı)
            query = {
                "$or": [
                    # O gün başlayan vardiyalar
                    {"shift_date": target_datetime},
                    # O gün devam eden çok günlük vardiyalar
                    {
                        "shift_date": {"$lte": target_datetime},
                        "end_date": {"$gte": target_datetime}
                    }
                ]
            }
            
            assignments = await shift_assignments_collection.find(query).to_list(1000)
            logger.info(f"Tarihe göre atama sorgusu: {date} - {len(assignments)} atama bulundu")
        except ValueError as e:
            logger.error(f"Geçersiz tarih formatı: {date} - {e}")
            raise HTTPException(status_code=400, detail="Geçersiz tarih formatı. YYYY-MM-DD kullanın.")
    else:
        # Tüm atamaları getir (yavaş)
        assignments = await shift_assignments_collection.find({}).sort("shift_date", -1).to_list(1000)
    
    # Kullanıcı ve araç bilgilerini toplu çek (performans için)
    user_ids = list(set(a.get("user_id") for a in assignments if a.get("user_id")))
    vehicle_ids = list(set(a.get("vehicle_id") for a in assignments if a.get("vehicle_id")))
    
    # User ID'leri için hem ObjectId hem de string olarak arama yap
    from bson import ObjectId
    
    # Farklı formatlardaki ID'leri ayır
    object_user_ids = []
    string_user_ids = []
    
    for uid in user_ids:
        if isinstance(uid, str):
            if len(uid) == 24:
                # ObjectId formatı olabilir
                try:
                    object_user_ids.append(ObjectId(uid))
                except:
                    string_user_ids.append(uid)
            else:
                # UUID veya başka string format
                string_user_ids.append(uid)
        else:
            object_user_ids.append(uid)
    
    # Hem ObjectId hem string _id'lerle kullanıcıları çek
    users_docs = []
    if object_user_ids:
        docs = await users_collection.find({"_id": {"$in": object_user_ids}}).to_list(1000)
        users_docs.extend(docs)
    if string_user_ids:
        docs = await users_collection.find({"_id": {"$in": string_user_ids}}).to_list(1000)
        users_docs.extend(docs)
    
    logger.info(f"Kullanıcı lookup: {len(user_ids)} ID, {len(users_docs)} kullanıcı bulundu")
    
    # Map'te hem ObjectId hem string key'leri tut
    users_map = {}
    for u in users_docs:
        users_map[u["_id"]] = u
        users_map[str(u["_id"])] = u
    
    vehicles_docs = await vehicles_collection.find({"_id": {"$in": vehicle_ids}}).to_list(len(vehicle_ids))
    vehicles_map = {v["_id"]: v for v in vehicles_docs}
    
    result = []
    for a in assignments:
        serialized = serialize_assignment(a)
        
        # Kullanıcı bilgilerini ekle
        user_doc = users_map.get(a.get("user_id"))
        if user_doc:
            serialized["user_name"] = user_doc.get("name", "Bilinmiyor")
            serialized["user_role"] = user_doc.get("role", "-")
            serialized["profile_photo"] = user_doc.get("profile_photo")
        else:
            # Kullanıcı bulunamazsa, shift'teki user_name'i kullan (import'tan gelenler için)
            serialized["user_name"] = a.get("user_name", "Bilinmiyor")
            serialized["user_role"] = a.get("user_role", "-")
        # Geçici görev rolü varsa ekle (assigned_role)
        serialized["assigned_role"] = a.get("assigned_role")
        
        # Araç bilgilerini ekle
        vehicle_doc = vehicles_map.get(a.get("vehicle_id"))
        if vehicle_doc:
            serialized["vehicle_plate"] = vehicle_doc.get("plate", "")
        elif a.get("vehicle_plate"):
            # Araç bulunamazsa, shift'teki vehicle_plate'i kullan
            serialized["vehicle_plate"] = a.get("vehicle_plate", "")
        
        result.append(serialized)
    
    return result

@router.post("/create-assignment-v2")
async def create_shift_assignment(data: ShiftAssignmentCreate, request: Request):
    """Create shift assignment (Admin only)"""
    current_user = await require_roles(["merkez_ofis", "operasyon_muduru", "bas_sofor"])(request)
    
    # Validate user exists
    from database import users_collection
    user = await users_collection.find_one({"_id": data.user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Validate vehicle if location type is vehicle
    if data.location_type == "arac":
        if not data.vehicle_id:
            raise HTTPException(status_code=400, detail="Vehicle ID is required for vehicle assignments")
        vehicle = await vehicles_collection.find_one({"_id": data.vehicle_id})
        if not vehicle:
            raise HTTPException(status_code=404, detail="Vehicle not found")
    elif data.location_type == "saglik_merkezi":
        if not data.health_center_name:
            data.health_center_name = "Sağlık Merkezi"
    
    # Parse shift_date
    try:
        if 'T' in data.shift_date:
            shift_date = datetime.fromisoformat(data.shift_date.replace('Z', '+00:00'))
        else:
            shift_date = datetime.strptime(data.shift_date, '%Y-%m-%d')
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid shift_date: {str(e)}")
    
    # Parse end_date
    end_date = None
    if data.end_date:
        try:
            if 'T' in data.end_date:
                end_date = datetime.fromisoformat(data.end_date.replace('Z', '+00:00'))
            else:
                end_date = datetime.strptime(data.end_date, '%Y-%m-%d')
                end_date = end_date.replace(hour=23, minute=59, second=59)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid end_date: {str(e)}")
    
    # Check for overlaps
    active_assignments = await shift_assignments_collection.find({
        "user_id": data.user_id,
        "status": {"$in": ["pending", "started"]}
    }).to_list(100)
    
    logger.info(f"Overlap kontrolü - Kullanıcı: {user.get('name')} ({data.user_id}), Aktif atama sayısı: {len(active_assignments)}")
    
    new_end = end_date if end_date else shift_date
    
    for existing in active_assignments:
        existing_start = existing.get("shift_date")
        existing_end = existing.get("end_date") or existing_start
        existing_id = existing.get("_id", "Bilinmiyor")
        existing_status = existing.get("status", "Bilinmiyor")
        
        # Convert to datetime if needed
        if isinstance(existing_start, str):
            existing_start = datetime.fromisoformat(existing_start.replace('Z', '+00:00'))
        if isinstance(existing_end, str):
            existing_end = datetime.fromisoformat(existing_end.replace('Z', '+00:00'))
        
        if not isinstance(existing_start, datetime) or not isinstance(existing_end, datetime):
            logger.warning(f"Geçersiz tarih formatı: {existing}")
            continue
        
        # Check overlap
        if shift_date <= existing_end and new_end >= existing_start:
            logger.warning(f"Çakışma tespit edildi - Mevcut atama: {existing_id}, Status: {existing_status}, Tarih: {existing_start} - {existing_end}")
            raise HTTPException(
                status_code=400,
                detail=f"{user.get('name')} için {existing_start.strftime('%d.%m.%Y')} - {existing_end.strftime('%d.%m.%Y')} tarihlerinde mevcut atama var (ID: {existing_id}, Durum: {existing_status}). Lütfen önce bu atamayı silin."
            )
    
    # Healmedy lokasyon adını bul
    healmedy_location_name = None
    if data.healmedy_location_id:
        from models import HEALMEDY_LOCATIONS
        location = next((l for l in HEALMEDY_LOCATIONS if l["id"] == data.healmedy_location_id), None)
        if location:
            healmedy_location_name = location["name"]
    
    # Create assignment
    assignment = ShiftAssignment(
        user_id=data.user_id,
        vehicle_id=data.vehicle_id if data.vehicle_id else None,
        location_type=data.location_type or "arac",
        health_center_name=data.health_center_name,
        assigned_by=current_user.id,
        shift_date=shift_date,
        start_time=data.start_time,
        end_time=data.end_time,
        end_date=end_date,
        healmedy_location_id=data.healmedy_location_id,
        healmedy_location_name=healmedy_location_name,
        assigned_role=data.assigned_role  # Geçici görev rolü
    )
    
    assignment_dict = assignment.model_dump(by_alias=True)
    result = await shift_assignments_collection.insert_one(assignment_dict)
    
    # Healmedy lokasyonu seçildiyse araç lokasyonunu güncelle
    if data.healmedy_location_id and data.vehicle_id and healmedy_location_name:
        from database import vehicle_current_locations_collection
        turkey_tz = pytz.timezone('Europe/Istanbul')
        turkey_now = datetime.now(turkey_tz)
        
        await vehicle_current_locations_collection.update_one(
            {"vehicle_id": data.vehicle_id},
            {"$set": {
                "vehicle_id": data.vehicle_id,
                "vehicle_plate": vehicle.get("plate") if vehicle else "",
                "assigned_location_id": data.healmedy_location_id,
                "assigned_location_name": healmedy_location_name,
                "current_location_id": data.healmedy_location_id,
                "current_location_name": healmedy_location_name,
                "updated_by": current_user.id,
                "updated_at": turkey_now
            }},
            upsert=True
        )
        logger.info(f"Araç lokasyonu güncellendi: {vehicle.get('plate') if vehicle else data.vehicle_id} → {healmedy_location_name}")
    
    inserted_doc = await shift_assignments_collection.find_one({"_id": result.inserted_id})
    if not inserted_doc:
        raise HTTPException(status_code=500, detail="Failed to create assignment")
    
    return serialize_assignment(inserted_doc)

@router.delete("/assignments/{assignment_id}")
async def delete_assignment(assignment_id: str, request: Request):
    """Delete shift assignment (Admin only)"""
    await require_roles(["merkez_ofis", "operasyon_muduru", "bas_sofor"])(request)
    
    result = await shift_assignments_collection.delete_one({"_id": assignment_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Assignment not found")
    
    return {"message": "Assignment deleted successfully"}

@router.post("/assignments/{assignment_id}/start")
async def admin_start_shift(assignment_id: str, request: Request):
    """Admin can start shift on behalf of user (Operasyon Müdürü, Merkez Ofis, Baş Şoför)"""
    admin_user = await require_roles(["merkez_ofis", "operasyon_muduru", "bas_sofor"])(request)
    
    from database import users_collection
    
    # Find the assignment
    assignment = await shift_assignments_collection.find_one({"_id": assignment_id})
    if not assignment:
        raise HTTPException(status_code=404, detail="Atama bulunamadı")
    
    if assignment.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Bu vardiya zaten başlatılmış veya tamamlanmış")
    
    # Get the assigned user
    assigned_user = await users_collection.find_one({"_id": assignment.get("user_id")})
    if not assigned_user:
        raise HTTPException(status_code=404, detail="Atanan kullanıcı bulunamadı")
    
    # Check if user already has an active shift
    active_shift = await shifts_collection.find_one({
        "user_id": assignment.get("user_id"),
        "status": {"$in": ["active", "on_break"]}
    })
    
    if active_shift:
        raise HTTPException(status_code=400, detail="Bu kullanıcının zaten aktif bir vardiyası var")
    
    # Get vehicle info if it's a vehicle assignment
    vehicle_plate = None
    vehicle_id = assignment.get("vehicle_id")
    if vehicle_id:
        vehicle = await vehicles_collection.find_one({"_id": vehicle_id})
        if vehicle:
            vehicle_plate = vehicle.get("plate")
            # Update vehicle status to "gorevde"
            await vehicles_collection.update_one(
                {"_id": vehicle_id},
                {"$set": {"status": "gorevde", "updated_at": get_turkey_time()}}
            )
    
    # Create shift
    new_shift = Shift(
        user_id=assignment.get("user_id"),
        vehicle_id=vehicle_id,
        vehicle_plate=vehicle_plate,
        start_km=0,  # Admin başlatınca km 0 olarak başlar
        status="active",
        location_type=assignment.get("location_type", "arac"),
        health_center_name=assignment.get("health_center_name"),
        started_by_admin=True,
        admin_id=admin_user.id,
        admin_note=f"{admin_user.name} tarafından manuel olarak başlatıldı"
    )
    
    shift_dict = new_shift.model_dump(by_alias=True)
    await shifts_collection.insert_one(shift_dict)
    
    # Update assignment status
    await shift_assignments_collection.update_one(
        {"_id": assignment_id},
        {
            "$set": {
                "status": "started",
                "started_at": get_turkey_time(),
                "started_by_admin": True,
                "admin_id": admin_user.id
            }
        }
    )
    
    return {
        "message": f"{assigned_user.get('name', 'Kullanıcı')} için vardiya başarıyla başlatıldı",
        "shift_id": str(shift_dict["_id"]),
        "user_name": assigned_user.get("name"),
        "vehicle_plate": vehicle_plate,
        "location_type": assignment.get("location_type")
    }


@router.post("/assignments/{assignment_id}/end")
async def admin_end_shift(assignment_id: str, request: Request):
    """Admin can end shift on behalf of user (Operasyon Müdürü, Merkez Ofis, Baş Şoför)"""
    admin_user = await require_roles(["merkez_ofis", "operasyon_muduru", "bas_sofor"])(request)
    
    from database import users_collection
    
    # Find the assignment
    assignment = await shift_assignments_collection.find_one({"_id": assignment_id})
    if not assignment:
        raise HTTPException(status_code=404, detail="Atama bulunamadı")
    
    if assignment.get("status") != "started":
        raise HTTPException(status_code=400, detail="Bu vardiya henüz başlatılmamış veya zaten tamamlanmış")
    
    # Get the assigned user
    assigned_user = await users_collection.find_one({"_id": assignment.get("user_id")})
    if not assigned_user:
        raise HTTPException(status_code=404, detail="Atanan kullanıcı bulunamadı")
    
    # Find the active shift for this user
    active_shift = await shifts_collection.find_one({
        "user_id": assignment.get("user_id"),
        "status": {"$in": ["active", "on_break"]}
    })
    
    if active_shift:
        # End the shift - Türkiye saati kullan
        end_time = get_turkey_time()
        start_time = active_shift.get("start_time", end_time)
        if isinstance(start_time, datetime):
            duration_minutes = int((end_time - start_time).total_seconds() / 60)
        else:
            duration_minutes = 0
        
        await shifts_collection.update_one(
            {"_id": active_shift["_id"]},
            {
                "$set": {
                    "status": "completed",
                    "end_time": end_time,
                    "duration_minutes": duration_minutes,
                    "notes": f"{admin_user.name} tarafından manuel olarak bitirildi",
                    "ended_by_admin": True,
                    "admin_end_id": admin_user.id
                }
            }
        )
        
        # Update vehicle status back to "musait"
        vehicle_id = assignment.get("vehicle_id")
        if vehicle_id:
            await vehicles_collection.update_one(
                {"_id": vehicle_id},
                {"$set": {"status": "musait", "updated_at": get_turkey_time()}}
            )
    
    # Update assignment status
    await shift_assignments_collection.update_one(
        {"_id": assignment_id},
        {
            "$set": {
                "status": "completed",
                "completed_at": get_turkey_time(),
                "ended_by_admin": True,
                "admin_end_id": admin_user.id
            }
        }
    )
    
    return {
        "message": f"{assigned_user.get('name', 'Kullanıcı')} için vardiya başarıyla bitirildi",
        "user_name": assigned_user.get("name"),
        "ended_at": get_turkey_time().isoformat()
    }


@router.post("/import-december-2024")
async def import_december_2024_shifts(request: Request):
    """Aralık 2024 nöbet listesini import et - Sadece operasyon müdürü ve merkez ofis"""
    import hashlib
    
    user = await require_roles(["merkez_ofis", "operasyon_muduru"])(request)
    
    ROLE_MAPPING = {
        "Dr.": "doktor", "Prm": "paramedik", "Tek.": "att", "Tek": "att",
        "Src.": "sofor", "Src": "sofor", "Hem.": "hemsire", "Hem": "hemsire",
        "Ç.M.": "cagri_merkezi", "T.G.": "temizlik"
    }
    
    def generate_email(name):
        tr_chars = {"ş": "s", "ğ": "g", "ü": "u", "ı": "i", "ö": "o", "ç": "c",
                    "Ş": "S", "Ğ": "G", "Ü": "U", "İ": "I", "Ö": "O", "Ç": "C"}
        clean_name = name.lower()
        for tr, en in tr_chars.items():
            clean_name = clean_name.replace(tr, en)
        parts = clean_name.split()
        email = f"{parts[0]}.{parts[-1]}@healmedy.com" if len(parts) >= 2 else f"{parts[0]}@healmedy.com"
        return email.replace(" ", "").lower()
    
    async def get_or_create_user(name, role_abbr):
        role = ROLE_MAPPING.get(role_abbr, "att")
        email = generate_email(name)
        existing = await users_collection.find_one({"email": email})
        if existing:
            return existing["_id"], False
        user_id = str(uuid.uuid4())
        await users_collection.insert_one({
            "_id": user_id, "email": email, "name": name, "role": role,
            "password_hash": hashlib.sha256("123456".encode()).hexdigest(),
            "is_active": True, "phone": "",
            "created_at": get_turkey_time(), "updated_at": get_turkey_time()
        })
        return user_id, True
    
    async def get_vehicle_id(plate):
        if not plate: return None
        vehicle = await vehicles_collection.find_one({"plate": plate})
        if vehicle: return vehicle["_id"]
        vehicle_id = str(uuid.uuid4())
        await vehicles_collection.insert_one({
            "_id": vehicle_id, "plate": plate, "type": "ambulans", "status": "musait",
            "km": 0, "qr_code": str(uuid.uuid4()),
            "created_at": get_turkey_time(), "updated_at": get_turkey_time()
        })
        return vehicle_id
    
    SCHEDULE = {
        "06_CHZ_142": {"plate": "06 CHZ 142", "personnel": [
            {"role": "Prm", "name": "Hatice Acar CANBAZ", "shifts": [1,4,7,10,13,16,19,22,25,28,31]},
            {"role": "Prm", "name": "Aleyna OZDEMIR", "shifts": [2,5,8,11,14,17,20,23,26,29]},
            {"role": "Prm", "name": "Hasan GUNEY", "shifts": [3,6,9,12,15,18,21,24,27,30]},
            {"role": "Tek.", "name": "Mustafa KARAGOL", "shifts": [1,4,7,10,13,16,19,22,25,28,31]},
            {"role": "Tek.", "name": "Derya GOMLEKSIZOGLU", "shifts": [3,6,9,12,15,18,21,24,27,30]},
            {"role": "Src.", "name": "Gorkem GURPUZER", "shifts": [1,4,7,10,13,16,19,22,25,28,31]},
            {"role": "Src.", "name": "Mert CINAR", "shifts": [2,5,8,11,14,17,20,23,26,29]},
            {"role": "Src.", "name": "Talha Dogukan KARTAL", "shifts": [3,6,9,12,15,18,21,24,27,30]},
        ]},
        "06_CHZ_146": {"plate": "06 CHZ 146", "personnel": [
            {"role": "Prm", "name": "Elif KURBAN", "shifts": [1,4,7,10,13,16,19,22,25,28,31]},
            {"role": "Prm", "name": "Burak ILIK", "shifts": [2,5,8,11,14,17,20,23,26,29]},
            {"role": "Prm", "name": "Busra Bahtiyar GUZEL", "shifts": [3,6,9,12,15,18,21,24,27,30]},
            {"role": "Tek.", "name": "Berkecan TURPCU", "shifts": [1,4,7,10,13,16,19,22,25,28,31]},
            {"role": "Tek.", "name": "Rumeysa UZUNAY", "shifts": [2,5,8,11,14,17,20,23,26,29]},
            {"role": "Tek.", "name": "Serkan KAMIT", "shifts": [3,6,9,12,15,18,21,24,27,30]},
            {"role": "Src.", "name": "Kubilay ELICORA", "shifts": [1,4,7,10,13,16,19,22,25,28,31]},
            {"role": "Src.", "name": "Samet KOCAPINAR", "shifts": [2,5,8,11,14,17,20,23,26,29]},
            {"role": "Src.", "name": "Oktay TUTUNCUOGLU", "shifts": [3,6,9,12,15,18,21,24,27,30]},
        ]},
        "06_CHZ_149": {"plate": "06 CHZ 149", "personnel": [
            {"role": "Prm", "name": "Aysegul Beyza YILMAZ", "shifts": [1,4,7,10,13,16,19,22,25,28,31]},
            {"role": "Prm", "name": "Ugur VAR", "shifts": [2,5,8,11,14,17,20,23,26,29]},
            {"role": "Prm", "name": "Mervenur GEDIK", "shifts": [3,6,9,12,15,18,21,24,27,30]},
            {"role": "Tek.", "name": "Efe Talha AKKAS", "shifts": [1,4,7,10,13,16,19,22,25,28,31]},
            {"role": "Tek.", "name": "Mugenur SOYKAN", "shifts": [2,5,8,11,14,17,20,23,26,29]},
            {"role": "Tek.", "name": "Tayfun KOCAMAN", "shifts": [3,6,9,12,15,18,21,24,27,30]},
            {"role": "Src.", "name": "Kadirhan ALKAN", "shifts": [1,4,7,10,13,16,19,22,25,28,31]},
            {"role": "Src.", "name": "Emirhan DOGAN", "shifts": [2,5,8,11,14,17,20,23,26,29]},
            {"role": "Src.", "name": "Melihcan DOGAN", "shifts": [3,6,9,12,15,18,21,24,27,30]},
        ]},
        "34_FTU_336": {"plate": "34 FTU 336", "personnel": [
            {"role": "Prm", "name": "Nesrin TUYSUZ", "shifts": [1,4,7,10,13,16,19,22,25,28,31]},
            {"role": "Prm", "name": "Gamze Hande BOZ", "shifts": [2,5,8,11,14,17,20,23,26,29]},
            {"role": "Prm", "name": "Alican TULUBAS", "shifts": [3,6,9,12,15,18,21,24,27,30]},
            {"role": "Tek.", "name": "Muzaffer OZCAN", "shifts": [1,4,7,10,13,16,19,22,25,28,31]},
            {"role": "Tek.", "name": "Fatih MEKIKCI", "shifts": [2,5,8,11,14,17,20,23,26,29]},
            {"role": "Tek.", "name": "Aysegul ORAL", "shifts": [3,6,9,12,15,18,21,24,27,30]},
            {"role": "Src.", "name": "Serkna Bilal BATTAL", "shifts": [1,4,7,10,13,16,19,22,25,28,31]},
            {"role": "Src.", "name": "Burak TIRYAKIOGLU", "shifts": [2,5,8,11,14,17,20,23,26,29]},
            {"role": "Src.", "name": "Feyzi FIDAN", "shifts": [3,6,9,12,15,18,21,24,27,30]},
        ]},
        "34_KMP_224": {"plate": "34 KMP 224", "personnel": [
            {"role": "Prm", "name": "Murat KESER", "shifts": [1,4,7,10,13,16,19,22,25,28,31]},
            {"role": "Prm", "name": "Melike KARATEPE", "shifts": [2,5,8,11,14,17,20,23,26,29]},
            {"role": "Prm", "name": "Burakcan SAHINTURK", "shifts": [3,6,9,12,15,18,21,24,27,30]},
            {"role": "Tek.", "name": "Busra AYDEMIR", "shifts": [1,4,7,10,13,16,19,22,25,28,31]},
            {"role": "Tek.", "name": "Muhammet BILICI", "shifts": [2,5,8,11,14,17,20,23,26,29]},
            {"role": "Tek.", "name": "Asli KOCOGLU", "shifts": [3,6,9,12,15,18,21,24,27,30]},
            {"role": "Src.", "name": "Onur YALIN", "shifts": [1,4,7,10,13,16,19,22,25,28,31]},
            {"role": "Src.", "name": "Mehmetcan SAVLI", "shifts": [2,5,8,11,14,17,20,23,26,29]},
            {"role": "Src.", "name": "Tuegay KOSE", "shifts": [3,6,9,12,15,18,21,24,27,30]},
        ]},
        "34_MHA_112": {"plate": "34 MHA 112", "personnel": [
            {"role": "Prm", "name": "Kadir ARTAR", "shifts": [1,4,7,10,13,16,19,22,25,28,31]},
            {"role": "Prm", "name": "Hamza Tarik ERMIS", "shifts": [2,5,8,11,14,17,20,23,26,29]},
            {"role": "Prm", "name": "Buse TOPCU", "shifts": [3,6,9,12,15,18,21,24,27,30]},
            {"role": "Tek.", "name": "Sule SATICI", "shifts": [1,4,7,10,13,16,19,22,25,28,31]},
            {"role": "Tek.", "name": "Ceren YIGIT", "shifts": [2,5,8,11,14,17,20,23,26,29]},
            {"role": "Tek.", "name": "Cem BALAT", "shifts": [3,6,9,12,15,18,21,24,27,30]},
            {"role": "Src.", "name": "Murat GULSEN", "shifts": [1,4,7,10,13,16,19,22,25,28,31]},
            {"role": "Src.", "name": "Anil BALCI", "shifts": [2,5,8,11,14,17,20,23,26,29]},
            {"role": "Src.", "name": "Mesut CINKAVUK", "shifts": [3,6,9,12,15,18,21,24,27,30]},
        ]},
    }
    
    new_users = existing_users = new_shifts = skipped_shifts = 0
    
    for loc_key, loc_data in SCHEDULE.items():
        vehicle_id = await get_vehicle_id(loc_data["plate"])
        for person in loc_data["personnel"]:
            user_id, is_new = await get_or_create_user(person["name"], person["role"])
            if is_new: new_users += 1
            else: existing_users += 1
            for day in person["shifts"]:
                # Aralık ayı, mevcut yıl kullan (2025)
                current_year = datetime.now().year
                shift_date = datetime(current_year, 12, day)
                if await shift_assignments_collection.find_one({"user_id": user_id, "shift_date": shift_date}):
                    skipped_shifts += 1
                    continue
                # 24 saat vardiya: 08:00 - ertesi gün 08:00
                next_day = shift_date + timedelta(days=1)
                await shift_assignments_collection.insert_one({
                    "_id": str(uuid.uuid4()), "user_id": user_id, "vehicle_id": vehicle_id,
                    "shift_date": shift_date, "start_time": "08:00", "end_time": "08:00",
                    "end_date": next_day,  # Ertesi gün
                    "shift_type": "saha_24",  # 24 saat vardiya
                    "location_type": "arac", "status": "pending", "created_at": get_turkey_time()
                })
                new_shifts += 1
    
    logger.info(f"December {current_year} import: {new_users} new users, {new_shifts} new shifts")
    return {"message": f"Aralık {current_year} nöbet listesi import edildi", "new_users": new_users,
            "existing_users": existing_users, "new_shifts": new_shifts, "skipped_shifts": skipped_shifts}


@router.post("/start", response_model=Shift)
async def start_shift(data: ShiftStart, request: Request):
    """Start shift by scanning vehicle QR code"""
    user = await get_current_user(request)
    turkey_now = datetime.utcnow() + timedelta(hours=3)
    
    # Log: Form açılma zamanı (eğer gönderilmişse)
    if hasattr(data, 'form_opened_at') and data.form_opened_at:
        shift_logs_collection = db["shift_logs"]
        log_entry = {
            "_id": str(uuid.uuid4()),
            "user_id": user.id,
            "user_name": user.name,
            "user_role": user.role,
            "action": "form_opened",
            "action_type": "shift_start",
            "form_opened_at": datetime.fromisoformat(data.form_opened_at.replace('Z', '+00:00')),
            "logged_at": turkey_now,
            "vehicle_qr": data.vehicle_qr
        }
        await shift_logs_collection.insert_one(log_entry)
        logger.info(f"Vardiya başlatma formu açıldı: {user.name} - {data.form_opened_at}")
    
    # Find vehicle by QR code
    vehicle = await vehicles_collection.find_one({"qr_code": data.vehicle_qr})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Araç bulunamadı. QR kodu geçersiz.")
    
    # Check if user has TODAY's assignment for this vehicle
    # Türkiye saati (UTC+3) kullan - timedelta zaten import edildi
    turkey_now = datetime.utcnow() + timedelta(hours=3)
    today = turkey_now.date()
    yesterday = today - timedelta(days=1)  # Dün de kabul et (gece yarısı toleransı)
    logger.info(f"Vardiya başlatma kontrolü - Bugün (TR): {today}, Dün (TR): {yesterday}, Kullanıcı: {user.id}, Araç: {vehicle.get('plate')}")
    
    # Get all active assignments for this user and vehicle (pending veya started)
    all_assignments = await shift_assignments_collection.find({
        "user_id": user.id,
        "vehicle_id": vehicle["_id"],
        "status": {"$in": ["pending", "started"]}  # Başlatılmış vardiyalar da kabul
    }).to_list(100)
    
    # Filter to find an assignment that is valid for today
    valid_assignment = None
    for assignment in all_assignments:
        shift_date_str = assignment.get("shift_date", "")
        end_date_str = assignment.get("end_date", "")
        
        # Parse shift date
        if isinstance(shift_date_str, datetime):
            shift_date = shift_date_str.date()
        elif isinstance(shift_date_str, str):
            try:
                if 'T' in shift_date_str:
                    shift_date = datetime.fromisoformat(shift_date_str.replace('Z', '+00:00')).date()
                else:
                    shift_date = datetime.strptime(shift_date_str, "%Y-%m-%d").date()
            except:
                continue
        else:
            continue
        
        # Parse end date
        end_date = shift_date  # Default to same day
        if end_date_str:
            if isinstance(end_date_str, datetime):
                end_date = end_date_str.date()
            elif isinstance(end_date_str, str):
                try:
                    if 'T' in end_date_str:
                        end_date = datetime.fromisoformat(end_date_str.replace('Z', '+00:00')).date()
                    else:
                        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
                except:
                    pass
        
        # Check if today OR yesterday falls within the assignment period
        # (Gece yarısı toleransı - vardiya gece yarısını geçebilir)
        is_valid_today = shift_date <= today <= end_date
        is_valid_yesterday = shift_date <= yesterday <= end_date
        
        logger.info(f"Atama kontrolü: {shift_date} - {end_date}, Bugün geçerli: {is_valid_today}, Dün geçerli: {is_valid_yesterday}")
        
        if is_valid_today or is_valid_yesterday:
            valid_assignment = assignment
            break
    
    if not valid_assignment:
        # Get user's actual assignment for today to show in error
        user_today_assignment = await shift_assignments_collection.find_one({
            "user_id": user.id,
            "status": "pending"
        })
        
        if user_today_assignment:
            assigned_vehicle = await vehicles_collection.find_one({"_id": user_today_assignment.get("vehicle_id")})
            assigned_plate = assigned_vehicle.get("plate", "Bilinmiyor") if assigned_vehicle else "Bilinmiyor"
            raise HTTPException(
                status_code=403, 
                detail=f"Bu araç ({vehicle.get('plate')}) size atanmamış. Bugün için atanan aracınız: {assigned_plate}"
            )
        else:
            raise HTTPException(
                status_code=403, 
                detail="Bu araç için bugün geçerli vardiya atamanız yok. Lütfen yöneticinizle iletişime geçin."
            )
    
    assignment = valid_assignment
    
    # Check if user already has an active shift
    active_shift = await shifts_collection.find_one({
        "user_id": user.id,
        "end_time": None
    })
    
    if active_shift:
        raise HTTPException(status_code=400, detail="Zaten aktif bir vardiyanz var")
    
    # Create new shift
    new_shift = Shift(
        assignment_id=assignment["_id"],
        user_id=user.id,
        vehicle_id=vehicle["_id"],
        vehicle_plate=vehicle.get("plate"),  # Plaka bilgisini ekle
        start_time=get_turkey_time(),
        photos=data.photos,
        daily_control=data.daily_control
    )
    
    shift_dict = new_shift.model_dump(by_alias=True)
    await shifts_collection.insert_one(shift_dict)
    
    # Update assignment status
    await shift_assignments_collection.update_one(
        {"_id": assignment["_id"]},
        {"$set": {"status": "started"}}
    )
    
    # Günlük kontrol formunu form geçmişine kaydet
    if data.daily_control:
        from database import forms_collection
        from utils.sanitize import sanitize_form_data
        
        # Form verilerini temizle
        sanitized_daily_control = sanitize_form_data(data.daily_control) if data.daily_control else {}
        
        daily_control_form = {
            "_id": str(uuid.uuid4()),
            "form_type": "daily_control",
            "submitted_by": user.id,
            "form_data": sanitized_daily_control,
            "vehicle_plate": vehicle.get("plate"),
            "vehicle_id": vehicle["_id"],
            "shift_id": new_shift.id,
            "created_at": get_turkey_time()
        }
        await forms_collection.insert_one(daily_control_form)
        logger.info(f"Günlük kontrol formu kaydedildi: {daily_control_form['_id']}")
        
        # Günlük formu doldurulmuş olarak işaretle - aynı araçta diğer ekip üyesi doldurmasın
        await shifts_collection.update_one(
            {"_id": new_shift.id},
            {"$set": {
                "daily_control_filled_by": user.id,
                "daily_control_filled_at": turkey_now
            }}
        )
        logger.info(f"Günlük kontrol formu dolduruldu işaretlendi: user={user.id}, shift={new_shift.id}")
    
    # Araç fotoğraflarını kaydet (ayrı bir collection'da)
    if data.photos:
        shift_photos_collection = db["shift_photos"]
        
        photos_doc = {
            "_id": str(uuid.uuid4()),
            "shift_id": new_shift.id,
            "user_id": user.id,
            "vehicle_id": vehicle["_id"],
            "vehicle_plate": vehicle.get("plate"),
            "photos": data.photos,
            "created_at": get_turkey_time()
        }
        await shift_photos_collection.insert_one(photos_doc)
        logger.info(f"Vardiya fotoğrafları kaydedildi: {photos_doc['_id']}")
    
    # Log: Vardiya başlatma işlemi
    shift_logs_collection = db["shift_logs"]
    action_taken_at = turkey_now
    if hasattr(data, 'action_taken_at') and data.action_taken_at:
        try:
            action_taken_at = datetime.fromisoformat(data.action_taken_at.replace('Z', '+00:00'))
        except:
            action_taken_at = turkey_now
    
    form_opened_at_val = None
    if hasattr(data, 'form_opened_at') and data.form_opened_at:
        try:
            form_opened_at_val = datetime.fromisoformat(data.form_opened_at.replace('Z', '+00:00'))
        except:
            pass
    
    action_log = {
        "_id": str(uuid.uuid4()),
        "user_id": user.id,
        "user_name": user.name,
        "user_role": user.role,
        "action": "shift_started",
        "action_type": "shift_start",
        "shift_id": new_shift.id,
        "vehicle_id": vehicle["_id"],
        "vehicle_plate": vehicle.get("plate"),
        "action_taken_at": action_taken_at,
        "logged_at": turkey_now,
        "form_opened_at": form_opened_at_val
    }
    await shift_logs_collection.insert_one(action_log)
    
    # Otomatik onay kaydı oluştur (pending olarak - yönetici onayı beklenmeden vardiya başlatıldı)
    role_type = "medical" if user.role.lower() in ['att', 'paramedik', 'hemsire'] else "driver"
    approval_id = str(uuid.uuid4())
    approval = {
        "_id": approval_id,
        "shift_id": new_shift.id,  # Shift referansı ekle
        "user_id": user.id,
        "user_name": user.name,
        "user_role": user.role,
        "vehicle_id": vehicle["_id"],
        "vehicle_plate": vehicle.get("plate"),
        "role_type": role_type,
        "daily_control_data": data.daily_control if hasattr(data, 'daily_control') else None,
        "photos": data.photos if hasattr(data, 'photos') else None,
        "status": "pending",  # Otomatik oluşturuldu, yönetici onayı bekliyor
        "created_at": turkey_now,
        "shift_started_at": action_taken_at,  # Vardiya başlatılma zamanı
        "auto_created": True,  # Otomatik oluşturuldu işareti
        "approved_by": None,
        "approved_by_name": None,
        "approved_at": None,
        "rejection_reason": None
    }
    
    # Eğer approval_id gönderilmişse (manuel onay isteği varsa), onu kullan
    if hasattr(data, 'approval_id') and data.approval_id:
        # Mevcut onay kaydını güncelle
        await shift_start_approvals_collection.update_one(
            {"_id": data.approval_id},
            {"$set": {
                "shift_started_at": action_taken_at,
                "auto_created": False
            }}
        )
    else:
        # Yeni onay kaydı oluştur
        await shift_start_approvals_collection.insert_one(approval)
    
    logger.info(f"Vardiya başlatıldı: {user.name} - {vehicle.get('plate')} - {action_taken_at} (Onay kaydı: {approval_id})")
    
    return new_shift

@router.post("/end", response_model=Shift)
async def end_shift(data: ShiftEnd, request: Request):
    """End shift"""
    user = await get_current_user(request)
    turkey_now = datetime.utcnow() + timedelta(hours=3)
    
    # Find active shift
    shift_doc = await shifts_collection.find_one({
        "_id": data.shift_id,
        "user_id": user.id,
        "end_time": None
    })
    
    if not shift_doc:
        raise HTTPException(status_code=404, detail="Active shift not found")
    
    # Log: Form açılma zamanı (eğer gönderilmişse)
    form_opened_at = None
    if hasattr(data, 'handover_form') and data.handover_form and data.handover_form.get('form_opened_at'):
        shift_logs_collection = db["shift_logs"]
        form_opened_at = datetime.fromisoformat(data.handover_form['form_opened_at'].replace('Z', '+00:00'))
        log_entry = {
            "_id": str(uuid.uuid4()),
            "user_id": user.id,
            "user_name": user.name,
            "user_role": user.role,
            "action": "form_opened",
            "action_type": "shift_end",
            "shift_id": data.shift_id,
            "form_opened_at": form_opened_at,
            "logged_at": turkey_now
        }
        await shift_logs_collection.insert_one(log_entry)
        logger.info(f"Vardiya bitirme formu açıldı: {user.name} - {form_opened_at}")
    
    # Calculate duration - Türkiye saati kullan
    end_time = get_turkey_time()
    start_time = shift_doc["start_time"]
    # start_time zaten UTC+3 olarak kaydedilmiş
    if isinstance(start_time, datetime):
        duration_minutes = int((end_time - start_time).total_seconds() / 60)
    else:
        duration_minutes = 0
    
    # Update shift
    result = await shifts_collection.find_one_and_update(
        {"_id": data.shift_id},
        {
            "$set": {
                "end_time": end_time,
                "duration_minutes": duration_minutes,
                "notes": data.notes
            }
        },
        return_document=True
    )
    
    # Assignment status'unu "completed" olarak güncelle
    if shift_doc.get("assignment_id"):
        await shift_assignments_collection.update_one(
            {"_id": shift_doc["assignment_id"]},
            {"$set": {"status": "completed"}}
        )
        logger.info(f"Assignment {shift_doc['assignment_id']} marked as completed")
    
    # Log: Vardiya bitirme işlemi
    shift_logs_collection = db["shift_logs"]
    action_taken_at = turkey_now
    if hasattr(data, 'handover_form') and data.handover_form and data.handover_form.get('action_taken_at'):
        action_taken_at = datetime.fromisoformat(data.handover_form['action_taken_at'].replace('Z', '+00:00'))
    
    action_log = {
        "_id": str(uuid.uuid4()),
        "user_id": user.id,
        "user_name": user.name,
        "user_role": user.role,
        "action": "shift_ended",
        "action_type": "shift_end",
        "shift_id": data.shift_id,
        "vehicle_id": shift_doc.get("vehicle_id"),
        "vehicle_plate": shift_doc.get("vehicle_plate"),
        "action_taken_at": action_taken_at,
        "logged_at": turkey_now,
        "form_opened_at": form_opened_at
    }
    await shift_logs_collection.insert_one(action_log)
    
    # Otomatik onay kaydı oluştur (pending olarak - yönetici onayı beklenmeden vardiya bitirildi)
    role_type = "driver" if user.role.lower() == "sofor" else "medical"
    approval_id = str(uuid.uuid4())
    
    # Devir teslim bilgilerini al
    devralan_adi = None
    end_km = None
    if hasattr(data, 'handover_form') and data.handover_form:
        devralan_adi = data.handover_form.get('devralan_adi')
        end_km = data.handover_form.get('teslimEttigimKm')
    
    approval_doc = {
        "_id": approval_id,
        "type": "end",
        "shift_id": data.shift_id,
        "vehicle_id": shift_doc.get("vehicle_id"),
        "vehicle_plate": shift_doc.get("vehicle_plate"),
        "user_id": user.id,
        "user_name": user.name,
        "user_role": user.role,
        "role_type": role_type,
        "end_km": end_km,
        "devralan_adi": devralan_adi,
        "form_opened_at": form_opened_at.isoformat() if form_opened_at else None,
        "request_sent_at": turkey_now.isoformat(),
        "shift_ended_at": action_taken_at,  # Vardiya bitirilme zamanı
        "status": "pending",  # Otomatik oluşturuldu, yönetici onayı bekliyor
        "auto_created": True,  # Otomatik oluşturuldu işareti
        "created_at": turkey_now,
        "approved_by": None,
        "approved_by_name": None,
        "approved_at": None
    }
    
    # Eğer approval_id gönderilmişse (manuel onay isteği varsa), onu kullan
    if hasattr(data, 'handover_form') and data.handover_form and data.handover_form.get('approval_request_id'):
        # Mevcut onay kaydını güncelle
        await shift_end_approvals_collection.update_one(
            {"_id": data.handover_form['approval_request_id']},
            {"$set": {
                "shift_ended_at": action_taken_at,
                "auto_created": False
            }}
        )
    else:
        # Yeni onay kaydı oluştur
        await shift_end_approvals_collection.insert_one(approval_doc)
    
    logger.info(f"Vardiya bitirildi: {user.name} - {shift_doc.get('vehicle_plate', 'N/A')} - {action_taken_at} (Onay kaydı: {approval_id})")
    
    result["id"] = result.pop("_id")
    return Shift(**result)

@router.get("/active")
async def get_active_shift(request: Request):
    """Get user's active shift with vehicle info"""
    user = await get_current_user(request)
    
    shift_doc = await shifts_collection.find_one({
        "user_id": user.id,
        "end_time": None
    })
    
    if not shift_doc:
        return None
    
    shift_doc["id"] = shift_doc.pop("_id")
    
    # Araç bilgisini ekle
    vehicle_id = shift_doc.get("vehicle_id")
    if vehicle_id and not shift_doc.get("vehicle_plate"):
        vehicle = await vehicles_collection.find_one({"_id": vehicle_id})
        if vehicle:
            shift_doc["vehicle_plate"] = vehicle.get("plate", "")
            shift_doc["vehicle_type"] = vehicle.get("type", "")
            shift_doc["vehicle"] = {
                "plate": vehicle.get("plate"),
                "type": vehicle.get("type"),
                "km": vehicle.get("km")
            }
    
    return shift_doc

@router.get("/history")
async def get_shift_history(
    request: Request,
    user_id: Optional[str] = None,
    limit: int = 50
):
    """Get shift history with vehicle info"""
    current_user = await get_current_user(request)
    
    # If user_id provided, check permission
    if user_id and user_id != current_user.id:
        if current_user.role not in ["merkez_ofis", "operasyon_muduru"]:
            raise HTTPException(status_code=403, detail="Permission denied")
    else:
        user_id = current_user.id
    
    shifts = await shifts_collection.find(
        {"user_id": user_id}
    ).sort("start_time", -1).limit(limit).to_list(limit)
    
    result = []
    for shift in shifts:
        shift["id"] = shift.pop("_id")
        # Araç bilgisini ekle
        vehicle_id = shift.get("vehicle_id")
        if vehicle_id and not shift.get("vehicle_plate"):
            vehicle = await vehicles_collection.find_one({"_id": vehicle_id})
            if vehicle:
                shift["vehicle_plate"] = vehicle.get("plate", "")
        result.append(shift)
    
    return result

@router.get("/photos")
async def get_shift_photos(
    request: Request,
    vehicle_plate: Optional[str] = None,
    limit: int = 50
):
    """
    Vardiya fotoğraflarını getir
    Tüm roller erişebilir (form geçmişi için)
    """
    try:
        current_user = await get_current_user(request)
        
        from database import db
        shift_photos_collection = db["shift_photos"]
        
        query = {}
        if vehicle_plate:
            query["vehicle_plate"] = vehicle_plate
        
        photos = await shift_photos_collection.find(query).sort("created_at", -1).limit(limit).to_list(limit)
        
        # Kullanıcı bilgisini ekle
        result = []
        for photo in photos:
            try:
                photo_id = photo.pop("_id", None)
                if photo_id:
                    photo["id"] = str(photo_id)
                
                user_id = photo.get("user_id")
                if user_id:
                    user_doc = await users_collection.find_one({"_id": user_id})
                    if user_doc:
                        photo["user_name"] = user_doc.get("name", "Bilinmiyor")
                        photo["user_role"] = user_doc.get("role", "")
                    else:
                        photo["user_name"] = "Bilinmiyor"
                        photo["user_role"] = ""
                else:
                    photo["user_name"] = "Bilinmiyor"
                    photo["user_role"] = ""
                
                # Tarihi string'e çevir (JSON serialization için)
                if "created_at" in photo and photo["created_at"]:
                    if hasattr(photo["created_at"], 'isoformat'):
                        photo["created_at"] = photo["created_at"].isoformat()
                
                result.append(photo)
            except Exception as e:
                logger.warning(f"Fotoğraf işleme hatası: {e}")
                continue
        
        return result
    except Exception as e:
        logger.error(f"Fotoğraf getirme hatası: {e}")
        return []

@router.get("/photos/{shift_id}")
async def get_shift_photos_by_id(shift_id: str, request: Request):
    """Belirli bir vardiyaya ait fotoğrafları getir - Tüm roller erişebilir
    Hem shift_photos collection'ından hem de shift'in kendisinden (end_photos) fotoğrafları getirir
    """
    current_user = await get_current_user(request)
    
    from database import db
    shift_photos_collection = db["shift_photos"]
    
    # Önce shift_photos collection'ından ara
    photo_doc = await shift_photos_collection.find_one({"shift_id": shift_id})
    
    # Shift'in kendisinden de fotoğrafları al (ATT/Paramedik için end_photos)
    shift = await shifts_collection.find_one({"_id": shift_id})
    
    if not photo_doc and not shift:
        return None
    
    # Eğer photo_doc yoksa ama shift varsa, shift'ten fotoğrafları al
    if not photo_doc and shift:
        end_photos = shift.get("end_photos") or {}
        photos = shift.get("photos") or {}
        
        # End photos ve photos'u birleştir
        all_photos = {**photos, **end_photos}
        
        if all_photos:
            photo_doc = {
                "id": shift_id,
                "shift_id": shift_id,
                "user_id": shift.get("user_id"),
                "vehicle_id": shift.get("vehicle_id"),
                "vehicle_plate": shift.get("vehicle_plate"),
                "photos": all_photos,
                "created_at": shift.get("start_time")
            }
    
    if photo_doc:
        photo_doc["id"] = photo_doc.pop("_id", photo_doc.get("id"))
        
        # Kullanıcı bilgisini ekle
        user_doc = await users_collection.find_one({"_id": photo_doc.get("user_id")})
        if user_doc:
            photo_doc["user_name"] = user_doc.get("name", "Bilinmiyor")
            photo_doc["user_role"] = user_doc.get("role", "")
    
    return photo_doc

@router.get("/stats/summary")
async def get_shift_stats(request: Request, user_id: Optional[str] = None):
    """Get shift statistics"""
    current_user = await get_current_user(request)
    
    if user_id and user_id != current_user.id:
        if current_user.role not in ["merkez_ofis", "operasyon_muduru"]:
            raise HTTPException(status_code=403, detail="Permission denied")
    else:
        user_id = current_user.id
    
    # Active shifts count
    active = await shifts_collection.count_documents({
        "user_id": user_id,
        "end_time": None
    })
    
    # Total shifts
    total = await shifts_collection.count_documents({"user_id": user_id})
    
    # Total hours worked
    completed_shifts = await shifts_collection.find({
        "user_id": user_id,
        "end_time": {"$ne": None}
    }).to_list(1000)
    
    total_minutes = sum(shift.get("duration_minutes", 0) for shift in completed_shifts)
    total_hours = total_minutes / 60
    
    return {
        "active_shifts": active,
        "total_shifts": total,
        "total_hours_worked": round(total_hours, 2)
    }


# ============================================================================
# GÜNLÜK FORM KONTROLÜ - EKİP BAZLI
# ============================================================================

@router.get("/check-daily-form/{vehicle_id}")
async def check_daily_form_filled(vehicle_id: str, request: Request, date: Optional[str] = None):
    """
    Bu araç için bugün günlük kontrol formu doldurulmuş mu?
    ATT/Paramedik'ten biri doldurduysa diğerinin doldurmasına gerek yok
    SADECE AYNI VARDİYA ATAMASINDA OLAN EKİP ÜYELERİ DOLDURABİLİR
    """
    from database import forms_collection
    
    user = await get_current_user(request)
    turkey_now = get_turkey_time()
    
    if date:
        try:
            target_date = datetime.strptime(date, "%Y-%m-%d")
            # Timezone bilgisi ekle
            turkey_tz = pytz.timezone('Europe/Istanbul')
            target_date = turkey_tz.localize(target_date)
        except ValueError:
            target_date = turkey_now
    else:
        target_date = turkey_now
    
    # Bugünün başlangıcı ve bitişi (timezone-aware)
    turkey_tz = pytz.timezone('Europe/Istanbul')
    day_start = turkey_tz.localize(datetime(target_date.year, target_date.month, target_date.day))
    day_end = day_start + timedelta(days=1)
    
    # Naive datetime versiyonları (bazı alanlar naive olarak kaydedilmiş olabilir)
    day_start_naive = datetime(target_date.year, target_date.month, target_date.day)
    day_end_naive = day_start_naive + timedelta(days=1)
    
    logger.info(f"Günlük form kontrolü: vehicle_id={vehicle_id}, user={user.id}, date={target_date.date()}")
    
    # Kullanıcının bugünkü vardiya assignment'ını bul
    today_naive = datetime(turkey_now.year, turkey_now.month, turkey_now.day)
    tomorrow_naive = today_naive + timedelta(days=1)
    
    user_assignment = await shift_assignments_collection.find_one({
        "user_id": user.id,
        "vehicle_id": vehicle_id,
        "status": {"$in": ["pending", "started"]},
        "$or": [
            {"shift_date": {"$gte": today_naive, "$lt": tomorrow_naive}},
            {"shift_date": {"$lte": today_naive}, "end_date": {"$gte": today_naive}}
        ]
    })
    
    if not user_assignment:
        logger.info(f"Kullanıcının ataması yok: user={user.id}, vehicle={vehicle_id}")
        return {
            "filled": False,
            "message": "Bu araç için aktif vardiya atamanız bulunmuyor",
            "no_assignment": True
        }
    
    # Bu vardiya assignment'ındaki diğer kullanıcıları bul (aynı araç, aynı tarih)
    assignment_date = user_assignment.get("shift_date")
    assignment_end_date = user_assignment.get("end_date") or assignment_date
    
    # Aynı assignment'taki diğer kullanıcıların user_id'lerini al
    same_assignment_users = await shift_assignments_collection.find({
        "vehicle_id": vehicle_id,
        "status": {"$in": ["pending", "started"]},
        "$or": [
            {"shift_date": assignment_date, "end_date": assignment_end_date},
            {"shift_date": {"$gte": today_naive, "$lt": tomorrow_naive}},
            {"shift_date": {"$lte": today_naive}, "end_date": {"$gte": today_naive}}
        ]
    }).to_list(100)
    
    team_user_ids = [a.get("user_id") for a in same_assignment_users]
    logger.info(f"Ekip üyeleri: {team_user_ids}")
    
    # 1. ÖNCELİKLE forms_collection'dan kontrol et (daha güvenilir)
    daily_form = await forms_collection.find_one({
        "vehicle_id": vehicle_id,
        "form_type": "daily_control",
        "submitted_by": {"$in": team_user_ids},
        "$or": [
            {"created_at": {"$gte": day_start, "$lt": day_end}},
            {"created_at": {"$gte": day_start_naive, "$lt": day_end_naive}}
        ]
    })
    
    if daily_form:
        filler_id = daily_form.get("submitted_by")
        filler = await users_collection.find_one({"_id": filler_id})
        filler_name = filler.get("name") if filler else "Bilinmiyor"
        filler_role = filler.get("role") if filler else ""
        
        logger.info(f"Form forms_collection'da bulundu: filler={filler_name}, role={filler_role}")
        
        # Dolduran kişi ATT veya Paramedik mi kontrol et
        if filler_role not in ["att", "paramedik"]:
            return {
                "filled": False,
                "message": f"Bu form {filler_name} tarafından doldurulmuş, ancak ATT/Paramedik tarafından doldurulması gerekiyor",
                "can_fill": True
            }
        
        return {
            "filled": True,
            "filled_by": filler_id,
            "filled_by_name": filler_name,
            "filled_at": daily_form.get("created_at"),
            "shift_id": daily_form.get("shift_id"),
            "message": f"Günlük kontrol formu {filler_name} tarafından doldurulmuş"
        }
    
    # 2. Yedek olarak shifts_collection'dan da kontrol et
    filled_shift = await shifts_collection.find_one({
        "vehicle_id": vehicle_id,
        "daily_control_filled_by": {"$in": team_user_ids},
        "$or": [
            {"daily_control_filled_at": {"$gte": day_start, "$lt": day_end}},
            {"daily_control_filled_at": {"$gte": day_start_naive, "$lt": day_end_naive}}
        ]
    })
    
    if filled_shift:
        filler = await users_collection.find_one({"_id": filled_shift.get("daily_control_filled_by")})
        filler_name = filler.get("name") if filler else "Bilinmiyor"
        filler_role = filler.get("role") if filler else ""
        
        logger.info(f"Form shifts_collection'da bulundu: filler={filler_name}, role={filler_role}")
        
        if filler_role not in ["att", "paramedik"]:
            return {
                "filled": False,
                "message": f"Bu form {filler_name} tarafından doldurulmuş, ancak ATT/Paramedik tarafından doldurulması gerekiyor",
                "can_fill": True
            }
        
        return {
            "filled": True,
            "filled_by": filled_shift.get("daily_control_filled_by"),
            "filled_by_name": filler_name,
            "filled_at": filled_shift.get("daily_control_filled_at"),
            "shift_id": filled_shift.get("_id"),
            "message": f"Günlük kontrol formu {filler_name} tarafından doldurulmuş"
        }
    
    logger.info(f"Form bulunamadı - doldurulmamış: vehicle={vehicle_id}")
    return {
        "filled": False,
        "message": "Bu araç için bugün günlük kontrol formu doldurulmamış"
    }


@router.post("/log-section-time")
async def log_section_time(request: Request):
    """
    Form doldurma sürelerini logla (ATT/Paramedik için zaman kısıtlamaları)
    """
    user = await get_current_user(request)
    body = await request.json()
    
    shift_id = body.get("shift_id")
    section_index = body.get("section_index")
    duration_seconds = body.get("duration_seconds")
    
    if not shift_id:
        raise HTTPException(status_code=400, detail="shift_id gerekli")
    
    turkey_now = datetime.utcnow() + timedelta(hours=3)
    
    # Mevcut section_times'ı al
    shift = await shifts_collection.find_one({"_id": shift_id})
    if not shift:
        raise HTTPException(status_code=404, detail="Vardiya bulunamadı")
    
    section_times = shift.get("section_times") or {}
    section_times[f"section_{section_index}"] = {
        "duration_seconds": duration_seconds,
        "completed_at": turkey_now.isoformat()
    }
    
    await shifts_collection.update_one(
        {"_id": shift_id},
        {"$set": {"section_times": section_times}}
    )
    
    return {"message": f"Section {section_index} süresi kaydedildi"}


@router.post("/mark-daily-form-filled")
async def mark_daily_form_filled(request: Request):
    """
    Günlük kontrol formunu doldurulmuş olarak işaretle
    """
    user = await get_current_user(request)
    body = await request.json()
    
    shift_id = body.get("shift_id")
    
    if not shift_id:
        raise HTTPException(status_code=400, detail="shift_id gerekli")
    
    turkey_now = datetime.utcnow() + timedelta(hours=3)
    
    await shifts_collection.update_one(
        {"_id": shift_id},
        {"$set": {
            "daily_control_filled_by": user.id,
            "daily_control_filled_at": turkey_now
        }}
    )
    
    logger.info(f"Günlük kontrol formu dolduruldu: {shift_id} - {user.name}")
    
    return {"message": "Günlük kontrol formu doldurulmuş olarak işaretlendi"}


# ============================================================================
# TOPLU VARDİYA ATAMA - EXCEL İLE
# ============================================================================

from fastapi import UploadFile, File
from io import BytesIO

@router.post("/bulk-upload")
async def bulk_upload_shifts(file: UploadFile = File(...), request: Request = None):
    """
    Excel dosyası ile toplu vardiya atama
    Sadece Mesul Müdür, Baş Şoför ve Merkez Ofis kullanabilir
    
    Excel formatı:
    | TC/Email | Araç Plaka | Lokasyon Tipi | Lokasyon Adı | Başlangıç Tarihi | Bitiş Tarihi | Vardiya Tipi |
    | 12345678901 | 34 ABC 123 | arac | - | 2025-01-15 | 2025-01-16 | saha_24 |
    | email@test.com | - | saglik_merkezi | Filyos SM | 2025-01-15 | 2025-01-15 | ofis_8 |
    """
    current_user = await require_roles(["merkez_ofis", "operasyon_muduru", "bas_sofor", "mesul_mudur"])(request)
    
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl kütüphanesi yüklü değil")
    
    # Dosyayı oku
    contents = await file.read()
    workbook = openpyxl.load_workbook(BytesIO(contents))
    sheet = workbook.active
    
    results = {
        "success": [],
        "errors": [],
        "total_rows": 0,
        "successful_count": 0,
        "error_count": 0
    }
    
    # İlk satır başlık, 2. satırdan itibaren veri
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:  # Boş satır
            continue
        
        results["total_rows"] += 1
        
        try:
            tc_or_email = str(row[0]).strip() if row[0] else None
            vehicle_plate = str(row[1]).strip() if row[1] else None
            location_type = str(row[2]).strip().lower() if row[2] else "arac"
            location_name = str(row[3]).strip() if row[3] else None
            start_date_str = str(row[4]).strip() if row[4] else None
            end_date_str = str(row[5]).strip() if row[5] else None
            shift_type = str(row[6]).strip() if row[6] else "saha_24"
            
            if not tc_or_email or not start_date_str:
                results["errors"].append({
                    "row": row_idx,
                    "error": "TC/Email ve başlangıç tarihi zorunlu"
                })
                results["error_count"] += 1
                continue
            
            # Kullanıcıyı bul (TC veya email ile)
            user_doc = None
            if "@" in tc_or_email:
                user_doc = await users_collection.find_one({"email": tc_or_email})
            else:
                user_doc = await users_collection.find_one({"tc_no": tc_or_email})
            
            if not user_doc:
                results["errors"].append({
                    "row": row_idx,
                    "error": f"Kullanıcı bulunamadı: {tc_or_email}"
                })
                results["error_count"] += 1
                continue
            
            # Araç bul (eğer araç atama ise)
            vehicle_id = None
            if location_type == "arac" and vehicle_plate:
                vehicle_doc = await vehicles_collection.find_one({"plate": vehicle_plate})
                if not vehicle_doc:
                    results["errors"].append({
                        "row": row_idx,
                        "error": f"Araç bulunamadı: {vehicle_plate}"
                    })
                    results["error_count"] += 1
                    continue
                vehicle_id = vehicle_doc["_id"]
            
            # Tarihleri parse et
            from datetime import datetime as dt
            try:
                if isinstance(row[4], dt):
                    start_date = row[4]
                else:
                    start_date = dt.strptime(start_date_str, "%Y-%m-%d")
                
                if end_date_str:
                    if isinstance(row[5], dt):
                        end_date = row[5]
                    else:
                        end_date = dt.strptime(end_date_str, "%Y-%m-%d")
                else:
                    end_date = start_date
            except ValueError as e:
                results["errors"].append({
                    "row": row_idx,
                    "error": f"Tarih formatı hatalı: {str(e)}"
                })
                results["error_count"] += 1
                continue
            
            # Vardiya tipi ve saatleri belirle
            if shift_type == "ofis_8":
                start_time = "08:00"
                end_time = "17:00"
            else:  # saha_24
                start_time = "08:00"
                end_time = "08:00"
                # 24 saat vardiya için bitiş tarihi bir gün sonra
                if end_date == start_date:
                    end_date = start_date + timedelta(days=1)
            
            # Atama oluştur
            assignment = ShiftAssignment(
                user_id=user_doc["_id"],
                vehicle_id=vehicle_id,
                location_type=location_type if location_type in ["arac", "saglik_merkezi"] else "arac",
                health_center_name=location_name if location_type == "saglik_merkezi" else None,
                assigned_by=current_user.id,
                shift_date=start_date,
                start_time=start_time,
                end_time=end_time,
                end_date=end_date,
                shift_type=shift_type if shift_type in ["saha_24", "ofis_8"] else "saha_24"
            )
            
            assignment_dict = assignment.model_dump(by_alias=True)
            await shift_assignments_collection.insert_one(assignment_dict)
            
            results["success"].append({
                "row": row_idx,
                "user": user_doc.get("name"),
                "vehicle": vehicle_plate,
                "date": start_date.strftime("%Y-%m-%d")
            })
            results["successful_count"] += 1
            
        except Exception as e:
            results["errors"].append({
                "row": row_idx,
                "error": str(e)
            })
            results["error_count"] += 1
    
    logger.info(f"Toplu vardiya atama: {results['successful_count']} başarılı, {results['error_count']} hatalı")
    
    return results


@router.get("/bulk-upload/template")
async def get_bulk_upload_template(request: Request):
    """Excel şablon dosyasını indir"""
    await require_roles(["merkez_ofis", "operasyon_muduru", "bas_sofor", "mesul_mudur"])(request)
    
    try:
        import openpyxl
        from fastapi.responses import StreamingResponse
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl kütüphanesi yüklü değil")
    
    # Yeni workbook oluştur
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vardiya Atama"
    
    # Başlıklar
    headers = [
        "TC/Email",
        "Araç Plaka",
        "Lokasyon Tipi (arac/saglik_merkezi)",
        "Lokasyon Adı (Sağlık Merkezi ise)",
        "Başlangıç Tarihi (YYYY-MM-DD)",
        "Bitiş Tarihi (YYYY-MM-DD)",
        "Vardiya Tipi (saha_24/ofis_8)"
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
    
    # Örnek satırlar
    examples = [
        ["12345678901", "34 ABC 123", "arac", "", "2025-01-15", "2025-01-16", "saha_24"],
        ["email@test.com", "", "saglik_merkezi", "Filyos SM", "2025-01-15", "2025-01-15", "ofis_8"]
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Dosyayı BytesIO'ya kaydet
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=vardiya_atama_sablonu.xlsx"}
    )


# ============================================================================
# DEVİR TESLİM OTURUMU
# ============================================================================

from database import db
handover_sessions_collection = db.handover_sessions

@router.post("/handover/start")
async def start_handover_session(request: Request):
    """
    Devir teslim oturumu başlat (Devreden şoför)
    Sadece 07:30-08:30 arasında başlatılabilir
    """
    from models import HandoverSession
    from datetime import time
    
    user = await get_current_user(request)
    body = await request.json()
    
    # Aktif vardiya kontrolü
    active_shift = await shifts_collection.find_one({
        "user_id": user.id,
        "status": "active"
    })
    
    if not active_shift:
        raise HTTPException(status_code=400, detail="Aktif vardiyanz yok. Önce vardiya başlatın.")
    
    # Saat kontrolü (07:30-08:30 arası) - Türkiye saati
    turkey_now = datetime.utcnow() + timedelta(hours=3)
    current_time = turkey_now.time()
    handover_start = time(7, 30)
    handover_end = time(8, 30)
    
    # Debug için saati logla
    logger.info(f"Devir teslim saat kontrolü - Şu an: {current_time}, Aralık: {handover_start}-{handover_end}")
    
    # Test modunda saat kontrolünü atla
    skip_time_check = body.get("skip_time_check", False)
    
    if not skip_time_check and not (handover_start <= current_time <= handover_end):
        raise HTTPException(
            status_code=400, 
            detail=f"Devir teslim sadece 07:30-08:30 arasında yapılabilir. Şu an: {current_time.strftime('%H:%M')}"
        )
    
    vehicle_id = active_shift.get("vehicle_id")
    if not vehicle_id:
        raise HTTPException(status_code=400, detail="Vardiyaya araç atanmamış")
    
    # Araç bilgisini al
    vehicle = await vehicles_collection.find_one({"_id": vehicle_id})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Araç bulunamadı")
    
    # Sonraki vardiyalı kişiyi bul
    today = turkey_now.date()
    tomorrow = today + timedelta(days=1)
    
    next_assignments = await shift_assignments_collection.find({
        "vehicle_id": vehicle_id,
        "user_id": {"$ne": user.id},
        "status": "pending"
    }).to_list(100)
    
    # Bugün veya yarın için geçerli atama
    receiver_assignment = None
    for assignment in next_assignments:
        shift_date = assignment.get("shift_date")
        if isinstance(shift_date, str):
            shift_date = datetime.fromisoformat(shift_date.replace('Z', '+00:00'))
        if isinstance(shift_date, datetime):
            assignment_date = shift_date.date()
            if assignment_date == today or assignment_date == tomorrow:
                receiver_assignment = assignment
                break
    
    if not receiver_assignment:
        raise HTTPException(status_code=400, detail="Bu araç için sonraki vardiya ataması bulunamadı")
    
    # Devralan kullanıcı bilgisi
    receiver_user = await users_collection.find_one({"_id": receiver_assignment.get("user_id")})
    if not receiver_user:
        raise HTTPException(status_code=400, detail="Devralan kullanıcı bulunamadı")
    
    # Mevcut bekleyen oturum var mı kontrol et
    existing_session = await handover_sessions_collection.find_one({
        "vehicle_id": vehicle_id,
        "status": {"$in": ["waiting_receiver", "waiting_manager"]}
    })
    
    if existing_session:
        # Varolan oturumu döndür
        existing_session["id"] = existing_session.pop("_id")
        return existing_session
    
    # Yeni oturum oluştur
    session = HandoverSession(
        giver_id=user.id,
        giver_name=user.name,
        receiver_id=receiver_user["_id"],
        receiver_name=receiver_user.get("name", ""),
        vehicle_id=vehicle_id,
        vehicle_plate=vehicle.get("plate", ""),
        giver_shift_id=active_shift["_id"],
        expires_at=turkey_now + timedelta(hours=2)  # 2 saat geçerli
    )
    
    session_dict = session.model_dump(by_alias=True)
    await handover_sessions_collection.insert_one(session_dict)
    
    logger.info(f"Devir teslim oturumu başlatıldı: {session.id} - {user.name} -> {receiver_user.get('name')}")
    
    session_dict["id"] = session_dict.pop("_id")
    return session_dict


@router.get("/handover/active")
async def get_active_handover(request: Request):
    """
    Kullanıcının aktif devir teslim oturumunu getir
    (Hem devreden hem devralan için)
    """
    user = await get_current_user(request)
    
    turkey_now = datetime.utcnow() + timedelta(hours=3)
    
    # Kullanıcının devreden veya devralan olduğu aktif oturum
    session = await handover_sessions_collection.find_one({
        "$or": [
            {"giver_id": user.id},
            {"receiver_id": user.id}
        ],
        "status": {"$in": ["waiting_receiver", "waiting_manager"]},
        "expires_at": {"$gt": turkey_now}
    })
    
    if not session:
        return None
    
    session["id"] = session.pop("_id")
    
    # Kullanıcının rolünü ekle
    session["user_role_in_session"] = "giver" if session["giver_id"] == user.id else "receiver"
    
    return session


@router.post("/handover/{session_id}/sign")
async def sign_handover(session_id: str, request: Request):
    """
    Devir teslim oturumunu imzala (Devralan şoför)
    İmza veya OTP ile onay
    """
    user = await get_current_user(request)
    body = await request.json()
    
    session = await handover_sessions_collection.find_one({"_id": session_id})
    if not session:
        raise HTTPException(status_code=404, detail="Oturum bulunamadı")
    
    if session["receiver_id"] != user.id:
        raise HTTPException(status_code=403, detail="Bu oturumu sadece devralan onaylayabilir")
    
    if session["status"] != "waiting_receiver":
        raise HTTPException(status_code=400, detail="Bu oturum zaten onaylanmış veya süresi dolmuş")
    
    signature = body.get("signature")  # Base64 imza
    otp_code = body.get("otp_code")
    
    turkey_now = datetime.utcnow() + timedelta(hours=3)
    
    update_data = {
        "receiver_signed_at": turkey_now,
        "status": "waiting_manager",
        "receiver_login_at": turkey_now,  # Sisteme giriş zamanı
        "updated_at": turkey_now
    }
    
    if signature:
        update_data["receiver_signature"] = signature
    
    if otp_code:
        # OTP doğrulama
        from services.otp_service import verify_user_otp
        
        otp_secret = user.otp_secret if hasattr(user, 'otp_secret') else None
        if not otp_secret:
            user_doc = await users_collection.find_one({"_id": user.id})
            otp_secret = user_doc.get("otp_secret") if user_doc else None
        
        if otp_secret and verify_user_otp(otp_secret, otp_code):
            update_data["receiver_otp_verified"] = True
        else:
            raise HTTPException(status_code=400, detail="Geçersiz OTP kodu")
    
    await handover_sessions_collection.update_one(
        {"_id": session_id},
        {"$set": update_data}
    )
    
    logger.info(f"Devir teslim imzalandı: {session_id} - Devralan: {user.name}")
    
    # Baş şoförlere bildirim gönder
    try:
        from services.onesignal_service import onesignal_service, NotificationType
        
        managers = await users_collection.find({
            "role": {"$in": ["bas_sofor", "operasyon_muduru"]},
            "is_active": True
        }).to_list(50)
        
        manager_ids = [m["_id"] for m in managers]
        
        if manager_ids:
            await onesignal_service.send_notification(
                NotificationType.SYSTEM_ALERT,
                manager_ids,
                {
                    "message": f"Devir teslim onayı bekliyor: {session.get('giver_name')} → {user.name} ({session.get('vehicle_plate')})"
                }
            )
    except Exception as e:
        logger.warning(f"Bildirim gönderilemedi: {e}")
    
    return {"message": "Devir teslim imzalandı. Yönetici onayı bekleniyor."}


@router.post("/handover/{session_id}/approve")
async def approve_handover(session_id: str, request: Request):
    """
    Devir teslim oturumunu onayla (Baş Şoför/Yönetici)
    """
    user = await require_roles(["bas_sofor", "operasyon_muduru", "merkez_ofis"])(request)
    body = await request.json()
    
    session = await handover_sessions_collection.find_one({"_id": session_id})
    if not session:
        raise HTTPException(status_code=404, detail="Oturum bulunamadı")
    
    if session["status"] != "waiting_manager":
        raise HTTPException(status_code=400, detail="Bu oturum yönetici onayı beklemeiyor")
    
    turkey_now = datetime.utcnow() + timedelta(hours=3)
    
    approve = body.get("approve", True)
    rejection_reason = body.get("rejection_reason")
    
    if approve:
        # Onay
        await handover_sessions_collection.update_one(
            {"_id": session_id},
            {"$set": {
                "status": "approved",
                "manager_id": user.id,
                "manager_name": user.name,
                "manager_action_at": turkey_now,
                "updated_at": turkey_now
            }}
        )
        
        # Devreden şoförün vardiyasını bitir
        if session.get("giver_shift_id"):
            giver_shift = await shifts_collection.find_one({"_id": session["giver_shift_id"]})
            if giver_shift:
                start_time = giver_shift.get("start_time")
                duration = int((turkey_now - start_time).total_seconds() / 60) if start_time else 0
                
                await shifts_collection.update_one(
                    {"_id": session["giver_shift_id"]},
                    {"$set": {
                        "status": "completed",
                        "end_time": turkey_now,
                        "duration_minutes": duration,
                        "handover_session_id": session_id
                    }}
                )
        
        # Devreden kişinin assignment'ını tamamla
        if giver_shift and giver_shift.get("assignment_id"):
            await shift_assignments_collection.update_one(
                {"_id": giver_shift["assignment_id"]},
                {"$set": {"status": "completed"}}
            )
        
        logger.info(f"Devir teslim onaylandı: {session_id} - Onaylayan: {user.name}")
        
        return {"message": "Devir teslim onaylandı. Devralan artık vardiyasını başlatabilir."}
    else:
        # Red
        await handover_sessions_collection.update_one(
            {"_id": session_id},
            {"$set": {
                "status": "rejected",
                "manager_id": user.id,
                "manager_name": user.name,
                "manager_action_at": turkey_now,
                "rejection_reason": rejection_reason,
                "updated_at": turkey_now
            }}
        )
        
        logger.info(f"Devir teslim reddedildi: {session_id} - Reddeden: {user.name}")
        
        return {"message": "Devir teslim reddedildi."}


@router.get("/handover/pending-approvals")
async def get_pending_handover_approvals(request: Request, date: Optional[str] = None):
    """
    Bekleyen devir teslim onaylarını getir (Baş Şoför için)
    """
    user = await require_roles(["bas_sofor", "operasyon_muduru", "merkez_ofis", "cagri_merkezi", "hemsire"])(request)
    
    turkey_now = datetime.utcnow() + timedelta(hours=3)
    
    query = {"status": "waiting_manager"}
    
    if date:
        try:
            target_date = datetime.strptime(date, "%Y-%m-%d")
            query["form_opened_at"] = {
                "$gte": target_date,
                "$lt": target_date + timedelta(days=1)
            }
        except ValueError:
            pass
    
    sessions = await handover_sessions_collection.find(query).sort("form_opened_at", -1).to_list(100)
    
    for session in sessions:
        session["id"] = session.pop("_id")
    
    return sessions


@router.get("/handover/logs")
async def get_handover_logs(request: Request, date: Optional[str] = None, vehicle_id: Optional[str] = None):
    """
    Devir teslim loglarını getir (Baş Şoför için)
    Tarih ve araç bazlı filtreleme
    """
    user = await require_roles(["bas_sofor", "operasyon_muduru", "merkez_ofis", "cagri_merkezi", "hemsire"])(request)
    
    turkey_now = datetime.utcnow() + timedelta(hours=3)
    
    query = {}
    
    if date:
        try:
            target_date = datetime.strptime(date, "%Y-%m-%d")
            query["form_opened_at"] = {
                "$gte": target_date,
                "$lt": target_date + timedelta(days=1)
            }
        except ValueError:
            # Bugün
            today = turkey_now.date()
            query["form_opened_at"] = {
                "$gte": datetime.combine(today, datetime.min.time()),
                "$lt": datetime.combine(today + timedelta(days=1), datetime.min.time())
            }
    else:
        # Bugün
        today = turkey_now.date()
        query["form_opened_at"] = {
            "$gte": datetime.combine(today, datetime.min.time()),
            "$lt": datetime.combine(today + timedelta(days=1), datetime.min.time())
        }
    
    if vehicle_id:
        query["vehicle_id"] = vehicle_id
    
    sessions = await handover_sessions_collection.find(query).sort("form_opened_at", -1).to_list(100)
    
    result = []
    for session in sessions:
        session["id"] = session.pop("_id")
        
        # Log detayları
        log = {
            "session": session,
            "logs": []
        }
        
        # Form açılış
        if session.get("form_opened_at"):
            log["logs"].append({
                "event": "form_opened",
                "time": session["form_opened_at"],
                "user": session.get("giver_name"),
                "description": "Devir teslim formu açıldı"
            })
        
        # Devralan giriş
        if session.get("receiver_login_at"):
            log["logs"].append({
                "event": "receiver_login",
                "time": session["receiver_login_at"],
                "user": session.get("receiver_name"),
                "description": "Devralan sisteme giriş yaptı"
            })
        
        # İmza
        if session.get("receiver_signed_at"):
            log["logs"].append({
                "event": "receiver_signed",
                "time": session["receiver_signed_at"],
                "user": session.get("receiver_name"),
                "description": "Devralan imzaladı" + (" (OTP)" if session.get("receiver_otp_verified") else " (İmza)")
            })
        
        # Yönetici işlemi
        if session.get("manager_action_at"):
            log["logs"].append({
                "event": "manager_action",
                "time": session["manager_action_at"],
                "user": session.get("manager_name"),
                "description": f"Yönetici {'onayladı' if session.get('status') == 'approved' else 'reddetti'}"
            })
        
        # Vardiya başlatma
        if session.get("shift_started_at"):
            log["logs"].append({
                "event": "shift_started",
                "time": session["shift_started_at"],
                "user": session.get("receiver_name"),
                "description": "Vardiya başlatıldı"
            })
        
        result.append(log)
    
    return result


# ============================================================================
# EKİP GRUPLAMA
# ============================================================================

@router.get("/today-team/{vehicle_id}")
async def get_today_team(vehicle_id: str, request: Request):
    """
    Bugün bu araçta/lokasyonda çalışacak ekibi getir
    Şoför, ATT, Paramedik otomatik gruplanır
    """
    user = await get_current_user(request)
    
    turkey_now = datetime.utcnow() + timedelta(hours=3)
    today = turkey_now.date()
    
    # Bu araç için bugünkü atamaları al
    assignments = await shift_assignments_collection.find({
        "vehicle_id": vehicle_id,
        "status": {"$in": ["pending", "started"]}
    }).to_list(100)
    
    # Bugün için geçerli atamaları filtrele
    today_assignments = []
    for assignment in assignments:
        shift_date = assignment.get("shift_date")
        end_date = assignment.get("end_date") or shift_date
        
        if isinstance(shift_date, str):
            shift_date = datetime.fromisoformat(shift_date.replace('Z', '+00:00'))
        if isinstance(end_date, str):
            end_date = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        
        if isinstance(shift_date, datetime) and isinstance(end_date, datetime):
            if shift_date.date() <= today <= end_date.date():
                today_assignments.append(assignment)
    
    # Ekip üyelerini topla
    team = []
    for assignment in today_assignments:
        user_doc = await users_collection.find_one({"_id": assignment.get("user_id")})
        if user_doc:
            team.append({
                "user_id": user_doc["_id"],
                "name": user_doc.get("name"),
                "role": user_doc.get("role"),
                "phone": user_doc.get("phone"),
                "profile_photo": user_doc.get("profile_photo"),
                "assignment_id": assignment["_id"],
                "status": assignment.get("status")
            })
    
    # Araç bilgisi
    vehicle = await vehicles_collection.find_one({"_id": vehicle_id})
    
    return {
        "vehicle_id": vehicle_id,
        "vehicle_plate": vehicle.get("plate") if vehicle else None,
        "date": today.isoformat(),
        "team": team,
        "team_count": len(team)
    }


# ============================================================================
# VARDİYA BAŞLATMA ONAY SİSTEMİ (ATT/Paramedik ve Şoför için ayrı)
# ============================================================================

from database import db
shift_start_approvals_collection = db.shift_start_approvals


class ShiftStartApprovalCreate(BaseModel):
    vehicle_id: str
    role_type: str  # "medical" (ATT/Paramedik) veya "driver" (Şoför)
    daily_control_data: Optional[dict] = None
    photos: Optional[dict] = None


@router.post("/start-approval/request")
async def request_shift_start_approval(data: ShiftStartApprovalCreate, request: Request):
    """
    Vardiya başlatma onayı iste (ATT/Paramedik veya Şoför)
    Onay shift-approvals sayfasına düşer
    """
    user = await get_current_user(request)
    turkey_now = datetime.utcnow() + timedelta(hours=3)
    
    # Araç bilgisini al
    vehicle = await vehicles_collection.find_one({"_id": data.vehicle_id})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Araç bulunamadı")
    
    # Aynı gün için bekleyen onay var mı?
    today_start = turkey_now.replace(hour=0, minute=0, second=0, microsecond=0)
    existing = await shift_start_approvals_collection.find_one({
        "user_id": user.id,
        "vehicle_id": data.vehicle_id,
        "status": "pending",
        "created_at": {"$gte": today_start}
    })
    
    if existing:
        return {"id": existing["_id"], "message": "Zaten bekleyen onay mevcut", "status": "pending"}
    
    approval_id = str(uuid.uuid4())
    approval = {
        "_id": approval_id,
        "user_id": user.id,
        "user_name": user.name,
        "user_role": user.role,
        "vehicle_id": data.vehicle_id,
        "vehicle_plate": vehicle.get("plate"),
        "role_type": data.role_type,  # "medical" veya "driver"
        "daily_control_data": data.daily_control_data,
        "photos": data.photos,
        "status": "pending",  # pending, approved, rejected
        "created_at": turkey_now,
        "approved_by": None,
        "approved_by_name": None,
        "approved_at": None,
        "rejection_reason": None
    }
    
    await shift_start_approvals_collection.insert_one(approval)
    
    logger.info(f"Vardiya başlatma onayı istendi: {user.name} - {vehicle.get('plate')} - {data.role_type}")
    
    return {"id": approval_id, "message": "Onay talebi oluşturuldu", "status": "pending"}


@router.get("/start-approval/pending")
async def get_pending_shift_start_approvals(request: Request, role_type: Optional[str] = None):
    """
    Bekleyen vardiya başlatma onaylarını getir
    role_type: "medical" (ATT/Paramedik) veya "driver" (Şoför) veya None (tümü)
    """
    user = await require_roles(["bas_sofor", "operasyon_muduru", "merkez_ofis", "mesul_mudur"])(request)
    
    query = {"status": "pending"}
    
    if role_type:
        query["role_type"] = role_type
    
    approvals = await shift_start_approvals_collection.find(query).sort("created_at", -1).to_list(100)
    
    # Onay kayıtlarını düzenle
    for approval in approvals:
        approval["id"] = approval.pop("_id")
        
        # Onay kaydında zaten photos ve daily_control_data var!
        # Ama frontend "daily_control" bekliyor, "daily_control_data" değil
        if "daily_control_data" in approval:
            approval["daily_control"] = approval["daily_control_data"]
        
        # Eğer onay kaydında photos yoksa, shift_photos collection'dan al
        shift_id = approval.get("shift_id")
        if shift_id and not approval.get("photos"):
            shift_photos_collection = db["shift_photos"]
            photo_doc = await shift_photos_collection.find_one({"shift_id": shift_id})
            if photo_doc:
                approval["photos"] = photo_doc.get("photos")
        
        # Eğer onay kaydında daily_control yoksa, forms collection'dan al
        if shift_id and not approval.get("daily_control"):
            from database import forms_collection
            daily_control_form = await forms_collection.find_one({
                "shift_id": shift_id,
                "form_type": "daily_control"
            })
            if daily_control_form:
                approval["daily_control"] = daily_control_form.get("form_data")
    
    return approvals


@router.get("/start-approval/check/{approval_id}")
async def check_shift_start_approval(approval_id: str, request: Request):
    """
    Belirli bir onay talebinin durumunu kontrol et
    """
    user = await get_current_user(request)
    
    approval = await shift_start_approvals_collection.find_one({"_id": approval_id})
    
    if not approval:
        raise HTTPException(status_code=404, detail="Onay talebi bulunamadı")
    
    return {
        "id": approval["_id"],
        "status": approval["status"],
        "approved_by_name": approval.get("approved_by_name"),
        "approved_at": approval.get("approved_at"),
        "rejection_reason": approval.get("rejection_reason")
    }


@router.post("/start-approval/{approval_id}/approve")
async def approve_shift_start(approval_id: str, request: Request):
    """
    Vardiya başlatma onayı ver
    """
    user = await require_roles(["bas_sofor", "operasyon_muduru", "merkez_ofis", "mesul_mudur"])(request)
    turkey_now = datetime.utcnow() + timedelta(hours=3)
    
    approval = await shift_start_approvals_collection.find_one({"_id": approval_id})
    
    if not approval:
        raise HTTPException(status_code=404, detail="Onay talebi bulunamadı")
    
    if approval["status"] != "pending":
        raise HTTPException(status_code=400, detail="Bu talep zaten işlenmiş")
    
    await shift_start_approvals_collection.update_one(
        {"_id": approval_id},
        {"$set": {
            "status": "approved",
            "approved_by": user.id,
            "approved_by_name": user.name,
            "approved_at": turkey_now
        }}
    )
    
    logger.info(f"Vardiya başlatma onaylandı: {approval['user_name']} - {approval['vehicle_plate']} - Onaylayan: {user.name}")
    
    return {"message": "Vardiya başlatma onaylandı", "status": "approved"}


@router.post("/start-approval/{approval_id}/reject")
async def reject_shift_start(approval_id: str, request: Request, reason: str = None):
    """
    Vardiya başlatma onayını reddet
    """
    user = await require_roles(["bas_sofor", "operasyon_muduru", "merkez_ofis", "mesul_mudur"])(request)
    turkey_now = datetime.utcnow() + timedelta(hours=3)
    
    approval = await shift_start_approvals_collection.find_one({"_id": approval_id})
    
    if not approval:
        raise HTTPException(status_code=404, detail="Onay talebi bulunamadı")
    
    if approval["status"] != "pending":
        raise HTTPException(status_code=400, detail="Bu talep zaten işlenmiş")
    
    await shift_start_approvals_collection.update_one(
        {"_id": approval_id},
        {"$set": {
            "status": "rejected",
            "approved_by": user.id,
            "approved_by_name": user.name,
            "approved_at": turkey_now,
            "rejection_reason": reason or "Belirtilmedi"
        }}
    )
    
    logger.info(f"Vardiya başlatma reddedildi: {approval['user_name']} - {approval['vehicle_plate']} - Reddeden: {user.name}")
    
    return {"message": "Vardiya başlatma reddedildi", "status": "rejected"}


# ===================== VARDİYA BİTİRME ONAY SİSTEMİ =====================

# Vardiya Bitirme Onayları Collection
shift_end_approvals_collection = db["shift_end_approvals"]


class EndApprovalRequest(BaseModel):
    shift_id: str
    vehicle_id: Optional[str] = None
    vehicle_plate: Optional[str] = None
    end_km: Optional[str] = None
    devralan_adi: Optional[str] = None
    form_opened_at: Optional[str] = None
    request_sent_at: Optional[str] = None


@router.post("/end-approval/request")
async def request_end_approval(data: EndApprovalRequest, request: Request):
    """
    Vardiya bitirme için yönetici onayı iste
    """
    user = await get_current_user(request)
    turkey_now = datetime.utcnow() + timedelta(hours=3)
    
    # Rol tipini belirle
    role_type = "driver" if user.role.lower() == "sofor" else "medical"
    
    approval_id = str(uuid.uuid4())
    approval_doc = {
        "_id": approval_id,
        "type": "end",
        "shift_id": data.shift_id,
        "vehicle_id": data.vehicle_id,
        "vehicle_plate": data.vehicle_plate,
        "user_id": user.id,
        "user_name": user.name,
        "user_role": user.role,
        "role_type": role_type,
        "end_km": data.end_km,
        "devralan_adi": data.devralan_adi,
        "form_opened_at": data.form_opened_at,
        "request_sent_at": data.request_sent_at or turkey_now.isoformat(),
        "status": "pending",
        "created_at": turkey_now
    }
    
    await shift_end_approvals_collection.insert_one(approval_doc)
    
    logger.info(f"Vardiya bitirme onayı istendi: {user.name} ({user.role}) - {data.vehicle_plate}")
    
    return {
        "message": "Vardiya bitirme onay talebi gönderildi",
        "request_id": approval_id
    }


@router.get("/shift-approvals/pending")
async def get_pending_shift_approvals(request: Request):
    """
    Tüm bekleyen vardiya onaylarını getir (başlatma + bitirme)
    """
    await require_roles(["bas_sofor", "operasyon_muduru", "merkez_ofis", "mesul_mudur"])(request)
    
    # Başlatma onayları
    start_approvals = await shift_start_approvals_collection.find({"status": "pending"}).sort("created_at", -1).to_list(100)
    for a in start_approvals:
        a["id"] = a.pop("_id")
        a["type"] = "start"
        
        # Onay kaydında zaten photos ve daily_control_data var
        if "daily_control_data" in a and a["daily_control_data"]:
            a["daily_control"] = a["daily_control_data"]
    
    # Bitirme onayları
    end_approvals = await shift_end_approvals_collection.find({"status": "pending"}).sort("created_at", -1).to_list(100)
    for a in end_approvals:
        a["id"] = a.pop("_id")
        a["type"] = "end"
        
        # Bitirme onayında shift_id var, shift'ten verileri alalım
        shift_id = a.get("shift_id")
        if shift_id:
            shift = await shifts_collection.find_one({"_id": shift_id})
            if shift:
                a["end_photos"] = shift.get("end_photos")  # Bitiş fotoğrafları (4 köşe)
                a["end_signature"] = shift.get("end_signature")  # Bitiş imzası
                
                # Başlatma fotoğraflarını shift_photos collection'dan al
                shift_photos_collection = db["shift_photos"]
                photo_doc = await shift_photos_collection.find_one({"shift_id": shift_id})
                if photo_doc:
                    a["photos"] = photo_doc.get("photos")
                
                # Günlük kontrol formunu forms collection'dan al
                from database import forms_collection
                daily_control_form = await forms_collection.find_one({
                    "shift_id": shift_id,
                    "form_type": "daily_control"
                })
                if daily_control_form:
                    a["daily_control"] = daily_control_form.get("form_data")
    
    return start_approvals + end_approvals


@router.post("/shift-approvals/{approval_id}/approve")
async def approve_shift_approval(approval_id: str, request: Request):
    """
    Vardiya onayı ver (başlatma veya bitirme)
    """
    user = await require_roles(["bas_sofor", "operasyon_muduru", "merkez_ofis", "mesul_mudur"])(request)
    turkey_now = datetime.utcnow() + timedelta(hours=3)
    
    # Önce başlatma onaylarına bak
    approval = await shift_start_approvals_collection.find_one({"_id": approval_id})
    collection = shift_start_approvals_collection
    approval_type = "başlatma"
    
    if not approval:
        # Bitirme onaylarına bak
        approval = await shift_end_approvals_collection.find_one({"_id": approval_id})
        collection = shift_end_approvals_collection
        approval_type = "bitirme"
    
    if not approval:
        raise HTTPException(status_code=404, detail="Onay talebi bulunamadı")
    
    if approval["status"] != "pending":
        raise HTTPException(status_code=400, detail="Bu talep zaten işlenmiş")
    
    await collection.update_one(
        {"_id": approval_id},
        {"$set": {
            "status": "approved",
            "approved_by": user.id,
            "approved_by_name": user.name,
            "approved_at": turkey_now
        }}
    )
    
    logger.info(f"Vardiya {approval_type} onaylandı: {approval['user_name']} - {approval.get('vehicle_plate', 'N/A')} - Onaylayan: {user.name}")
    
    return {"message": f"Vardiya {approval_type} onaylandı", "status": "approved"}


@router.post("/shift-approvals/{approval_id}/reject")
async def reject_shift_approval(approval_id: str, request: Request, reason: str = None):
    """
    Vardiya onayını reddet (başlatma veya bitirme)
    """
    user = await require_roles(["bas_sofor", "operasyon_muduru", "merkez_ofis", "mesul_mudur"])(request)
    turkey_now = datetime.utcnow() + timedelta(hours=3)
    
    # Önce başlatma onaylarına bak
    approval = await shift_start_approvals_collection.find_one({"_id": approval_id})
    collection = shift_start_approvals_collection
    approval_type = "başlatma"
    
    if not approval:
        # Bitirme onaylarına bak
        approval = await shift_end_approvals_collection.find_one({"_id": approval_id})
        collection = shift_end_approvals_collection
        approval_type = "bitirme"
    
    if not approval:
        raise HTTPException(status_code=404, detail="Onay talebi bulunamadı")
    
    if approval["status"] != "pending":
        raise HTTPException(status_code=400, detail="Bu talep zaten işlenmiş")
    
    await collection.update_one(
        {"_id": approval_id},
        {"$set": {
            "status": "rejected",
            "approved_by": user.id,
            "approved_by_name": user.name,
            "approved_at": turkey_now,
            "rejection_reason": reason or "Belirtilmedi"
        }}
    )
    
    logger.info(f"Vardiya {approval_type} reddedildi: {approval['user_name']} - {approval.get('vehicle_plate', 'N/A')} - Reddeden: {user.name}")
    
    return {"message": f"Vardiya {approval_type} reddedildi", "status": "rejected"}


@router.get("/shift-approvals/logs")
async def get_shift_approval_logs(
    request: Request,
    date: str = None,
    user_id: str = None,
    vehicle_id: str = None,
    limit: int = 100
):
    """
    Vardiya onay loglarını getir (başlatma + bitirme)
    Kişi ve araç bazında filtreleme yapılabilir
    """
    await require_roles(["bas_sofor", "operasyon_muduru", "merkez_ofis", "mesul_mudur"])(request)
    
    query = {}
    if date:
        try:
            target_date = datetime.strptime(date, "%Y-%m-%d")
            query["created_at"] = {
                "$gte": target_date,
                "$lt": target_date + timedelta(days=1)
            }
        except ValueError:
            pass
    
    if user_id:
        query["user_id"] = user_id
    
    if vehicle_id:
        query["vehicle_id"] = vehicle_id
    
    # Başlatma logları
    start_logs = await shift_start_approvals_collection.find(query).sort("created_at", -1).limit(limit).to_list(limit)
    for log in start_logs:
        log["id"] = log.pop("_id")
        log["type"] = "start"
        log["type_label"] = "Vardiya Başlatma"
        # Vardiya başlatılma zamanı (shift_started_at varsa onu kullan, yoksa created_at)
        log["action_time"] = log.get("shift_started_at") or log.get("created_at")
    
    # Bitirme logları
    end_logs = await shift_end_approvals_collection.find(query).sort("created_at", -1).limit(limit).to_list(limit)
    for log in end_logs:
        log["id"] = log.pop("_id")
        log["type"] = "end"
        log["type_label"] = "Vardiya Bitirme"
        # Vardiya bitirilme zamanı (shift_ended_at varsa onu kullan, yoksa created_at)
        log["action_time"] = log.get("shift_ended_at") or log.get("created_at")
    
    # Birleştir ve sırala (action_time'e göre)
    all_logs = start_logs + end_logs
    all_logs.sort(key=lambda x: x.get("action_time", x.get("created_at", datetime.min)), reverse=True)
    
    return all_logs[:limit]


@router.get("/shift-logs")
async def get_shift_logs(
    request: Request,
    date: str = None,
    action_type: str = None,
    limit: int = 100
):
    """
    Vardiya işlem loglarını getir (form açılma, başlatma, bitirme)
    """
    await require_roles(["bas_sofor", "operasyon_muduru", "merkez_ofis", "mesul_mudur"])(request)
    
    shift_logs_collection = db["shift_logs"]
    
    query = {}
    if date:
        try:
            target_date = datetime.strptime(date, "%Y-%m-%d")
            query["logged_at"] = {
                "$gte": target_date,
                "$lt": target_date + timedelta(days=1)
            }
        except ValueError:
            pass
    
    if action_type:
        query["action_type"] = action_type
    
    logs = await shift_logs_collection.find(query).sort("logged_at", -1).limit(limit).to_list(limit)
    
    # ID'leri düzelt ve formatla
    for log in logs:
        log["id"] = str(log.pop("_id"))
        # Tarih formatla
        if log.get("logged_at"):
            log["logged_at_formatted"] = log["logged_at"].strftime("%d.%m.%Y %H:%M:%S") if isinstance(log["logged_at"], datetime) else str(log["logged_at"])
        if log.get("form_opened_at"):
            log["form_opened_at_formatted"] = log["form_opened_at"].strftime("%d.%m.%Y %H:%M:%S") if isinstance(log["form_opened_at"], datetime) else str(log["form_opened_at"])
        if log.get("action_taken_at"):
            log["action_taken_at_formatted"] = log["action_taken_at"].strftime("%d.%m.%Y %H:%M:%S") if isinstance(log["action_taken_at"], datetime) else str(log["action_taken_at"])
    
    return logs


# ============================================================================
# OTOMATIK VARDIYA BAŞLATMA - SAĞLIK MERKEZİ ÇALIŞANLARI İÇİN
# ============================================================================

async def auto_start_health_center_shifts():
    """
    Sağlık merkezinde çalışanlar için otomatik vardiya başlatma
    Çalışma saati geldiğinde pending olan vardiyaları otomatik başlatır
    Sadece sağlık merkezi lokasyonunda görevli olanlar için (location_type="saglik_merkezi")
    """
    try:
        turkey_now = get_turkey_time()
        today = turkey_now.date()
        current_time = turkey_now.time()
        
        logger.info(f"[AUTO_SHIFT_START] Kontrol başlatıldı - Tarih: {today}, Saat: {current_time}")
        
        # Bugün için pending olan ve sağlık merkezi atamalarını bul
        today_start = datetime.combine(today, datetime.min.time())
        today_end = datetime.combine(today + timedelta(days=1), datetime.min.time())
        
        assignments = await shift_assignments_collection.find({
            "shift_date": {
                "$gte": today_start,
                "$lt": today_end
            },
            "status": "pending",
            "location_type": "saglik_merkezi"
        }).to_list(1000)
        
        if not assignments:
            logger.debug(f"[AUTO_SHIFT_START] Başlatılacak vardiya bulunamadı")
            return
        
        logger.info(f"[AUTO_SHIFT_START] {len(assignments)} adet pending sağlık merkezi ataması bulundu")
        
        started_count = 0
        skipped_count = 0
        
        for assignment in assignments:
            try:
                # Kullanıcı bilgisini al ve rolünü kontrol et
                user_id = assignment.get("user_id")
                user = await users_collection.find_one({"_id": user_id})
                if not user:
                    logger.warning(f"[AUTO_SHIFT_START] Kullanıcı {user_id} bulunamadı, atlanıyor")
                    skipped_count += 1
                    continue
                
                user_role = user.get("role")
                # Sadece sağlık merkezi çalışanları için otomatik başlat (hemsire, cagri_merkezi, bas_sofor, doktor, operasyon_muduru)
                health_center_roles = ["hemsire", "cagri_merkezi", "bas_sofor", "doktor", "operasyon_muduru", "merkez_ofis"]
                if user_role not in health_center_roles:
                    logger.debug(f"[AUTO_SHIFT_START] Kullanıcı {user.get('name')} ({user_role}) sağlık merkezi çalışanı değil, atlanıyor")
                    skipped_count += 1
                    continue
                
                # start_time kontrolü
                start_time_str = assignment.get("start_time")
                if not start_time_str:
                    logger.debug(f"[AUTO_SHIFT_START] Atama {assignment['_id']} için start_time yok, atlanıyor")
                    skipped_count += 1
                    continue
                
                # start_time'ı parse et (HH:MM format)
                try:
                    start_hour, start_minute = map(int, start_time_str.split(':'))
                    from datetime import time as dt_time
                    shift_start_time = dt_time(hour=start_hour, minute=start_minute)
                except (ValueError, AttributeError) as e:
                    logger.warning(f"[AUTO_SHIFT_START] Atama {assignment['_id']} için geçersiz start_time: {start_time_str}, hata: {e}")
                    skipped_count += 1
                    continue
                
                # Çalışma saati geldi mi? (0-30 dakika aralığında)
                time_diff_minutes = (current_time.hour * 60 + current_time.minute) - (shift_start_time.hour * 60 + shift_start_time.minute)
                
                if time_diff_minutes < 0:
                    # Henüz çalışma saati gelmemiş
                    logger.debug(f"[AUTO_SHIFT_START] Atama {assignment['_id']} için henüz zaman gelmedi. Başlangıç: {shift_start_time}, Şu an: {current_time}")
                    skipped_count += 1
                    continue
                
                if time_diff_minutes > 30:
                    # Çalışma saati 30 dakikadan fazla geçmiş, atla (muhtemelen manuel başlatılmalı)
                    logger.debug(f"[AUTO_SHIFT_START] Atama {assignment['_id']} için çalışma saati çok geçmiş ({time_diff_minutes} dakika), atlanıyor")
                    skipped_count += 1
                    continue
                
                # Kullanıcının aktif vardiyası var mı kontrol et
                active_shift = await shifts_collection.find_one({
                    "user_id": user_id,
                    "status": {"$in": ["active", "on_break"]}
                })
                
                if active_shift:
                    logger.info(f"[AUTO_SHIFT_START] Kullanıcı {user.get('name')} zaten aktif vardiyaya sahip, atlanıyor")
                    skipped_count += 1
                    continue
                
                # Vardiyayı otomatik başlat
                logger.info(f"[AUTO_SHIFT_START] Vardiya otomatik başlatılıyor - Kullanıcı: {user.get('name')}, Rol: {user_role}, Atama ID: {assignment['_id']}")
                
                # Shift oluştur
                new_shift = Shift(
                    assignment_id=assignment["_id"],
                    user_id=user_id,
                    vehicle_id=assignment.get("vehicle_id"),
                    vehicle_plate=None,  # Sağlık merkezi için araç yok
                    start_time=get_turkey_time(),
                    status="active",
                    location_type="saglik_merkezi",
                    health_center_name=assignment.get("health_center_name"),
                    started_by_admin=False,  # Otomatik başlatıldı
                    admin_note="Çalışma saati geldiğinde otomatik olarak başlatıldı"
                )
                
                shift_dict = new_shift.model_dump(by_alias=True)
                await shifts_collection.insert_one(shift_dict)
                
                # Assignment durumunu güncelle
                await shift_assignments_collection.update_one(
                    {"_id": assignment["_id"]},
                    {
                        "$set": {
                            "status": "started",
                            "started_at": get_turkey_time(),
                            "auto_started": True,
                            "auto_started_at": get_turkey_time()
                        }
                    }
                )
                
                started_count += 1
                logger.info(f"[AUTO_SHIFT_START] Vardiya başarıyla başlatıldı - Shift ID: {shift_dict['_id']}, Kullanıcı: {user.get('name')}")
                
            except Exception as e:
                logger.error(f"[AUTO_SHIFT_START] Atama {assignment.get('_id')} için hata: {e}", exc_info=True)
                skipped_count += 1
                continue
        
        if started_count > 0:
            logger.info(f"[AUTO_SHIFT_START] Tamamlandı - Başlatılan: {started_count}, Atlanan: {skipped_count}")
        
    except Exception as e:
        logger.error(f"[AUTO_SHIFT_START] Genel hata: {e}", exc_info=True)