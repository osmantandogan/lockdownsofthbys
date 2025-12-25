from fastapi import APIRouter, HTTPException, Request, BackgroundTasks
from typing import List, Optional, Any
from database import cases_collection, vehicles_collection, users_collection, shift_assignments_collection, medication_usage_collection
from models import Case, CaseCreate, CaseAssignTeam, CaseUpdateStatus, CaseStatusUpdate, MedicalFormData, DoctorApproval, CaseParticipant
from auth_utils import get_current_user, require_roles
from datetime import datetime, timedelta
from utils.timezone import get_turkey_time
from email_service import send_case_notifications
from pydantic import BaseModel
import uuid
import os
import logging

# Bildirim servisi
try:
    from services.notification_service import notification_service, NotificationType, NotificationChannel
    NOTIFICATIONS_ENABLED = True
except ImportError:
    NOTIFICATIONS_ENABLED = False

# FCM Bildirim servisi
try:
    from services.firebase_service import send_fcm_to_multiple, send_case_notification_fcm
    FCM_ENABLED = True
except ImportError:
    FCM_ENABLED = False
    
logger = logging.getLogger(__name__)

router = APIRouter()

# Vaka numarasƒ± ba≈ülangƒ±√ß deƒüeri - 1'den ba≈ülar, 6 haneli format (000001)
CASE_NUMBER_START = 1

# Counter collection - g√ºnl√ºk vaka numarasƒ± i√ßin
from database import db
counters_collection = db["counters"]

async def peek_next_case_sequence() -> int:
    """G√ºnl√ºk sƒ±ralƒ± vaka numarasƒ±nƒ± √ñNƒ∞ZLE - Counter'ƒ± ARTIRMAZ"""
    # T√ºrkiye saati (UTC+3)
    turkey_now = get_turkey_time()
    today = turkey_now.strftime("%Y%m%d")
    counter_id = f"case_number_{today}"
    
    # Sadece oku, artƒ±rma
    result = await counters_collection.find_one({"_id": counter_id})
    
    # Eƒüer yoksa ilk vaka 1 olacak
    current_seq = result["seq"] if result else 0
    next_seq = current_seq + 1
    
    logger.info(f"[CASE_NUMBER_PEEK] Bug√ºn ({today}) i√ßin sonraki sƒ±ra: {next_seq}")
    return next_seq

async def get_next_case_sequence() -> int:
    """G√ºnl√ºk sƒ±ralƒ± vaka numarasƒ± al - Atomic counter kullanarak (ARTIRIR)"""
    from pymongo import ReturnDocument
    
    # T√ºrkiye saati (UTC+3)
    turkey_now = get_turkey_time()
    today = turkey_now.strftime("%Y%m%d")
    counter_id = f"case_number_{today}"
    
    # Atomic increment - findAndModify ile
    result = await counters_collection.find_one_and_update(
        {"_id": counter_id},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=ReturnDocument.AFTER
    )
    
    seq = result["seq"] if result else 1
    logger.info(f"[CASE_NUMBER] Bug√ºn ({today}) i√ßin sƒ±ra: {seq}")
    
    return seq

async def generate_case_number() -> str:
    """Generate case number in format YYYYMMDD-XXXXXX (starting from 000001)"""
    # T√ºrkiye saati (UTC+3)
    now = get_turkey_time()
    date_str = now.strftime("%Y%m%d")
    seq = await get_next_case_sequence()
    # Tam 6 haneli format: 000001, 000002, ...
    seq_str = str(seq).zfill(6)
    return f"{date_str}-{seq_str}"

@router.get("/next-case-number")
async def get_next_case_number(request: Request):
    """Sonraki vaka numarasƒ±nƒ± d√∂nd√ºr (SADECE √ñNƒ∞ZLEME - Counter artƒ±rmaz)"""
    user = await get_current_user(request)
    
    # T√ºrkiye saati (UTC+3)
    now = get_turkey_time()
    date_str = now.strftime("%Y%m%d")
    
    # PEEK kullan - counter'ƒ± artƒ±rmaz!
    seq = await peek_next_case_sequence()
    seq_str = str(seq).zfill(6)
    
    return {"next_case_number": f"{date_str}-{seq_str}"}

@router.post("", response_model=Case)
async def create_case(data: CaseCreate, request: Request):
    """Create new case (Call Center)"""
    user = await get_current_user(request)
    
    # Generate case number (async - sƒ±ralƒ± numara)
    case_number = await generate_case_number()
    
    # Timestamps olu≈ütur - √ßaƒürƒ± saatini kaydet
    timestamps_data = None
    if hasattr(data, 'timestamps') and data.timestamps:
        timestamps_data = data.timestamps.model_dump() if hasattr(data.timestamps, 'model_dump') else data.timestamps
    else:
        # Varsayƒ±lan olarak √ßaƒürƒ± saatini ≈üimdi olarak ayarla
        timestamps_data = {"call_received": datetime.now().isoformat()}
    
    # Kaynak belirle: case_details.type == 'ayaktan_basvuru' ise 'registration', yoksa 'call_center'
    case_details = data.case_details if hasattr(data, 'case_details') else None
    source = "registration" if (case_details and case_details.get("type") == "ayaktan_basvuru") else "call_center"
    
    # Olu≈üturulma lokasyonu - case_details'tan al veya lokasyondan
    created_location = None
    if case_details:
        created_location = case_details.get("user_location")
    if not created_location and data.location:
        created_location = data.location.address
    
    # Create case
    new_case = Case(
        case_number=case_number,
        caller=data.caller,
        patient=data.patient,
        location=data.location,
        priority=data.priority,
        case_details=case_details,
        source=source,
        timestamps=timestamps_data,
        created_by=user.id,
        created_by_name=user.name,
        created_by_role=user.role,
        created_location=created_location
    )
    
    # Add initial status to history - kullanƒ±cƒ± adƒ± ve rol√º ile
    initial_status = CaseStatusUpdate(
        status="acildi",
        note="Vaka olu≈üturuldu",
        updated_by=user.id,
        updated_by_name=user.name,
        updated_by_role=user.role
    )
    new_case.status_history.append(initial_status)
    
    case_dict = new_case.model_dump(by_alias=True)
    await cases_collection.insert_one(case_dict)
    
    # Vaka olu≈üturma bildirimi g√∂nder (arka planda)
    if NOTIFICATIONS_ENABLED:
        try:
            from services.notification_service import get_users_by_role
            
            # ƒ∞lgili rollere bildirim g√∂nder
            managers = await get_users_by_role(["merkez_ofis", "operasyon_muduru"])
            recipients = []
            for manager in managers:
                recipients.append({
                    "user_id": manager["id"],
                    "phone": manager.get("phone"),
                    "push_subscriptions": manager.get("push_subscriptions", []),
                    "fcm_token": manager.get("fcm_token"),
                    "notification_preferences": manager.get("notification_preferences", {})
                })
            
            if recipients:
                patient_info = case_dict.get("patient", {})
                location_info = case_dict.get("location", {})
                
                await notification_service.send_notification(
                    NotificationType.CASE_CREATED,
                    recipients,
                    {
                        "case_number": case_number,
                        "patient_name": f"{patient_info.get('name', '')} {patient_info.get('surname', '')}".strip() or "Belirtilmemi≈ü",
                        "location": location_info.get("address", "Belirtilmemi≈ü"),
                        "priority": case_dict.get("priority", "Normal"),
                        "created_at": datetime.now().strftime("%d.%m.%Y %H:%M"),
                        "url": f"/dashboard/cases/{case_dict['_id']}"
                    }
                )
                logger.info(f"Sent case creation notifications to {len(recipients)} recipients")
        except Exception as e:
            logger.error(f"Error sending case creation notifications: {e}", exc_info=True)
    
    # Return with 'id' field for frontend
    case_dict["id"] = case_dict.pop("_id")
    return case_dict

@router.get("", response_model=List[Case])
async def get_cases(
    request: Request,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    search: Optional[str] = None,
    source: Optional[str] = None,  # 'call_center' or 'registration'
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    last_24h: Optional[bool] = False,
    user_id: Optional[str] = None,  # Hangi kullanƒ±cƒ± i≈ülem yapmƒ±≈ü
    medication_name: Optional[str] = None,  # Hangi ila√ß kullanƒ±lmƒ±≈ü
    has_hospital_transfer: Optional[bool] = None,  # Hastaneye sevk var mƒ±
    page: Optional[int] = 1,
    limit: Optional[int] = 30
):
    """Get all cases with filters"""
    user = await get_current_user(request)
    
    # Build query
    query = {}
    filters = []
    
    # Role-based filtering - saha personeli sadece atandƒ±klarƒ± veya olu≈üturduklarƒ± vakalarƒ± g√∂r√ºr
    # Hem≈üire t√ºm vakalarƒ± g√∂rebilir
    if user.role in ["paramedik", "att", "sofor", "bas_sofor"]:
        # Son 24 saat i√ßinde atandƒ±ƒüƒ± VEYA olu≈üturduƒüu vakalarƒ± g√∂rs√ºn
        last_24h_for_assignment = get_turkey_time() - timedelta(hours=24)
        
        # Kullanƒ±cƒ±nƒ±n bug√ºnk√º vardiya atamasƒ±ndaki ara√ßlarƒ± bul
        user_vehicle_ids = []
        try:
            today = get_turkey_time().replace(hour=0, minute=0, second=0, microsecond=0)
            tomorrow = today + timedelta(days=1)
            
            user_shift_assignments = await shift_assignments_collection.find({
                "user_id": user.id,
                "status": {"$in": ["pending", "started"]},
            "$or": [
                    {"shift_date": {"$gte": today, "$lt": tomorrow}},
                    {"shift_date": {"$lte": today}, "end_date": {"$gte": today}}
                ]
            }).to_list(10)
            
            for assignment in user_shift_assignments:
                vehicle_id = assignment.get("vehicle_id")
                if vehicle_id and vehicle_id not in user_vehicle_ids:
                    user_vehicle_ids.append(vehicle_id)
            
            logger.info(f"[Cases] User {user.id} has {len(user_vehicle_ids)} vehicle assignments: {user_vehicle_ids}")
        except Exception as e:
            logger.warning(f"[Cases] Error getting user vehicle assignments: {e}")
        
        # Temel filtreler
        or_conditions = [
            # assigned_team (tekil atama) kontrol√º
                {"assigned_team.driver_id": user.id},
                {"assigned_team.paramedic_id": user.id},
                {"assigned_team.att_id": user.id},
                {"assigned_team.nurse_id": user.id},
            # assigned_teams (√ßoklu atama) kontrol√º - array i√ßinde arama
            {"assigned_teams.driver_id": user.id},
            {"assigned_teams.paramedic_id": user.id},
            {"assigned_teams.att_id": user.id},
            {"assigned_teams.nurse_id": user.id},
            # Kendi olu≈üturduƒüu vakalar
            {"created_by": user.id},
            # Son 24 saat i√ßinde olu≈üturulan vakalar (atama ge√ßmi≈üi i√ßin)
            {
                "created_at": {"$gte": last_24h_for_assignment},
                "status_history.updated_by": user.id
            }
        ]
        
        # Kullanƒ±cƒ±nƒ±n vardiya atamasƒ±ndaki ara√ßlara atanan vakalarƒ± da ekle
        for vehicle_id in user_vehicle_ids:
            or_conditions.append({"assigned_team.vehicle_id": vehicle_id})
            or_conditions.append({"assigned_teams.vehicle_id": vehicle_id})
        
        filters.append({"$or": or_conditions})
    # Hem≈üire t√ºm vakalarƒ± g√∂rebilir - filtre ekleme
    
    if status:
        filters.append({"status": status})
    
    if priority:
        filters.append({"priority": priority})
    
    if search:
        filters.append({
            "$or": [
                {"case_number": {"$regex": search, "$options": "i"}},
                {"patient.name": {"$regex": search, "$options": "i"}},
                {"patient.surname": {"$regex": search, "$options": "i"}}
            ]
        })
    
    # Source filter (kaynak filtresi)
    if source:
        if source == "call_center":
            # √áaƒürƒ± merkezi - case_details.type yoksa veya 'ayaktan_basvuru' deƒüilse
            filters.append({
                "$or": [
                    {"case_details.type": {"$exists": False}},
                    {"case_details.type": {"$ne": "ayaktan_basvuru"}}
                ]
            })
        elif source == "registration":
            # Kayƒ±t ekranƒ± - case_details.type == 'ayaktan_basvuru'
            filters.append({"case_details.type": "ayaktan_basvuru"})
    
    # Date filters
    if last_24h:
        last_24h_time = get_turkey_time() - timedelta(hours=24)
        filters.append({"created_at": {"$gte": last_24h_time}})
    else:
        if start_date:
            try:
                start = datetime.fromisoformat(start_date.replace("Z", "+00:00"))
                start = start.replace(hour=0, minute=0, second=0, microsecond=0)
                filters.append({"created_at": {"$gte": start}})
            except:
                pass
        if end_date:
            try:
                end = datetime.fromisoformat(end_date.replace("Z", "+00:00"))
                end = end.replace(hour=23, minute=59, second=59, microsecond=999999)
                filters.append({"created_at": {"$lte": end}})
            except:
                pass
    
    # Combine all filters with $and
    if filters:
        query = {"$and": filters} if len(filters) > 1 else filters[0]
    
    # Pagination
    skip = (page - 1) * limit if page > 0 else 0
    
    # Get cases - vaka numarasƒ±na g√∂re azalan sƒ±ralama (en yeni vaka en √ºstte)
    cases_cursor = cases_collection.find(query).sort("case_number", -1).skip(skip).limit(limit)
    cases = await cases_cursor.to_list(limit)
    
    # Advanced filters (user_id, medication_name, has_hospital_transfer)
    # These require additional lookups
    if user_id or medication_name is not None or has_hospital_transfer is not None:
        # Get case IDs that match advanced filters
        advanced_query = {}
        if user_id:
            # Kullanƒ±cƒ±nƒ±n i≈ülem yaptƒ±ƒüƒ± vakalar
            # Status history'de veya medication usage'da veya participants'da
            advanced_query["$or"] = [
                {"status_history.updated_by": user_id},
                {"participants.user_id": user_id},
                {"assigned_team.driver_id": user_id},
                {"assigned_team.paramedic_id": user_id},
                {"assigned_team.att_id": user_id},
                {"assigned_team.nurse_id": user_id}
            ]
        
        if medication_name:
            # Medication usage'dan ila√ß adƒ±na g√∂re vakalarƒ± bul
            med_cases = await medication_usage_collection.find({
                "name": {"$regex": medication_name, "$options": "i"}
            }).to_list(1000)
            med_case_ids = [m.get("case_id") for m in med_cases if m.get("case_id")]
            if med_case_ids:
                if "$or" in advanced_query:
                    advanced_query["$or"].append({"_id": {"$in": med_case_ids}})
                else:
                    advanced_query["_id"] = {"$in": med_case_ids}
            else:
                # Eƒüer ila√ß bulunamazsa bo≈ü liste d√∂nd√ºr
                return []
        
        if has_hospital_transfer is not None:
            # Transfer sekmesinde "Hastaneye Nakil" veya "Hastaneler Arasƒ± Nakil" var mƒ±?
            if has_hospital_transfer:
                # Medical form'da transfer bilgisi var mƒ± kontrol et
                # Bu durumda medical_form.transfers array'inde "Hastaneye Nakil" veya "Hastaneler Arasƒ± Nakil" olmalƒ±
                pass  # Bu filtreyi case query'ye ekleyemeyiz √ß√ºnk√º medical_form nested, sonra filtreleyeceƒüiz
        
        # Apply advanced filters to cases
        if advanced_query:
            advanced_cases = await cases_collection.find(advanced_query).to_list(1000)
            advanced_case_ids = [c["_id"] for c in advanced_cases]
            cases = [c for c in cases if c["_id"] in advanced_case_ids]
        
        # Hospital transfer filter
        if has_hospital_transfer is not None:
            filtered_cases = []
            for case in cases:
                medical_form = case.get("medical_form", {})
                transfers = medical_form.get("transfers", {})
                has_transfer = (
                    transfers.get("Hastaneye Nakil") or 
                    transfers.get("Hastaneler Arasƒ± Nakil")
                )
                if has_hospital_transfer == has_transfer:
                    filtered_cases.append(case)
            cases = filtered_cases
    
    for case in cases:
        case["id"] = case.pop("_id")
        
        # Add source info for frontend
        case_details = case.get("case_details", {})
        if case_details.get("type") == "ayaktan_basvuru":
            case["source"] = "registration"
        else:
            case["source"] = "call_center"
    
    return cases

@router.get("/{case_id}")
async def get_case(case_id: str, request: Request):
    """Get case by ID with 36-hour access restriction"""
    user = await get_current_user(request)
    
    case_doc = await cases_collection.find_one({"_id": case_id})
    if not case_doc:
        raise HTTPException(status_code=404, detail="Case not found")
    
    # 36 saat eri≈üim kƒ±sƒ±tƒ± kontrol√º
    access_info = await check_case_access(case_doc, user)
    
    case_doc["id"] = case_doc.pop("_id")
    case_doc["access_info"] = access_info
    
    return case_doc


async def check_case_access(case_doc: dict, user) -> dict:
    """
    36 saat eri≈üim kƒ±sƒ±tƒ± kontrol√º
    Returns: access_info dict with can_edit, is_restricted, reason
    """
    # Varsayƒ±lan: tam eri≈üim
    access_info = {
        "can_view": True,
        "can_edit": True,
        "is_restricted": False,
        "reason": None,
        "requires_approval": False
    }
    
    # Operasyon M√ºd√ºr√º ve Merkez Ofis i√ßin her zaman tam eri≈üim
    exempt_roles = ['operasyon_muduru', 'merkez_ofis']
    if user.role in exempt_roles:
        return access_info
    
    # Vakanƒ±n olu≈üturulma zamanƒ±nƒ± kontrol et
    created_at = case_doc.get("created_at")
    if not created_at:
        return access_info
    
    # 36 saat kontrol√º
    hours_36 = timedelta(hours=36)
    now = get_turkey_time()
    
    # Eƒüer created_at string ise parse et
    if isinstance(created_at, str):
        try:
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        except:
            return access_info
    
    time_elapsed = now - created_at.replace(tzinfo=None)
    
    if time_elapsed > hours_36:
        # 36 saat ge√ßmi≈ü - kƒ±sƒ±tlƒ± eri≈üim
        access_info["can_edit"] = False
        access_info["is_restricted"] = True
        access_info["reason"] = "36 saat ge√ßtiƒüi i√ßin d√ºzenleme yetkisi kaldƒ±rƒ±ldƒ±"
        access_info["requires_approval"] = True
        access_info["hours_elapsed"] = int(time_elapsed.total_seconds() / 3600)
    
    return access_info


class CaseAccessApprovalRequest(BaseModel):
    """Vaka eri≈üim onayƒ± isteƒüi"""
    case_id: str
    otp_code: str
    approver_id: str  # Onaylayan m√ºd√ºr√ºn ID'si


@router.post("/{case_id}/request-access")
async def request_case_access(case_id: str, data: CaseAccessApprovalRequest, request: Request):
    """
    36 saat sonrasƒ± vaka eri≈üimi i√ßin OTP onayƒ± al
    M√ºd√ºr√ºn OTP koduyla onay gerektirir
    """
    user = await get_current_user(request)
    
    # Vakayƒ± kontrol et
    case_doc = await cases_collection.find_one({"_id": case_id})
    if not case_doc:
        raise HTTPException(status_code=404, detail="Vaka bulunamadƒ±")
    
    # Onaylayƒ±cƒ±yƒ± kontrol et (m√ºd√ºr veya merkez ofis olmalƒ±)
    approver = await users_collection.find_one({"_id": data.approver_id})
    if not approver:
        raise HTTPException(status_code=404, detail="Onaylayƒ±cƒ± bulunamadƒ±")
    
    if approver.get("role") not in ['operasyon_muduru', 'merkez_ofis']:
        raise HTTPException(status_code=403, detail="Sadece m√ºd√ºr veya merkez ofis onaylayabilir")
    
    # OTP kodunu doƒürula
    from services.otp_service import verify_user_otp
    
    otp_secret = approver.get("otp_secret")
    if not otp_secret:
        raise HTTPException(status_code=400, detail="Onaylayƒ±cƒ±nƒ±n OTP'si olu≈üturulmamƒ±≈ü")
    
    if not verify_user_otp(otp_secret, data.otp_code):
        raise HTTPException(status_code=400, detail="Ge√ßersiz onay kodu")
    
    # Eri≈üim izni ver - vakaya ge√ßici eri≈üim kaydƒ± ekle
    access_grant = {
        "user_id": user.id,
        "granted_by": data.approver_id,
        "granted_at": get_turkey_time(),
        "expires_at": get_turkey_time() + timedelta(hours=4),  # 4 saatlik eri≈üim
        "otp_verified": True
    }
    
    await cases_collection.update_one(
        {"_id": case_id},
        {"$push": {"access_grants": access_grant}}
    )
    
    logger.info(f"Case {case_id} access granted to {user.id} by {data.approver_id}")
    
    return {
        "success": True,
        "message": "Eri≈üim onaylandƒ±. 4 saat boyunca d√ºzenleme yapabilirsiniz.",
        "expires_at": access_grant["expires_at"].isoformat()
    }


class PatientInfoUpdate(BaseModel):
    """Hasta bilgisi g√ºncelleme"""
    name: Optional[str] = None
    surname: Optional[str] = None
    tc_no: Optional[str] = None
    age: Optional[int] = None
    birth_date: Optional[str] = None
    gender: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    status: Optional[str] = None


@router.patch("/{case_id}/patient")
async def update_patient_info(case_id: str, data: PatientInfoUpdate, request: Request):
    """
    Hasta bilgilerini g√ºncelle (Ad-Soyad, TC, Ya≈ü, Cinsiyet)
    D√ºzenleme yetkisi olan t√ºm roller eri≈üebilir
    """
    user = await get_current_user(request)
    
    # D√ºzenleme yetkisi kontrol√º
    edit_roles = ['operasyon_muduru', 'merkez_ofis', 'doktor', 'hemsire', 'paramedik', 'att']
    if user.role not in edit_roles:
        raise HTTPException(status_code=403, detail="Bu i≈ülemi yapmaya yetkiniz yok")
    
    # Vakayƒ± kontrol et
    case_doc = await cases_collection.find_one({"_id": case_id})
    if not case_doc:
        raise HTTPException(status_code=404, detail="Vaka bulunamadƒ±")
    
    # G√ºncellenecek alanlarƒ± hazƒ±rla
    update_fields = {}
    if data.name is not None:
        update_fields["patient.name"] = data.name
    if data.surname is not None:
        update_fields["patient.surname"] = data.surname
    if data.tc_no is not None:
        update_fields["patient.tc_no"] = data.tc_no
    if data.age is not None:
        update_fields["patient.age"] = data.age
    if data.birth_date is not None:
        update_fields["patient.birth_date"] = data.birth_date
    if data.gender is not None:
        update_fields["patient.gender"] = data.gender
    if data.phone is not None:
        update_fields["patient.phone"] = data.phone
    if data.address is not None:
        update_fields["patient.address"] = data.address
    if data.status is not None:
        update_fields["patient.status"] = data.status
    
    if not update_fields:
        return {"message": "G√ºncellenecek alan yok"}
    
    update_fields["updated_at"] = get_turkey_time()
    
    await cases_collection.update_one(
        {"_id": case_id},
        {"$set": update_fields}
    )
    
    logger.info(f"Patient info updated for case {case_id} by {user.id}")
    
    return {
        "success": True,
        "message": "Hasta bilgileri g√ºncellendi"
    }


@router.post("/{case_id}/assign-team")
async def assign_team(case_id: str, data: CaseAssignTeam, request: Request):
    """Assign team to case (Operation Manager, Call Center, Nurse, Doctor)"""
    user = await get_current_user(request)
    
    # √áaƒürƒ± merkezi, hem≈üire, doktor ve ba≈ü ≈üof√∂r de ekip atayabilir
    allowed_roles = ["merkez_ofis", "operasyon_muduru", "cagri_merkezi", "hemsire", "doktor", "bas_sofor"]
    if user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="Bu i≈ülemi yapmaya yetkiniz yok")
    
    # Check if vehicle exists and is available
    vehicle = await vehicles_collection.find_one({"_id": data.vehicle_id})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    # Sadece ambulans tipindeki ara√ßlara vaka atamasƒ± yapƒ±labilir
    if vehicle.get("type") != "ambulans":
        raise HTTPException(status_code=400, detail="Sadece ambulans tipindeki ara√ßlara vaka atamasƒ± yapƒ±labilir")
    
    if vehicle["status"] != "musait":
        raise HTTPException(status_code=400, detail="Vehicle is not available")
    
    # Bug√ºnk√º vardiya atamalarƒ±ndan ekibi otomatik bul
    turkey_now = get_turkey_time()
    today = turkey_now.date()
    current_hour = turkey_now.hour
    
    # Vardiya deƒüi≈üim saati: 08:00
    # Saat 08:00'dan √∂nce ise, d√ºnk√º gece vardiyasƒ± hala ge√ßerli olabilir
    # Saat 08:00'dan sonra ise, sadece bug√ºnk√º vardiyalar ge√ßerli
    is_after_shift_change = current_hour >= 8
    yesterday = today - timedelta(days=1)
    
    logger.info(f"[Shift] Current time: {turkey_now}, is_after_shift_change: {is_after_shift_change}")
    
    # O araca atanmƒ±≈ü personeli bul (t√ºm pending/started)
    all_vehicle_assignments = await shift_assignments_collection.find({
        "vehicle_id": data.vehicle_id,
        "status": {"$in": ["pending", "started"]}
    }).to_list(100)
    
    logger.info(f"Found {len(all_vehicle_assignments)} total assignments for vehicle {data.vehicle_id}")
    
    # BUG√úN i√ßin ge√ßerli atamalarƒ± filtrele
    vehicle_assignments = []
    for assignment in all_vehicle_assignments:
        shift_date = assignment.get("shift_date")
        end_date = assignment.get("end_date") or shift_date
        
        # Tarih formatƒ±nƒ± d√ºzelt
        if isinstance(shift_date, str):
            try:
                shift_date = datetime.fromisoformat(shift_date.replace('Z', '+00:00'))
            except:
                continue
        if isinstance(end_date, str):
            try:
                end_date = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            except:
                end_date = shift_date
        
        if shift_date is None:
            continue
            
        # Tarih kontrol√º
        shift_date_only = shift_date.date() if isinstance(shift_date, datetime) else shift_date
        end_date_only = end_date.date() if isinstance(end_date, datetime) else (end_date or shift_date_only)
        
        # KATIL KURAL: Eƒüer saat 08:00'dan sonra ise, sadece BUG√úN ba≈ülayan vardiyalarƒ± al
        # Bu, eski vardiyalarƒ±n atamada kullanƒ±lmasƒ±nƒ± engeller
        if is_after_shift_change:
            # Saat 08:00'dan sonra - sadece bug√ºn ba≈ülayan vardiyalar
            if shift_date_only == today:
                vehicle_assignments.append(assignment)
                logger.info(f"‚úÖ Assignment {assignment.get('_id')} valid (today only): shift_date={shift_date_only}")
            else:
                logger.info(f"‚ùå Assignment {assignment.get('_id')} SKIPPED (not today): shift_date={shift_date_only} != today={today}")
        else:
            # Saat 08:00'dan √∂nce - d√ºnk√º veya bug√ºnk√º vardiyalar
            if shift_date_only == today or shift_date_only == yesterday:
                vehicle_assignments.append(assignment)
                logger.info(f"‚úÖ Assignment {assignment.get('_id')} valid (early morning): shift_date={shift_date_only}")
            else:
                logger.info(f"‚ùå Assignment {assignment.get('_id')} SKIPPED: shift_date={shift_date_only} not in [{yesterday}, {today}]")
    
    logger.info(f"Found {len(vehicle_assignments)} TODAY's assignments for vehicle {data.vehicle_id}")
    
    # Ekip ID'lerini doldur (eƒüer g√∂nderilmemi≈üse)
    assigned_team = data.model_dump()
    
    for assignment in vehicle_assignments:
        user_id = assignment.get("user_id")
        logger.info(f"Processing assignment: user_id={user_id}")
        if user_id:
            # Kullanƒ±cƒ±nƒ±n rol√ºn√º bul
            assigned_user = await users_collection.find_one({"_id": user_id})
            if assigned_user:
                role = assigned_user.get("role")
                user_name = assigned_user.get("name", "")
                logger.info(f"User {user_id} has role: {role}, name: {user_name}")
                if role == "sofor" or role == "bas_sofor":
                    if not assigned_team.get("driver_id"):
                        assigned_team["driver_id"] = user_id
                        assigned_team["driver_name"] = user_name
                        logger.info(f"Set driver_id to {user_id}, name: {user_name}")
                elif role == "paramedik":
                    if not assigned_team.get("paramedic_id"):
                        assigned_team["paramedic_id"] = user_id
                        assigned_team["paramedic_name"] = user_name
                        logger.info(f"Set paramedic_id to {user_id}, name: {user_name}")
                elif role == "att":
                    if not assigned_team.get("att_id"):
                        assigned_team["att_id"] = user_id
                        assigned_team["att_name"] = user_name
                        logger.info(f"Set att_id to {user_id}, name: {user_name}")
                elif role == "hemsire":
                    if not assigned_team.get("nurse_id"):
                        assigned_team["nurse_id"] = user_id
                        assigned_team["nurse_name"] = user_name
                        logger.info(f"Set nurse_id to {user_id}, name: {user_name}")
                elif role == "doktor":
                    if not assigned_team.get("doctor_id"):
                        assigned_team["doctor_id"] = user_id
                        assigned_team["doctor_name"] = user_name
                        logger.info(f"Set doctor_id to {user_id}, name: {user_name}")
    
    logger.info(f"Final assigned_team: {assigned_team}")
    assigned_team["assigned_at"] = get_turkey_time()
    assigned_team["vehicle_plate"] = vehicle.get("plate", "")  # Ara√ß plakasƒ±nƒ± ekle
    assigned_team["assigned_by"] = user.id  # Kim atadƒ±
    assigned_team["assigned_by_name"] = user.name  # Kim atadƒ± (isim)
    assigned_team["assigned_by_role"] = user.role  # Kim atadƒ± (rol)
    
    status_update = CaseStatusUpdate(
        status="ekip_bilgilendirildi",
        note="Ekip g√∂revlendirildi",
        updated_by=user.id,
        updated_by_name=user.name,
        updated_by_role=user.role
    )
    
    await cases_collection.update_one(
        {"_id": case_id},
        {
            "$set": {
                "assigned_team": assigned_team,
                "status": "ekip_bilgilendirildi",
                "updated_at": get_turkey_time()
            },
            "$push": {"status_history": status_update.model_dump()}
        }
    )
    
    # Update vehicle status
    await vehicles_collection.update_one(
        {"_id": data.vehicle_id},
        {
            "$set": {
                "status": "gorevde",
                "current_case_id": case_id,
                "updated_at": get_turkey_time()
            }
        }
    )
    
    # Bildirim g√∂nder (arka planda)
    if NOTIFICATIONS_ENABLED:
        try:
            # Vaka bilgilerini al
            case_doc = await cases_collection.find_one({"_id": case_id})
            
            # Alƒ±cƒ±larƒ± belirle: ekip √ºyeleri + m√ºd√ºr + doktor
            recipient_ids = []
            
            # Ekip √ºyelerini ekle
            for key in ["driver_id", "paramedic_id", "att_id", "nurse_id"]:
                if assigned_team.get(key):
                    recipient_ids.append(assigned_team[key])
            
            # Doktor ve operasyon m√ºd√ºr√º ekle
            managers = await users_collection.find({
                "role": {"$in": ["doktor", "operasyon_muduru", "merkez_ofis"]},
                "is_active": True
            }).to_list(50)
            
            for mgr in managers:
                if mgr["_id"] not in recipient_ids:
                    recipient_ids.append(mgr["_id"])
            
            # Alƒ±cƒ± bilgilerini topla
            recipients = []
            for rid in recipient_ids:
                user_doc = await users_collection.find_one({"_id": rid})
                if user_doc:
                    recipients.append({
                        "user_id": rid,
                        "phone": user_doc.get("phone"),
                        "push_subscriptions": user_doc.get("push_subscriptions", []),
                        "fcm_token": user_doc.get("fcm_token"),
                        "notification_preferences": user_doc.get("notification_preferences", {})
                    })
            
            # Vaka olu≈üturma bildirimi g√∂nder
            if recipients:
                patient_info = case_doc.get("patient", {})
                location_info = case_doc.get("location", {})
                
                await notification_service.send_notification(
                    NotificationType.CASE_ASSIGNED,
                    recipients,
                    {
                        "case_number": case_doc.get("case_number"),
                        "patient_name": f"{patient_info.get('name', '')} {patient_info.get('surname', '')}".strip() or "Belirtilmemi≈ü",
                        "location": location_info.get("address", "Belirtilmemi≈ü"),
                        "priority": case_doc.get("priority", "Normal"),
                        "vehicle_plate": vehicle.get("plate", ""),
                        "url": f"/dashboard/cases/{case_id}"
                    }
                )
                logger.info(f"Sent notifications to {len(recipients)} recipients for case {case_id}")
        except Exception as e:
            logger.error(f"Error sending case notifications: {e}")
    
    # ========== FCM Bƒ∞LDƒ∞Rƒ∞Mƒ∞ - ATT/PARAMEDƒ∞K/≈ûOF√ñR ƒ∞√áƒ∞N ACƒ∞L ALARM ==========
    logger.info(f"[FCM] FCM_ENABLED: {FCM_ENABLED}")
    if FCM_ENABLED:
        try:
            case_doc = await cases_collection.find_one({"_id": case_id})
            patient_info = case_doc.get("patient", {})
            location_info = case_doc.get("location", {})
            
            # Saha personelinin FCM token'larƒ±nƒ± topla (≈üof√∂r, att, paramedik, hem≈üire)
            # NOT: Doktor hari√ß - doktorlar web √ºzerinden takip ediyor
            field_personnel_ids = []
            for key in ["driver_id", "paramedic_id", "att_id", "nurse_id"]:
                if assigned_team.get(key):
                    field_personnel_ids.append(assigned_team[key])
                    logger.info(f"[FCM] Field personnel {key}: {assigned_team[key]}")
            
            logger.info(f"[FCM] Field personnel IDs: {field_personnel_ids}")
            
            fcm_tokens = []
            for user_id in field_personnel_ids:
                user_doc = await users_collection.find_one({"_id": user_id})
                if user_doc:
                    user_fcm_tokens = user_doc.get("fcm_tokens", [])
                    logger.info(f"[FCM] User {user_doc.get('name')} ({user_id}) has {len(user_fcm_tokens)} FCM tokens")
                    if user_fcm_tokens:
                        # Token objelerinden sadece token string'lerini √ßƒ±kar
                        for token_obj in user_fcm_tokens:
                            if isinstance(token_obj, dict) and token_obj.get("token"):
                                fcm_tokens.append(token_obj["token"])
                                logger.info(f"[FCM] Added token: ...{token_obj['token'][-20:]}")
                            elif isinstance(token_obj, str):
                                fcm_tokens.append(token_obj)
                                logger.info(f"[FCM] Added token (string): ...{token_obj[-20:]}")
                    else:
                        logger.warning(f"[FCM] User {user_doc.get('name')} has NO FCM tokens!")
                else:
                    logger.warning(f"[FCM] User {user_id} not found in database!")
            
            logger.info(f"[FCM] Total FCM tokens collected: {len(fcm_tokens)}")
            
            if fcm_tokens:
                patient_name = f"{patient_info.get('name', '')} {patient_info.get('surname', '')}".strip() or "Belirtilmemi≈ü"
                patient_phone = patient_info.get("phone") or case_doc.get("caller", {}).get("phone", "Belirtilmemi≈ü")
                patient_complaint = patient_info.get("complaint", "Belirtilmemi≈ü")
                address = location_info.get("address", "Belirtilmemi≈ü")
                priority = case_doc.get("priority", "Normal")
                
                # Acil/Kritik vakalar i√ßin emergency channel kullan
                notification_type = "new_case"  # Android'de emergency alarm tetikler
                
                result = await send_fcm_to_multiple(
                    tokens=fcm_tokens,
                    title=f"üö® YENƒ∞ VAKA - {priority.upper()}",
                    body=f"{patient_name}\nüìç {address}",
                    data={
                        "case_id": case_id,
                        "case_number": case_doc.get("case_number", ""),
                        "patient_name": patient_name,
                        "patient_phone": patient_phone,
                        "patient_complaint": patient_complaint,
                        "address": address,
                        "navigate_to": f"/dashboard/cases/{case_id}",
                        "target_roles": "att,paramedik,sofor"
                    },
                    notification_type=notification_type,
                    priority="high"
                )
                logger.info(f"FCM emergency sent to {len(fcm_tokens)} tokens: {result}")
            else:
                logger.warning(f"[FCM] No FCM tokens found for any field personnel! Notifications NOT sent.")
        except Exception as e:
            logger.error(f"Error sending FCM notification: {e}", exc_info=True)
    else:
        logger.warning("[FCM] FCM is disabled, skipping notifications")
    
    return {"message": "Team assigned successfully"}


@router.post("/{case_id}/assign-multiple-teams")
async def assign_multiple_teams(case_id: str, request: Request):
    """
    Vakaya birden fazla ara√ß/ekip ata
    Body: { "vehicle_ids": ["id1", "id2", ...] }
    """
    user = await get_current_user(request)
    
    # Yetki kontrol√º - ba≈ü ≈üof√∂r de ekleyebilsin
    allowed_roles = ["merkez_ofis", "operasyon_muduru", "cagri_merkezi", "hemsire", "doktor", "bas_sofor"]
    if user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="Bu i≈ülemi yapmaya yetkiniz yok")
    
    body = await request.json()
    vehicle_ids = body.get("vehicle_ids", [])
    
    if not vehicle_ids:
        raise HTTPException(status_code=400, detail="En az bir ara√ß se√ßmelisiniz")
    
    # Vakayƒ± kontrol et
    case_doc = await cases_collection.find_one({"_id": case_id})
    if not case_doc:
        raise HTTPException(status_code=404, detail="Vaka bulunamadƒ±")
    
    assigned_teams = []
    all_recipient_ids = []
    
    # Bug√ºn√ºn tarihini al
    turkey_now = get_turkey_time()
    today = turkey_now.date()
    current_hour = turkey_now.hour
    yesterday = today - timedelta(days=1)
    
    # Vardiya deƒüi≈üim saati: 08:00
    is_after_shift_change = current_hour >= 8
    
    logger.info(f"[MultiAssign] Current time: {turkey_now}, is_after_shift_change: {is_after_shift_change}")
    
    for vehicle_id in vehicle_ids:
        # Ara√ß kontrol√º
        vehicle = await vehicles_collection.find_one({"_id": vehicle_id})
        if not vehicle:
            logger.warning(f"Vehicle {vehicle_id} not found, skipping")
            continue
        
        # Sadece ambulans tipindeki ara√ßlara vaka atamasƒ± yapƒ±labilir
        if vehicle.get("type") != "ambulans":
            logger.warning(f"Vehicle {vehicle_id} is not an ambulance, skipping")
            continue
        
        # Ara√ß m√ºsait deƒüilse uyarƒ± ver ama engelleme (√ßoklu atama olabilir)
        if vehicle["status"] != "musait":
            logger.warning(f"Vehicle {vehicle_id} is not available (status: {vehicle['status']})")
        
        # Ekip bilgilerini al (t√ºm pending/started)
        all_vehicle_assignments = await shift_assignments_collection.find({
            "vehicle_id": vehicle_id,
            "status": {"$in": ["pending", "started"]}
        }).to_list(100)
        
        # BUG√úN i√ßin ge√ßerli atamalarƒ± filtrele (strict mode)
        vehicle_assignments = []
        for asgn in all_vehicle_assignments:
            shift_date = asgn.get("shift_date")
            end_date = asgn.get("end_date") or shift_date
            
            if isinstance(shift_date, str):
                try:
                    shift_date = datetime.fromisoformat(shift_date.replace('Z', '+00:00'))
                except:
                    continue
            if isinstance(end_date, str):
                try:
                    end_date = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                except:
                    end_date = shift_date
            
            if shift_date is None:
                continue
            
            shift_date_only = shift_date.date() if isinstance(shift_date, datetime) else shift_date
            
            # KATIL KURAL: Saat 08:00'dan sonra sadece bug√ºnk√º vardiyalarƒ± al
            if is_after_shift_change:
                if shift_date_only == today:
                    vehicle_assignments.append(asgn)
                    logger.info(f"‚úÖ [MultiAssign] Assignment {asgn.get('_id')} valid (today only): {shift_date_only}")
            else:
                # Saat 08:00'dan √∂nce - d√ºnk√º veya bug√ºnk√º
                if shift_date_only == today or shift_date_only == yesterday:
                    vehicle_assignments.append(asgn)
                    logger.info(f"‚úÖ [MultiAssign] Assignment {asgn.get('_id')} valid (early): {shift_date_only}")
        
        logger.info(f"Vehicle {vehicle_id}: {len(vehicle_assignments)} assignments valid for today")
        
        team_data = {
            "vehicle_id": vehicle_id,
            "vehicle_plate": vehicle.get("plate"),
            "assigned_at": get_turkey_time()
        }
        
        for assignment in vehicle_assignments:
            user_id = assignment.get("user_id")
            if user_id:
                assigned_user = await users_collection.find_one({"_id": user_id})
                if assigned_user:
                    role = assigned_user.get("role")
                    user_name = assigned_user.get("name", "")
                    if role in ["sofor", "bas_sofor"] and not team_data.get("driver_id"):
                        team_data["driver_id"] = user_id
                        team_data["driver_name"] = user_name
                    elif role == "paramedik" and not team_data.get("paramedic_id"):
                        team_data["paramedic_id"] = user_id
                        team_data["paramedic_name"] = user_name
                    elif role == "att" and not team_data.get("att_id"):
                        team_data["att_id"] = user_id
                        team_data["att_name"] = user_name
                    elif role == "hemsire" and not team_data.get("nurse_id"):
                        team_data["nurse_id"] = user_id
                        team_data["nurse_name"] = user_name
                    elif role == "doktor" and not team_data.get("doctor_id"):
                        team_data["doctor_id"] = user_id
                        team_data["doctor_name"] = user_name
                    
                    # Bildirim listesine ekle
                    if user_id not in all_recipient_ids:
                        all_recipient_ids.append(user_id)
        
        assigned_teams.append(team_data)
        
        # Ara√ß durumunu g√ºncelle
        await vehicles_collection.update_one(
            {"_id": vehicle_id},
            {
                "$set": {
                    "status": "gorevde",
                    "current_case_id": case_id,
                    "updated_at": get_turkey_time()
                }
            }
        )
    
    if not assigned_teams:
        raise HTTPException(status_code=400, detail="Hi√ßbir ara√ß atanamadƒ±")
    
    # Status update
    status_update = CaseStatusUpdate(
        status="ekip_bilgilendirildi",
        note=f"{len(assigned_teams)} ara√ß g√∂revlendirildi",
        updated_by=user.id,
        updated_by_name=user.name,
        updated_by_role=user.role
    )
    
    # Vakayƒ± g√ºncelle - ilk ekibi assigned_team olarak, t√ºm√ºn√º assigned_teams olarak kaydet
    update_data = {
        "assigned_team": assigned_teams[0],  # Geriye uyumluluk i√ßin ilk ekip
        "assigned_teams": assigned_teams,     # T√ºm ekipler
        "status": "ekip_bilgilendirildi",
        "updated_at": get_turkey_time()
    }
    
    await cases_collection.update_one(
        {"_id": case_id},
        {
            "$set": update_data,
            "$push": {"status_history": status_update.model_dump()}
        }
    )
    
    # Bildirimleri g√∂nder (OneSignal)
    if NOTIFICATIONS_ENABLED and all_recipient_ids:
        try:
            # notification_service kullan (onesignal_service yerine)
            await notification_service.send_to_users(
                NotificationType.CASE_ASSIGNED,
                all_recipient_ids,
                {
                    "case_number": case_doc.get("case_number"),
                    "patient_name": f"{case_doc.get('patient', {}).get('name', '')} {case_doc.get('patient', {}).get('surname', '')}".strip() or "Belirtilmemi≈ü",
                    "vehicle_plate": ", ".join([t.get("vehicle_plate", "") for t in assigned_teams]),
                    "case_id": case_id
                },
                url=f"/dashboard/cases/{case_id}"
            )
            logger.info(f"[OneSignal] Notification sent to {len(all_recipient_ids)} users for multiple team assignment")
        except Exception as e:
            logger.error(f"Error sending OneSignal notifications: {e}")
    
    # ========== FCM Bƒ∞LDƒ∞Rƒ∞Mƒ∞ - ATT/PARAMEDƒ∞K/≈ûOF√ñR ƒ∞√áƒ∞N ACƒ∞L ALARM ==========
    logger.info(f"[FCM-MultiTeam] FCM_ENABLED: {FCM_ENABLED}, all_recipient_ids: {all_recipient_ids}")
    if FCM_ENABLED and all_recipient_ids:
        try:
            patient_info = case_doc.get("patient", {})
            location_info = case_doc.get("location", {})
            
            # Saha personelinin FCM token'larƒ±nƒ± topla
            fcm_tokens = []
            for user_id in all_recipient_ids:
                user_doc = await users_collection.find_one({"_id": user_id})
                if user_doc:
                    # Sadece saha personeli i√ßin alarm g√∂nder
                    user_role = user_doc.get("role", "")
                    if user_role in ["sofor", "bas_sofor", "att", "paramedik"]:
                        user_fcm_tokens = user_doc.get("fcm_tokens", [])
                        logger.info(f"[FCM-MultiTeam] User {user_doc.get('name')} ({user_role}) has {len(user_fcm_tokens)} FCM tokens")
                        for token_obj in user_fcm_tokens:
                            if isinstance(token_obj, dict) and token_obj.get("token"):
                                fcm_tokens.append(token_obj["token"])
                                logger.info(f"[FCM-MultiTeam] Added token: ...{token_obj['token'][-20:]}")
                            elif isinstance(token_obj, str):
                                fcm_tokens.append(token_obj)
                                logger.info(f"[FCM-MultiTeam] Added token (string): ...{token_obj[-20:]}")
            
            logger.info(f"[FCM-MultiTeam] Total FCM tokens collected: {len(fcm_tokens)}")
            
            if fcm_tokens:
                patient_name = f"{patient_info.get('name', '')} {patient_info.get('surname', '')}".strip() or "Belirtilmemi≈ü"
                patient_phone = patient_info.get("phone") or case_doc.get("caller", {}).get("phone", "Belirtilmemi≈ü")
                patient_complaint = patient_info.get("complaint", "Belirtilmemi≈ü")
                address = location_info.get("address", "Belirtilmemi≈ü")
                priority = case_doc.get("priority", "Normal")
                
                result = await send_fcm_to_multiple(
                    tokens=fcm_tokens,
                    title=f"üö® YENƒ∞ VAKA - {priority.upper()}",
                    body=f"{patient_name}\nüìç {address}",
                    data={
                        "case_id": case_id,
                        "case_number": case_doc.get("case_number", ""),
                        "patient_name": patient_name,
                        "patient_phone": patient_phone,
                        "patient_complaint": patient_complaint,
                        "address": address,
                        "navigate_to": f"/dashboard/cases/{case_id}",
                        "target_roles": "att,paramedik,sofor"
                    },
                    notification_type="new_case",
                    priority="high"
                )
                logger.info(f"[FCM-MultiTeam] Emergency sent to {len(fcm_tokens)} tokens: {result}")
            else:
                logger.warning(f"[FCM-MultiTeam] No FCM tokens found for field personnel!")
        except Exception as e:
            logger.error(f"[FCM-MultiTeam] Error sending FCM: {e}", exc_info=True)
    else:
        if not FCM_ENABLED:
            logger.warning("[FCM-MultiTeam] FCM is disabled")
        if not all_recipient_ids:
            logger.warning("[FCM-MultiTeam] No recipients to notify")
    
    return {
        "message": f"{len(assigned_teams)} ara√ß ba≈üarƒ±yla g√∂revlendirildi",
        "assigned_teams": assigned_teams
    }


@router.patch("/{case_id}/status")
async def update_case_status(case_id: str, data: CaseUpdateStatus, request: Request):
    """Update case status"""
    user = await get_current_user(request)
    
    case_doc = await cases_collection.find_one({"_id": case_id})
    if not case_doc:
        raise HTTPException(status_code=404, detail="Case not found")
    
    # Create status update with user info
    status_update = CaseStatusUpdate(
        status=data.status,
        note=data.note,
        updated_by=user.id,
        updated_by_name=user.name,
        updated_by_role=user.role
    )
    
    update_data = {
        "status": data.status,
        "updated_at": get_turkey_time()
    }
    
    # If case is completed or cancelled, free the vehicle
    if data.status in ["tamamlandi", "iptal"]:
        if case_doc.get("assigned_team"):
            vehicle_id = case_doc["assigned_team"]["vehicle_id"]
            await vehicles_collection.update_one(
                {"_id": vehicle_id},
                {
                    "$set": {
                        "status": "musait",
                        "current_case_id": None,
                        "updated_at": get_turkey_time()
                    }
                }
            )
    
    await cases_collection.update_one(
        {"_id": case_id},
        {
            "$set": update_data,
            "$push": {"status_history": status_update.model_dump()}
        }
    )
    
    return {"message": "Case status updated successfully"}

@router.get("/stats/dashboard")
async def get_dashboard_stats(request: Request):
    """Get dashboard statistics"""
    await get_current_user(request)
    
    # Active cases
    active_cases = await cases_collection.count_documents({
        "status": {"$nin": ["tamamlandi", "iptal"]}
    })
    
    # Available vehicles
    available_vehicles = await vehicles_collection.count_documents({
        "status": "musait"
    })
    
    # Priority breakdown
    high_priority = await cases_collection.count_documents({
        "priority": "yuksek",
        "status": {"$nin": ["tamamlandi", "iptal"]}
    })
    
    return {
        "active_cases": active_cases,
        "available_vehicles": available_vehicles,
        "high_priority_cases": high_priority
    }

@router.post("/{case_id}/send-notification")
async def send_notification(case_id: str, vehicle_id: Optional[str] = None, request: Request = None):
    """Send notification about case to relevant users"""
    user = await get_current_user(request)
    
    # Get case
    case_doc = await cases_collection.find_one({"_id": case_id})
    if not case_doc:
        raise HTTPException(status_code=404, detail="Case not found")
    
    # Get vehicle plate if vehicle_id provided
    vehicle_plate = None
    if vehicle_id:
        vehicle = await vehicles_collection.find_one({"_id": vehicle_id})
        if vehicle:
            vehicle_plate = vehicle.get("plate")
    
    # Prepare notification data
    notification_data = {
        "case_number": case_doc.get("case_number"),
        "priority": case_doc.get("priority"),
        "patient": case_doc.get("patient"),
        "caller": case_doc.get("caller"),
        "location": case_doc.get("location"),
        "assigned_team": case_doc.get("assigned_team"),
        "vehicle_plate": vehicle_plate
    }
    
    # Send notifications
    sent_count = await send_case_notifications(notification_data, users_collection, vehicle_id)
    
    return {
        "message": "Notifications sent successfully",
        "sent_count": sent_count
    }

# ============================================================================
# REAL-TIME COLLABORATION ENDPOINTS
# ============================================================================

class MedicalFormUpdate(BaseModel):
    """Partial update for medical form"""
    field: str
    value: Any

@router.post("/{case_id}/join")
async def join_case(case_id: str, request: Request):
    """Join case as participant (real-time collaboration)"""
    try:
        user = await get_current_user(request)
        
        case_doc = await cases_collection.find_one({"_id": case_id})
        if not case_doc:
            raise HTTPException(status_code=404, detail="Vaka bulunamadƒ±")
        
        # Check if user is authorized (assigned to case or doctor/nurse/admin)
        is_authorized = False
        assigned_team = case_doc.get("assigned_team") or {}
        
        if user.role in ["merkez_ofis", "operasyon_muduru", "doktor", "hemsire", "att", "paramedik", "sofor", "bas_sofor"]:
            is_authorized = True
        elif user.id in [
            assigned_team.get("driver_id"),
            assigned_team.get("paramedic_id"),
            assigned_team.get("att_id"),
            assigned_team.get("nurse_id")
        ]:
            is_authorized = True
        
        if not is_authorized:
            raise HTTPException(status_code=403, detail="Bu vakaya eri≈üim yetkiniz yok")
        
        # Add or update participant
        participant = CaseParticipant(
            user_id=user.id or "",
            user_name=user.name or "Bilinmeyen",
            user_role=user.role or "personel",
            joined_at=get_turkey_time(),
            last_activity=get_turkey_time()
        )
        
        # Remove old entry if exists, then add new
        await cases_collection.update_one(
            {"_id": case_id},
            {"$pull": {"participants": {"user_id": user.id}}}
        )
        await cases_collection.update_one(
            {"_id": case_id},
            {"$push": {"participants": participant.model_dump()}}
        )
        
        return {"message": "Vakaya katƒ±ldƒ±nƒ±z", "participant": participant.model_dump()}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error joining case {case_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Vakaya katƒ±lƒ±rken hata: {str(e)}")

@router.post("/{case_id}/leave")
async def leave_case(case_id: str, request: Request):
    """Leave case as participant"""
    user = await get_current_user(request)
    
    await cases_collection.update_one(
        {"_id": case_id},
        {"$pull": {"participants": {"user_id": user.id}}}
    )
    
    return {"message": "Vakadan ayrƒ±ldƒ±nƒ±z"}

@router.get("/{case_id}/participants")
async def get_participants(case_id: str, request: Request):
    """Get active participants in case with profile photos"""
    await get_current_user(request)
    
    case_doc = await cases_collection.find_one({"_id": case_id})
    if not case_doc:
        raise HTTPException(status_code=404, detail="Vaka bulunamadƒ±")
    
    # Filter out inactive participants (last activity > 5 minutes ago)
    cutoff = get_turkey_time() - timedelta(minutes=5)
    participants = case_doc.get("participants", [])
    
    active_participants = []
    for p in participants:
        last_activity = p.get("last_activity")
        if isinstance(last_activity, str):
            last_activity = datetime.fromisoformat(last_activity.replace('Z', '+00:00'))
        if last_activity and last_activity > cutoff:
            # Kullanƒ±cƒ±nƒ±n profil fotoƒürafƒ±nƒ± al
            user_id = p.get("user_id")
            if user_id:
                user_doc = await users_collection.find_one({"_id": user_id})
                if user_doc:
                    p["profile_photo"] = user_doc.get("profile_photo")
            active_participants.append(p)
    
    return {"participants": active_participants}

def deep_merge(base: dict, updates: dict) -> dict:
    """
    ƒ∞√ß i√ße dictionary'leri birle≈ütirir (deep merge).
    √ñzellikle offline sync i√ßin √∂nemli - farklƒ± kullanƒ±cƒ±larƒ±n farklƒ± alanlarƒ±
    g√ºncellediƒüinde verilerin kaybolmasƒ±nƒ± √∂nler.
    
    √ñrnek:
    base = {"inline_consents": {"a": 1, "b": 2}, "vital_signs": [...]}
    updates = {"inline_consents": {"c": 3}}
    result = {"inline_consents": {"a": 1, "b": 2, "c": 3}, "vital_signs": [...]}
    """
    result = base.copy()
    
    for key, value in updates.items():
        if key in ["_id", "id"]:
            continue
            
        if key in result and isinstance(result[key], dict) and isinstance(value, dict):
            # Her iki deƒüer de dict ise, derin birle≈ütirme yap
            result[key] = deep_merge(result[key], value)
        elif key in result and isinstance(result[key], list) and isinstance(value, list):
            # Liste birle≈ütirme - yeni deƒüer bo≈ü deƒüilse kullan
            # (vital_signs gibi listeler i√ßin son g√∂nderilen ge√ßerli olsun)
            if value:  # Bo≈ü olmayan listeyi al
                result[key] = value
            # Yeni deƒüer bo≈üsa mevcut deƒüeri koru
        else:
            # Diƒüer durumlarda: yeni deƒüer None veya bo≈ü string deƒüilse g√ºncelle
            # None veya bo≈ü string ise mevcut deƒüeri koru
            if value is not None and value != "":
                result[key] = value
            elif key not in result:
                result[key] = value
    
    return result


@router.patch("/{case_id}/medical-form")
async def update_medical_form(case_id: str, request: Request):
    """
    Update medical form (real-time collaboration)
    DEEP MERGE: Farklƒ± kullanƒ±cƒ±larƒ±n farklƒ± alanlarƒ± offline g√ºncellemesi desteklenir.
    - ƒ∞mzalar ayrƒ±, vital signs ayrƒ±, clinical_obs ayrƒ± birle≈ütirilir
    - Hi√ßbir veri kaybolmaz
    """
    user = await get_current_user(request)
    
    case_doc = await cases_collection.find_one({"_id": case_id})
    if not case_doc:
        raise HTTPException(status_code=404, detail="Vaka bulunamadƒ±")
    
    # Get form data from request body
    form_data = await request.json()
    
    # Update medical form - DEEP MERGE
    current_form = case_doc.get("medical_form") or {}
    
    # Derin birle≈ütirme yap - farklƒ± kullanƒ±cƒ±larƒ±n verileri korunur
    merged_form = deep_merge(current_form, form_data)
    
    logger.info(f"[MedicalForm] Deep merge for case {case_id} by {user.name} ({user.role})")
    logger.info(f"[MedicalForm] Incoming keys: {list(form_data.keys())}")
    
    # Update case
    await cases_collection.update_one(
        {"_id": case_id},
        {
            "$set": {
                "medical_form": merged_form,
                "last_form_update": get_turkey_time(),
                "last_form_updater": user.id,
                "last_form_updater_name": user.name,
                "last_form_updater_role": user.role,
                "updated_at": get_turkey_time()
            }
        }
    )
    
    # Update participant's last activity
    await cases_collection.update_one(
        {"_id": case_id, "participants.user_id": user.id},
        {"$set": {"participants.$.last_activity": get_turkey_time()}}
    )
    
    return {
        "message": "Form g√ºncellendi",
        "updated_by": user.name,
        "updated_by_role": user.role,
        "updated_at": get_turkey_time().isoformat(),
        "merged_keys": list(form_data.keys())
    }

@router.get("/{case_id}/medical-form")
async def get_medical_form(case_id: str, request: Request):
    """Get medical form data"""
    await get_current_user(request)
    
    case_doc = await cases_collection.find_one({"_id": case_id})
    if not case_doc:
        raise HTTPException(status_code=404, detail="Vaka bulunamadƒ±")
    
    return {
        "medical_form": case_doc.get("medical_form"),
        "last_update": case_doc.get("last_form_update"),
        "last_updater": case_doc.get("last_form_updater"),
        "doctor_approval": case_doc.get("doctor_approval"),
        "participants": case_doc.get("participants", [])
    }

# ============================================================================
# DOCTOR APPROVAL ENDPOINTS
# ============================================================================

class DoctorApprovalRequest(BaseModel):
    status: str  # "approved" or "rejected"
    notes: Optional[str] = None
    rejection_reason: Optional[str] = None

@router.post("/{case_id}/doctor-approval")
async def doctor_approval(case_id: str, data: DoctorApprovalRequest, request: Request):
    """Doctor approves or rejects the case treatment"""
    user = await get_current_user(request)
    
    # Only doctors can approve
    if user.role != "doktor":
        raise HTTPException(status_code=403, detail="Sadece doktorlar onay verebilir")
    
    case_doc = await cases_collection.find_one({"_id": case_id})
    if not case_doc:
        raise HTTPException(status_code=404, detail="Vaka bulunamadƒ±")
    
    approval = DoctorApproval(
        status=data.status,
        doctor_id=user.id,
        doctor_name=user.name,
        approved_at=get_turkey_time(),
        rejection_reason=data.rejection_reason if data.status == "rejected" else None,
        notes=data.notes
    )
    
    # Update case
    await cases_collection.update_one(
        {"_id": case_id},
        {
            "$set": {
                "doctor_approval": approval.model_dump(),
                "updated_at": get_turkey_time()
            }
        }
    )
    
    # Add to status history with user info
    status_note = f"Doktor {user.name} tarafƒ±ndan {'onaylandƒ±' if data.status == 'approved' else 'reddedildi'}"
    if data.rejection_reason:
        status_note += f": {data.rejection_reason}"
    
    status_update = CaseStatusUpdate(
        status=case_doc.get("status"),  # Keep current status
        note=status_note,
        updated_by=user.id,
        updated_by_name=user.name,
        updated_by_role=user.role
    )
    
    await cases_collection.update_one(
        {"_id": case_id},
        {"$push": {"status_history": status_update.model_dump()}}
    )
    
    return {
        "message": f"ƒ∞≈ülem {'onaylandƒ±' if data.status == 'approved' else 'reddedildi'}",
        "approval": approval.model_dump()
    }

# ============================================================================
# VIDEO CALL ENDPOINTS
# ============================================================================

@router.post("/{case_id}/start-video-call")
async def start_video_call(case_id: str, request: Request):
    """Start video call for case using Jitsi Meet
    
    Kendi Jitsi sunucunuzu kullanmak i√ßin .env dosyasƒ±na ekleyin:
    JITSI_DOMAIN=jitsi.your-domain.com
    """
    user = await get_current_user(request)
    
    case_doc = await cases_collection.find_one({"_id": case_id})
    if not case_doc:
        raise HTTPException(status_code=404, detail="Vaka bulunamadƒ±")
    
    # Create a unique room name for Jitsi (always use case_number for consistency)
    case_number = case_doc.get("case_number", case_id[:8])
    # Clean room name - remove special characters for Jitsi
    video_room_id = f"HealMedy{case_number}".replace("-", "").replace(" ", "")
    
    # Get Jitsi domain from environment or use default
    jitsi_domain = os.environ.get("JITSI_DOMAIN", "meet.jit.si")
    
    # Update case with video call info
    await cases_collection.update_one(
        {"_id": case_id},
        {"$set": {
            "video_room_id": video_room_id, 
            "video_call_active": True,
            "jitsi_domain": jitsi_domain
        }}
    )
    
    # Build Jitsi URL
    jitsi_url = f"https://{jitsi_domain}/{video_room_id}"
    
    return {
        "room_id": video_room_id,
        "room_url": jitsi_url,
        "jitsi_domain": jitsi_domain,
        "started_by": user.name,
        "message": "G√∂r√ºnt√ºl√º g√∂r√º≈üme ba≈ülatƒ±ldƒ±."
    }


# ============================================================================
# EXCEL EXPORT ENDPOINT
# ============================================================================

from fastapi.responses import StreamingResponse
from services.excel_export_service import export_case_to_excel

@router.get("/{case_id}/export-excel")
async def export_case_excel(case_id: str, request: Request):
    """Vaka formunu Excel ≈üablonuna doldurarak indir"""
    user = await get_current_user(request)
    
    # Vakayƒ± getir
    case_doc = await cases_collection.find_one({"_id": case_id})
    if not case_doc:
        raise HTTPException(status_code=404, detail="Vaka bulunamadƒ±")
    
    # Vaka verilerini hazƒ±rla
    case_data = {
        "case_number": case_doc.get("case_number", ""),
        "created_at": case_doc.get("created_at"),
        "priority": case_doc.get("priority", ""),
        "status": case_doc.get("status", ""),
        "patient": case_doc.get("patient", {}),
        "caller": case_doc.get("caller", {}),
        "location": case_doc.get("location", {}),
        "assigned_team": case_doc.get("assigned_team", {}),
        "vehicle_info": case_doc.get("vehicle_info", {}),
        "time_info": case_doc.get("time_info", {}),
        "company": case_doc.get("company", ""),
        "call_type": case_doc.get("call_type", ""),
        "call_reason": case_doc.get("call_reason", ""),
        "complaint": case_doc.get("complaint", case_doc.get("patient", {}).get("complaint", "")),
        "chronic_diseases": case_doc.get("chronic_diseases", ""),
        "blood_sugar": case_doc.get("blood_sugar", ""),
        "body_temperature": case_doc.get("body_temperature", ""),
        "is_forensic": case_doc.get("is_forensic", False),
        "case_result": case_doc.get("case_result", ""),
        "transfer_hospital": case_doc.get("transfer_hospital", ""),
        "transfer_type": case_doc.get("transfer_type", ""),
        "referring_institution": case_doc.get("referring_institution", ""),
    }
    
    # Medical form verilerini al
    medical_form = case_doc.get("medical_form", {})
    
    # Vital signs
    vital_signs = case_doc.get("vital_signs", [])
    
    # Clinical observations
    clinical_obs = case_doc.get("clinical_observations", {})
    
    # CPR data
    cpr_data = case_doc.get("cpr_data", {})
    
    # Procedures
    procedures = case_doc.get("procedures", [])
    
    # Medications
    medications = case_doc.get("medications", [])
    
    # Materials
    materials = case_doc.get("materials", [])
    
    # Signatures
    signatures = case_doc.get("signatures", {})
    
    # Excel olu≈ütur
    try:
        excel_buffer = export_case_to_excel(
            case_data=case_data,
            medical_form=medical_form,
            vital_signs=vital_signs,
            clinical_obs=clinical_obs,
            cpr_data=cpr_data,
            procedures=procedures,
            medications=medications,
            materials=materials,
            signatures=signatures
        )
        
        # Dosya adƒ±
        case_number = case_doc.get("case_number", case_id[:8])
        date_str = get_turkey_time().strftime("%Y-%m-%d")
        filename = f"VAKA_FORMU_{case_number}_{date_str}.xlsx"
        
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Excel export hatasƒ±: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Excel olu≈üturma hatasƒ±: {str(e)}")


# ============================================================================
# Dƒ∞NAMƒ∞K EXCEL EXPORT - Mapping Kullanan
# ============================================================================

from services.dynamic_excel_export import export_case_with_template

@router.get("/{case_id}/export-excel-template/{template_id}")
async def export_case_with_excel_template(case_id: str, template_id: str, request: Request):
    """
    Vakayƒ± belirli bir Excel ≈üablonu ve onun data_mappings'i ile doldurarak indir
    """
    user = await get_current_user(request)
    
    # Vakayƒ± getir
    case_doc = await cases_collection.find_one({"_id": case_id})
    if not case_doc:
        raise HTTPException(status_code=404, detail="Vaka bulunamadƒ±")
    
    # Excel ≈üablonunu getir
    excel_templates_collection = db["excel_templates"]
    template = await excel_templates_collection.find_one({"_id": template_id})
    if not template:
        raise HTTPException(status_code=404, detail="Excel ≈üablonu bulunamadƒ±")
    
    # T√ºm vaka verilerini hazƒ±rla
    case_data = {
        "case_number": case_doc.get("case_number", ""),
        "created_at": case_doc.get("created_at"),
        "priority": case_doc.get("priority", ""),
        "status": case_doc.get("status", ""),
        "patient": case_doc.get("patient", {}),
        "caller": case_doc.get("caller", {}),
        "location": case_doc.get("location", {}),
        "assigned_team": case_doc.get("assigned_team", {}),
        "vehicle_info": case_doc.get("vehicle_info", {}),
        "time_info": case_doc.get("time_info", {}),
        "company": case_doc.get("company", ""),
        "call_type": case_doc.get("call_type", ""),
        "call_reason": case_doc.get("call_reason", ""),
        "complaint": case_doc.get("complaint", case_doc.get("patient", {}).get("complaint", "")),
        "chronic_diseases": case_doc.get("chronic_diseases", ""),
        "blood_sugar": case_doc.get("blood_sugar", ""),
        "body_temperature": case_doc.get("body_temperature", ""),
        "is_forensic": case_doc.get("is_forensic", False),
        "case_result": case_doc.get("case_result", ""),
        "transfer_hospital": case_doc.get("transfer_hospital", ""),
        "transfer_type": case_doc.get("transfer_type", ""),
        "referring_institution": case_doc.get("referring_institution", ""),
        "medical_form": case_doc.get("medical_form", {}),
        "vital_signs": case_doc.get("vital_signs", []),
        "clinical_observations": case_doc.get("clinical_observations", {}),
        "cpr_data": case_doc.get("cpr_data", {}),
        "procedures": case_doc.get("procedures", []),
        "medications": case_doc.get("medications", []),
        "materials": case_doc.get("materials", []),
        "fluids": case_doc.get("fluids", []) or case_doc.get("iv_fluids", []),
        "signatures": case_doc.get("signatures", {}),
        "hospital_rejection": case_doc.get("hospital_rejection", {}),
        "patient_rejection": case_doc.get("patient_rejection", {}),
        "escort": case_doc.get("escort", {}),
        "notes": case_doc.get("notes", ""),
    }
    
    try:
        # Dinamik export
        excel_buffer = export_case_with_template(template, case_data)
        
        # Dosya adƒ±
        case_number = case_doc.get("case_number", case_id[:8])
        template_name = template.get("name", "sablon")
        date_str = get_turkey_time().strftime("%Y-%m-%d")
        
        # T√ºrk√ße karakterleri temizle
        safe_name = template_name.replace('≈ü', 's').replace('≈û', 'S').replace('ƒ±', 'i').replace('ƒ∞', 'I')
        safe_name = safe_name.replace('ƒü', 'g').replace('ƒû', 'G').replace('√º', 'u').replace('√ú', 'U')
        safe_name = safe_name.replace('√∂', 'o').replace('√ñ', 'O').replace('√ß', 'c').replace('√á', 'C')
        
        filename = f"VAKA_{case_number}_{safe_name}_{date_str}.xlsx"
        
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
        
    except Exception as e:
        logger.error(f"Dinamik Excel export hatasƒ±: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Excel olu≈üturma hatasƒ±: {str(e)}")


# ============================================================================
# VAKA FORM MAPPING EXPORT - G√∂rsel Edit√∂r Mapping'i ile
# ============================================================================

from services.dynamic_excel_export import get_case_field_value

def build_export_case_data(case_doc: dict, medical_form: dict = None) -> dict:
    """
    Vaka ve medical_form verilerini export i√ßin birle≈ütirir
    Frontend'deki CaseDetail.js'de kaydedilen t√ºm verileri i√ßerir
    """
    if medical_form is None:
        medical_form = case_doc.get("medical_form") or {}
    
    extended_form = medical_form.get("extended_form") or {}
    clinical_obs = medical_form.get("clinical_obs") or {}
    
    # Vehicle info birle≈ütir - None kontrol√º
    case_vehicle_info = case_doc.get("vehicle_info") or {}
    form_vehicle_info = medical_form.get("vehicle_info") or {}
    vehicle_info = {**case_vehicle_info, **form_vehicle_info}
    
    # Time info birle≈ütir - None kontrol√º
    case_time_info = case_doc.get("time_info") or {}
    form_time_info = medical_form.get("time_info") or {}
    time_info = {**case_time_info, **form_time_info}
    
    # Assigned team - None kontrol√º
    assigned_team = case_doc.get("assigned_team") or {}
    
    return {
        "case_number": case_doc.get("case_number", ""),
        "created_at": case_doc.get("created_at"),
        "priority": case_doc.get("priority", ""),
        "status": case_doc.get("status", ""),
        "patient": case_doc.get("patient", {}),
        "caller": case_doc.get("caller", {}),
        "location": case_doc.get("location", {}),
        "assigned_team": assigned_team,
        
        # ATN NO - birden fazla yerden kontrol
        "atn_no": case_doc.get("atn_no", "") or extended_form.get("atnNo", "") or medical_form.get("atn_no", "") or vehicle_info.get("atn_no", ""),
        
        # PLAKA - assigned_team.vehicle veya vehicle_info.plate
        "vehicle_plate": assigned_team.get("vehicle", "") or vehicle_info.get("plate", "") or extended_form.get("vehiclePlate", "") or vehicle_info.get("plaka", ""),
        
        # VAKAYI VEREN KURUM - frontend'de referralSource olarak kaydediliyor
        "referring_institution": case_doc.get("referring_institution", "") or case_doc.get("company", "") or extended_form.get("referralSource", "") or extended_form.get("referringInstitution", "") or extended_form.get("vakayiVerenKurum", ""),
        
        # KM bilgileri
        "start_km": vehicle_info.get("start_km", "") or vehicle_info.get("baslangic_km", "") or extended_form.get("startKm", ""),
        "end_km": vehicle_info.get("end_km", "") or vehicle_info.get("bitis_km", "") or extended_form.get("endKm", ""),
        "total_km": vehicle_info.get("total_km", "") or vehicle_info.get("toplam_km", "") or extended_form.get("totalKm", ""),
        
        # ƒ∞stasyon - frontend'de stationCode olarak kaydediliyor
        "station_name": case_doc.get("station_name", "") or assigned_team.get("station", "") or extended_form.get("stationCode", "") or extended_form.get("stationName", ""),
        
        # Kodu
        "case_code": case_doc.get("case_code", "") or extended_form.get("caseCode", "") or case_doc.get("code", ""),
        
        # TRƒ∞YAJ KODU - frontend'de triageCode olarak kaydediliyor
        "triage_code": extended_form.get("triageCode", "") or case_doc.get("priority", ""),
        
        # vehicle_info ve time_info
        "vehicle_info": vehicle_info,
        "time_info": time_info,
        "company": case_doc.get("company", ""),
        
        # call_type ve call_reason: extended_form i√ßinden de al
        "call_type": case_doc.get("call_type", "") or extended_form.get("callType", ""),
        "call_reason": case_doc.get("call_reason", "") or extended_form.get("callReason", ""),
        "complaint": case_doc.get("complaint", "") or (case_doc.get("patient") or {}).get("complaint", ""),
        "chronic_diseases": case_doc.get("chronic_diseases", "") or medical_form.get("chronic_diseases", "") or extended_form.get("chronicDiseases", ""),
        "is_forensic": case_doc.get("is_forensic", False) or extended_form.get("isForensic", False),
        "case_result": case_doc.get("case_result", "") or extended_form.get("outcome", ""),
        "transfer_hospital": case_doc.get("transfer_hospital", "") or extended_form.get("transferHospital", "") or extended_form.get("nakledilenHastane", ""),
        
        # Vital signs - clinical_obs i√ßinden de al
        "vital_signs": case_doc.get("vital_signs", []) or medical_form.get("vital_signs", []) or clinical_obs.get("vital_signs", []),
        
        # Clinical observations - t√ºm kaynaklardan birle≈ütir (√∂ncelik sƒ±rasƒ± √∂nemli)
        "clinical_observations": {
            **(case_doc.get("clinical_observations") or {}), 
            **clinical_obs,
            # Kan ≈üekeri - birden fazla kaynak
            "blood_sugar": (clinical_obs.get("blood_sugar") or 
                           clinical_obs.get("bloodSugar") or 
                           extended_form.get("kanSekeri") or 
                           extended_form.get("bloodSugar") or ""),
            # Ate≈ü - birden fazla kaynak
            "temperature": (clinical_obs.get("temperature") or 
                           clinical_obs.get("temp") or 
                           clinical_obs.get("bodyTemp") or
                           extended_form.get("ates") or 
                           extended_form.get("bodyTemp") or ""),
            # Nabƒ±z tipi - object format
            "pulseType": clinical_obs.get("pulseType", {}) or extended_form.get("pulseType", {}) or {},
            "pulse_type": clinical_obs.get("pulseType", {}) or extended_form.get("pulseType", {}) or {},
            # Solunum tipi - object format
            "respType": clinical_obs.get("respType", {}) or extended_form.get("respType", {}) or {},
            "resp_type": clinical_obs.get("respType", {}) or extended_form.get("respType", {}) or {},
            # Pupil - object format
            "pupils": clinical_obs.get("pupils", {}) or extended_form.get("pupils", {}) or {},
            # Deri - object format
            "skin": clinical_obs.get("skin", {}) or extended_form.get("skin", {}) or {},
            # GKS - motor, verbal, eye ayrƒ± ayrƒ±
            "gks": clinical_obs.get("gks", {}) or extended_form.get("gks", {}) or {},
            "motorResponse": clinical_obs.get("motorResponse") or (clinical_obs.get("gks") or {}).get("motor") or extended_form.get("motorResponse") or 0,
            "verbalResponse": clinical_obs.get("verbalResponse") or (clinical_obs.get("gks") or {}).get("verbal") or extended_form.get("verbalResponse") or 0,
            "eyeOpening": clinical_obs.get("eyeOpening") or (clinical_obs.get("gks") or {}).get("eye") or extended_form.get("eyeOpening") or 0,
        },
        
        # √ñn tanƒ± ve a√ßƒ±klama
        "on_tani": medical_form.get("on_tani", "") or extended_form.get("onTani", "") or case_doc.get("preliminary_diagnosis", ""),
        "aciklamalar": medical_form.get("aciklamalar", "") or extended_form.get("aciklamalar", "") or case_doc.get("notes", ""),
        
        # Notlar ve medical_form (on_tani ve aciklamalar i√ßin fallback)
        "notes": medical_form.get("notes", "") or case_doc.get("notes", ""),
        "preliminary_diagnosis": medical_form.get("preliminary_diagnosis", []) or case_doc.get("preliminary_diagnosis", []),
        "medical_form": medical_form,  # T√ºm medical_form'u da aktar
        
        # Extended form (checkbox'lar i√ßin) - tam olarak aktar
        "extended_form": extended_form,
        
        # PROCEDURES - Dictionary format: {"Muayene (Acil)": {checked: true, adet: 2}}
        "procedures": medical_form.get("procedures", {}),
        
        # MEDICATIONS - hem case hem medical_form'dan, hem list hem dict format
        "medications": case_doc.get("medications", []) or medical_form.get("medications", []) or [],
        
        # PDF MEDICATIONS - Uygulamalar tabƒ±ndan se√ßilen ila√ßlar (PDF i√ßin)
        "pdf_medications": medical_form.get("pdf_medications", {}),
        
        # MATERIALS - Dictionary format: {"Enjekt√∂r 1-2 cc": {checked: true, adet: 5}}
        "materials": medical_form.get("materials", {}),
        
        # Transfers (nakil bilgileri)
        "transfers": medical_form.get("transfers", {}),
        
        # Fluids
        "fluids": case_doc.get("fluids", []) or case_doc.get("iv_fluids", []) or medical_form.get("fluids", []),
        
        # ƒ∞mzalar - hem signatures hem extended_form hem inline_consents'dan
        "signatures": {
            **(case_doc.get("signatures") or {}), 
            **(medical_form.get("signatures") or {}),
            # Extended form'dan da al
            "doctor_name": extended_form.get("hekimAdi", ""),
            "paramedic_name": extended_form.get("saglikPerAdi", ""),
            "driver_name": extended_form.get("soforAdi", ""),
            "receiver_name": extended_form.get("teslimAlanAdi", ""),
            "receiver_title": extended_form.get("teslimAlanUnvan", ""),
            "patient_name": extended_form.get("hastaYakiniAdi", ""),
        },
        
        # Ekip √ºyeleri isimleri (PDF i√ßin)
        "team_members": {
            "doctor_name": assigned_team.get("doctor_name", "") or extended_form.get("hekimAdi", ""),
            "paramedic_name": assigned_team.get("paramedic_name", "") or assigned_team.get("att_name", "") or extended_form.get("saglikPerAdi", ""),
            "driver_name": assigned_team.get("driver_name", "") or assigned_team.get("sofor_name", "") or extended_form.get("soforAdi", ""),
            "hemsire_name": assigned_team.get("hemsire_name", "") or assigned_team.get("nurse_name", ""),
        },
        
        # INLINE CONSENTS - ƒ∞mza tab'ƒ±ndaki veriler
        "inline_consents": medical_form.get("inline_consents", {}),
        
        "hospital_rejection": case_doc.get("hospital_rejection", {}),
        "patient_rejection": case_doc.get("patient_rejection", {}),
        
        # CPR data - t√ºm kaynaklardan birle≈ütir
        "cpr_data": {
            **(case_doc.get("cpr_data") or {}),
            **(medical_form.get("cpr_data") or {}),
            **(extended_form.get("cprData") or {}),
            # Explicit field mappings
            "start_time": ((medical_form.get("cpr_data") or {}).get("start_time") or 
                          (medical_form.get("cpr_data") or {}).get("startTime") or
                          extended_form.get("cprStartTime") or
                          extended_form.get("cpr_baslama") or ""),
            "stop_time": ((medical_form.get("cpr_data") or {}).get("stop_time") or 
                         (medical_form.get("cpr_data") or {}).get("stopTime") or
                         extended_form.get("cprStopTime") or
                         extended_form.get("cpr_bitis") or ""),
            "stop_reason": ((medical_form.get("cpr_data") or {}).get("stop_reason") or 
                           (medical_form.get("cpr_data") or {}).get("stopReason") or
                           extended_form.get("cprStopReason") or
                           extended_form.get("cpr_neden") or ""),
            "performed": ((medical_form.get("cpr_data") or {}).get("performed") or 
                         (medical_form.get("cpr_data") or {}).get("yapildi") or
                         extended_form.get("cprYapildi") or False),
        },
        
        # Isolation
        "isolation": medical_form.get("isolation", {}),
        
        # Scene type (olay yeri)
        "scene_type": extended_form.get("sceneType", "") or extended_form.get("scene", {}),
        
        # Transfer type
        "transfer_type": extended_form.get("transferType", ""),
        
        # Kazaya karƒ±≈üan ara√ßlar
        "crash_vehicles": extended_form.get("crashVehicles", []),
    }

@router.get("/{case_id}/export-excel-mapped")
async def export_case_with_vaka_form_mapping(case_id: str, request: Request):
    """Vakayƒ± Vaka Form Mapping kullanarak Excel'e export et (≈üablon formatƒ±yla)"""
    user = await get_current_user(request)
    
    case_doc = await cases_collection.find_one({"_id": case_id})
    if not case_doc:
        raise HTTPException(status_code=404, detail="Vaka bulunamadƒ±")
    
    mapping_doc = await db.vaka_form_mappings.find_one({"_id": "default"})
    if not mapping_doc or not mapping_doc.get("flat_mappings"):
        raise HTTPException(status_code=404, detail="Vaka form mapping bulunamadƒ±. √ñnce mapping olu≈üturun.")
    
    flat_mappings = mapping_doc.get("flat_mappings", {})
    logo_info = mapping_doc.get("logo", {})
    
    # Medical form verilerini al (CaseDetail'de kaydedilen)
    medical_form = case_doc.get("medical_form", {})
    
    case_data = build_export_case_data(case_doc, medical_form)
    
    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as XLImage
        import base64
        import re
        import os
        
        # VAKA FORMU ≈ûABLONUNU Y√úKLE
        backend_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        template_path = os.path.join(backend_dir, "templates", "VAKA_FORMU_TEMPLATE.xlsx")
        
        if os.path.exists(template_path):
            wb = load_workbook(template_path)
            ws = wb.active
            logger.info(f"≈ûablon y√ºklendi: {template_path}")
        else:
            # ≈ûablon yoksa bo≈ü workbook
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Vaka Formu"
            logger.warning(f"≈ûablon bulunamadƒ±, bo≈ü olu≈üturuluyor: {template_path}")
        
        # Logo - #VALUE! hatasƒ±nƒ± temizle ve logoyu d√ºzg√ºn ekle
        # Logo - √ñnce dosyadan, yoksa mapping'den y√ºkle
        logo_cell = logo_info.get("cell", "A1") if logo_info else "A1"
        
        # Logo h√ºcrelerindeki i√ßeriƒüi temizle (A1:C6 arasƒ± - sadece logo alanƒ±)
        # D s√ºtunundan ba≈ülayan metinlere dokunma
        for row in range(1, 7):
            for col in range(1, 4):  # A-C s√ºtunlarƒ± (D dahil deƒüil)
                try:
                    cell = ws.cell(row=row, column=col)
                    cell.value = None
                except:
                    pass
        
        # √ñnce dosyadan logo y√ºkle
        logo_path = os.path.join(backend_dir, "templates", "healmedy_logo.png")
        logo_added = False
        
        if os.path.exists(logo_path):
            try:
                img = XLImage(logo_path)
                img.width = 200   # Daha b√ºy√ºk ve net
                img.height = 75   # Orantƒ±lƒ±
                img.anchor = logo_cell
                ws.add_image(img, logo_cell)
                logo_added = True
            except Exception as e:
                logger.warning(f"Logo dosyadan eklenemedi: {e}")
        
        # Dosyadan eklenemezse mapping'deki base64'√º dene
        if not logo_added and logo_info and logo_info.get("url"):
            try:
                logo_url = logo_info["url"]
                if logo_url.startswith("data:image"):
                    header, encoded = logo_url.split(",", 1)
                    logo_data = base64.b64decode(encoded)
                    from io import BytesIO
                    img_buffer = BytesIO(logo_data)
                    img = XLImage(img_buffer)
                    img.width = 200   # Daha b√ºy√ºk ve net
                    img.height = 75   # Orantƒ±lƒ±
                    img.anchor = logo_cell
                    ws.add_image(img, logo_cell)
            except Exception as e:
                logger.warning(f"Logo base64'ten eklenemedi: {e}")
        
        # Mapping'leri uygula - ≈üablondaki h√ºcrelere deƒüer yaz
        for cell_address, field_key in flat_mappings.items():
            if field_key == "__LOGO__":
                # Logo h√ºcresini de temizle
                match = re.match(r'^([A-Z]+)(\d+)$', cell_address.upper())
                if match:
                    try:
                        ws[cell_address] = None
                    except:
                        pass
                continue
            
            value = get_case_field_value(case_data, field_key)
            
            # ƒ∞mza alanƒ± ise ve base64 g√∂r√ºnt√º varsa, g√∂r√ºnt√º olarak ekle
            if '_imza' in field_key or '_signature' in field_key:
                # inline_consents'tan ger√ßek imza verisini al
                sig_key_map = {
                    'sig.hekim_prm_imza': 'doctor_paramedic_signature',
                    'sig.saglik_per_imza': 'health_personnel_signature',
                    'sig.sofor_teknisyen_imza': 'driver_pilot_signature',
                    'sig.teslim_alan_imza': 'receiver_signature',
                    'sig.hasta_yakin_imza': 'patient_info_consent_signature',
                    'sig.hasta_reddi_imza': 'patient_rejection_signature',
                    'sig.hastane_reddi_imza': 'hospital_rejection_doctor_signature',
                }
                
                inline_key = sig_key_map.get(field_key)
                if inline_key:
                    sig_data = case_data.get('inline_consents', {}).get(inline_key)
                    if sig_data and isinstance(sig_data, str) and sig_data.startswith('data:image'):
                        try:
                            # Base64'√º g√∂r√ºnt√ºye √ßevir
                            header, encoded = sig_data.split(",", 1)
                            sig_bytes = base64.b64decode(encoded)
                            sig_buffer = BytesIO(sig_bytes)
                            sig_img = XLImage(sig_buffer)
                            sig_img.width = 150   # B√ºy√ºt√ºld√º
                            sig_img.height = 55   # B√ºy√ºt√ºld√º
                            ws.add_image(sig_img, cell_address)
                            continue  # G√∂r√ºnt√º eklendi, metin ekleme
                        except Exception as e:
                            logger.warning(f"ƒ∞mza g√∂r√ºnt√ºs√º eklenemedi {cell_address}: {e}")
                            # Fallback: ‚úì yaz
                            value = '‚úì'
            
            match = re.match(r'^([A-Z]+)(\d+)$', cell_address.upper())
            if match:
                try:
                    ws[cell_address] = value
                except Exception as e:
                    logger.warning(f"H√ºcre yazma hatasƒ± {cell_address}: {e}")
        
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        case_number = case_doc.get("case_number", case_id[:8])
        date_str = get_turkey_time().strftime("%Y-%m-%d")
        filename = f"VAKA_FORMU_{case_number}_{date_str}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
        
    except Exception as e:
        logger.error(f"Mapped Excel export hatasƒ±: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Excel olu≈üturma hatasƒ±: {str(e)}")


@router.get("/{case_id}/export-pdf-debug")
async def export_case_pdf_debug(case_id: str, request: Request):
    """PDF export debug - t√ºm verileri g√∂ster"""
    user = await get_current_user(request)
    
    case_doc = await cases_collection.find_one({"_id": case_id})
    if not case_doc:
        return {"error": "Case not found"}
    
    medical_form = case_doc.get("medical_form") or {}
    extended_form = medical_form.get("extended_form") or {}
    assigned_team = case_doc.get("assigned_team") or {}
    vehicle_info = case_doc.get("vehicle_info") or {}
    mf_vehicle_info = medical_form.get("vehicle_info") or {}  # medical_form i√ßindeki vehicle_info
    mf_time_info = medical_form.get("time_info") or {}  # medical_form i√ßindeki time_info
    mf_cpr_data = medical_form.get("cpr_data") or {}  # medical_form i√ßindeki cpr_data
    clinical_obs = medical_form.get("clinical_obs") or {}
    
    # √ñnemli alanlarƒ± √ßƒ±kar
    return {
        "case_id": case_id,
        "case_number": case_doc.get("case_number"),
        
        # Plaka kaynaklarƒ±
        "plaka_sources": {
            "assigned_team.vehicle": assigned_team.get("vehicle"),  # EN DOƒûRU KAYNAK
            "vehicle_info.plate": vehicle_info.get("plate"),
        },
        
        # KM kaynaklarƒ± - medical_form.vehicle_info i√ßinde
        "km_sources": {
            "mf_vehicle_info.startKm": mf_vehicle_info.get("startKm"),  # EN DOƒûRU KAYNAK
            "mf_vehicle_info.endKm": mf_vehicle_info.get("endKm"),      # EN DOƒûRU KAYNAK
        },
        
        # Saat bilgileri - medical_form.time_info i√ßinde
        "time_sources": {
            "mf_time_info.callTime": mf_time_info.get("callTime"),  # EN DOƒûRU KAYNAK
            "mf_time_info.arrivalTime": mf_time_info.get("arrivalTime"),
            "mf_time_info.departureTime": mf_time_info.get("departureTime"),
        },
        
        # CPR verileri - medical_form.cpr_data i√ßinde
        "cpr_sources": {
            "mf_cpr_data.cprBy": mf_cpr_data.get("cprBy"),  # EN DOƒûRU KAYNAK
            "mf_cpr_data.cprStart": mf_cpr_data.get("cprStart"),
            "mf_cpr_data.cprEnd": mf_cpr_data.get("cprEnd"),
            "mf_cpr_data.cprReason": mf_cpr_data.get("cprReason"),
        },
        
        # Kan ≈üekeri - extended_form i√ßinde
        "bloodSugar_sources": {
            "extended_form.bloodSugar": extended_form.get("bloodSugar"),  # EN DOƒûRU KAYNAK
            "clinical_obs.blood_sugar": clinical_obs.get("blood_sugar"),
        },
        
        # Nakledilen hastane
        "transfer_hospital": extended_form.get("transferHospital"),  # EN DOƒûRU KAYNAK
        
        # Triyaj kodu
        "triage_code": extended_form.get("triageCode"),  # EN DOƒûRU KAYNAK
        
        # Notes/Aciklamalar
        "notes": medical_form.get("notes"),  # EN DOƒûRU KAYNAK
        
        # Assigned team
        "assigned_team": assigned_team,
        
        # T√ºm medical_form.vehicle_info
        "mf_vehicle_info": mf_vehicle_info,
        
        # T√ºm medical_form.time_info
        "mf_time_info": mf_time_info,
        
        # T√ºm medical_form.cpr_data
        "mf_cpr_data": mf_cpr_data,
        
        # extended_form i√ßindeki t√ºm veriler
        "extended_form": extended_form,
    }


@router.get("/{case_id}/export-pdf-mapped")
async def export_case_pdf_with_mapping(case_id: str, request: Request):
    """Vakayƒ± Vaka Form Mapping kullanarak PDF olarak export et (Tek Sayfa)"""
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment
    from io import BytesIO
    import base64
    import re
    import os
    import subprocess
    import tempfile
    import uuid as uuid_module
    
    user = await get_current_user(request)
    logger.info(f"PDF export ba≈ülƒ±yor: case_id={case_id}")
    
    case_doc = await cases_collection.find_one({"_id": case_id})
    if not case_doc:
        raise HTTPException(status_code=404, detail="Vaka bulunamadƒ±")
    logger.info("Vaka bulundu")
    
    mapping_doc = await db.vaka_form_mappings.find_one({"_id": "default"})
    if not mapping_doc or not mapping_doc.get("flat_mappings"):
        raise HTTPException(status_code=404, detail="Vaka form mapping bulunamadƒ±. √ñnce mapping olu≈üturun.")
    logger.info(f"Mapping bulundu: {len(mapping_doc.get('flat_mappings', {}))} h√ºcre")
    
    flat_mappings = mapping_doc.get("flat_mappings", {})
    logo_info = mapping_doc.get("logo", {})
    
    # Medical form verilerini al
    medical_form = case_doc.get("medical_form", {})
    case_data = build_export_case_data(case_doc, medical_form)
    logger.info("Case data hazƒ±rlandƒ±")
    
    try:
        backend_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        template_path = os.path.join(backend_dir, "templates", "VAKA_FORMU_TEMPLATE.xlsx")
        logger.info(f"Template path: {template_path}, exists: {os.path.exists(template_path)}")
        
        if os.path.exists(template_path):
            wb = load_workbook(template_path)
            ws = wb.active
        else:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Vaka Formu"
        
        # Logo - √ñnce dosyadan, yoksa mapping'den y√ºkle
        logo_cell = logo_info.get("cell", "A1") if logo_info else "A1"
        
        # Logo h√ºcrelerindeki t√ºm i√ßeriƒüi temizle (A1:C6 arasƒ± - sadece logo alanƒ±)
        # D s√ºtunundan ba≈ülayan metinlere dokunma
        for row in range(1, 7):  # 1-6 satƒ±rlarƒ±
            for col in range(1, 4):  # A-C s√ºtunlarƒ± (D dahil deƒüil)
                try:
                    cell = ws.cell(row=row, column=col)
                    cell.value = None
                except:
                    pass
        
        # √ñnce dosyadan logo y√ºkle
        logo_path = os.path.join(backend_dir, "templates", "healmedy_logo.png")
        logo_added = False
        
        if os.path.exists(logo_path):
            try:
                img = XLImage(logo_path)
                # Logo boyutlarƒ± - orantƒ±lƒ± ve g√∂r√ºn√ºr (A1:C5 alanƒ±na sƒ±ƒüacak)
                # A4 Portrait i√ßin uygun boyut
                img.width = 200   # ~4 s√ºtun geni≈üliƒüi (daha b√ºy√ºk ve net)
                img.height = 75   # ~5 satƒ±r y√ºksekliƒüi (orantƒ±lƒ±)
                
                # Anchor ile konumlandƒ±r - sol √ºst k√∂≈üeye sabitle
                from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
                img.anchor = 'A1'  # A1 h√ºcresine sabitle
                
                ws.add_image(img, "A1")  # Sol √ºst k√∂≈üeye sabitle
                logger.info(f"Logo dosyadan eklendi: A1, {img.width}x{img.height}")
                logo_added = True
            except Exception as e:
                logger.warning(f"Logo dosyadan eklenemedi: {e}")
        
        # Dosyadan eklenemezse mapping'deki base64'√º dene
        if not logo_added and logo_info and logo_info.get("url"):
            try:
                logo_url = logo_info["url"]
                if logo_url.startswith("data:image"):
                    header, encoded = logo_url.split(",", 1)
                    logo_data = base64.b64decode(encoded)
                    img_buffer = BytesIO(logo_data)
                    img = XLImage(img_buffer)
                    img.width = 200   # Daha b√ºy√ºk ve net
                    img.height = 75   # Orantƒ±lƒ±
                    img.anchor = 'A1'
                    ws.add_image(img, "A1")
                    logger.info(f"Logo base64'ten eklendi: A1, {img.width}x{img.height}")
            except Exception as e:
                logger.warning(f"Logo base64'ten eklenemedi: {e}")
        
        # Mapping'leri uygula
        logger.info(f"Mapping uygulanƒ±yor: {len(flat_mappings)} h√ºcre")
        for cell_address, field_key in flat_mappings.items():
            try:
                if field_key == "__LOGO__":
                    # Logo h√ºcresini de temizle
                    match = re.match(r'^([A-Z]+)(\d+)$', cell_address.upper())
                    if match:
                        try:
                            ws[cell_address] = None
                        except:
                            pass
                    continue
                
                value = get_case_field_value(case_data, field_key)
            except Exception as field_err:
                logger.warning(f"Alan deƒüeri alƒ±namadƒ±: {field_key} -> {cell_address}: {field_err}")
                value = ""
            
            # ƒ∞mza alanƒ± ise ve base64 g√∂r√ºnt√º varsa, g√∂r√ºnt√º olarak ekle
            if '_imza' in field_key or '_signature' in field_key:
                sig_key_map = {
                    'sig.hekim_prm_imza': 'doctor_paramedic_signature',
                    'sig.saglik_per_imza': 'health_personnel_signature',
                    'sig.sofor_teknisyen_imza': 'driver_pilot_signature',
                    'sig.teslim_alan_imza': 'receiver_signature',
                    'sig.hasta_yakin_imza': 'patient_info_consent_signature',
                    'sig.hasta_reddi_imza': 'patient_rejection_signature',
                    'sig.hastane_reddi_imza': 'hospital_rejection_doctor_signature',
                }
                
                inline_key = sig_key_map.get(field_key)
                if inline_key:
                    sig_data = case_data.get('inline_consents', {}).get(inline_key)
                    if sig_data and isinstance(sig_data, str) and sig_data.startswith('data:image') and ',' in sig_data:
                        try:
                            header, encoded = sig_data.split(",", 1)
                            if encoded:  # Base64 verisi var mƒ± kontrol et
                                sig_bytes = base64.b64decode(encoded)
                                sig_buffer = BytesIO(sig_bytes)
                                sig_img = XLImage(sig_buffer)
                                # ƒ∞mza boyutlarƒ± - b√ºy√ºk (yazƒ±cƒ±da net g√∂r√ºns√ºn)
                                sig_img.width = 150   # ~5cm geni≈ülik
                                sig_img.height = 55   # ~1.7cm y√ºkseklik
                                ws.add_image(sig_img, cell_address)
                                continue
                        except Exception as e:
                            logger.warning(f"ƒ∞mza g√∂r√ºnt√ºs√º eklenemedi {cell_address}: {e}")
                            value = '‚úì'
                    elif sig_data:
                        # ƒ∞mza var ama g√∂r√ºnt√º olarak eklenemedi
                        value = '‚úì'
            
            match = re.match(r'^([A-Z]+)(\d+)$', cell_address.upper())
            if match:
                try:
                    cell = ws[cell_address]
                    cell.value = value
                    
                    from openpyxl.styles import Alignment
                    
                    # T√ºm h√ºcrelere text wrap uygula (uzun i√ßerik varsa satƒ±ra sƒ±ƒümasƒ± i√ßin)
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                    
                    # Uzun i√ßerik i√ßin satƒ±r y√ºksekliƒüini artƒ±r
                    if value and len(str(value)) > 25:
                        row_num = int(match.group(2))
                        
                        # S√ºtun geni≈üliƒüini hesapla (yakla≈üƒ±k karakter sayƒ±sƒ±)
                        col_letter = match.group(1)
                        col_width = ws.column_dimensions[col_letter].width or 10
                        chars_per_line = int(col_width * 1.2)  # Yakla≈üƒ±k
                        
                        # Gerekli satƒ±r sayƒ±sƒ±nƒ± hesapla
                        content_lines = max(1, len(str(value)) // max(chars_per_line, 10) + 1)
                        
                        # Mevcut y√ºksekliƒüi al
                        current_height = ws.row_dimensions[row_num].height or 15
                        
                        # Yeni y√ºkseklik hesapla (satƒ±r ba≈üƒ±na ~14px)
                        needed_height = content_lines * 14
                        new_height = max(current_height, needed_height, 15)
                        
                        # Maksimum sƒ±nƒ±r koy (√ßok uzun i√ßerikler i√ßin)
                        ws.row_dimensions[row_num].height = min(new_height, 100)
                        
                except Exception as e:
                    logger.warning(f"H√ºcre yazma hatasƒ± {cell_address}: {e}")
        
        # SAYFA AYARLARI: A4 Dikey (Portrait), TEK SAYFAYA SIƒûDIR
        # ≈ûablonun kendi boyutlandƒ±rmasƒ±nƒ± kullan - V4 ≈üablonu A4'e uygun hazƒ±rlandƒ±
        try:
            from openpyxl.worksheet.properties import PageSetupProperties
            
            # Sayfa ayarlarƒ± - A4 Portrait
            ws.page_setup.orientation = 'portrait'  # Dikey
            ws.page_setup.paperSize = 9  # A4 (9 = A4)
            ws.page_setup.fitToWidth = 1  # 1 sayfa geni≈üliƒüine sƒ±ƒüdƒ±r
            ws.page_setup.fitToHeight = 1  # 1 sayfa y√ºksekliƒüine sƒ±ƒüdƒ±r
            ws.page_setup.fitToPage = True  # FitToPage modunu etkinle≈ütir
            
            # Scale ayarƒ±nƒ± kullanma - fitToPage ile otomatik √∂l√ßeklenir
            # ws.page_setup.scale = None  # √ñl√ßek kullanƒ±lmƒ±yor, FitToPage aktif
            
            # Sheet properties
            if ws.sheet_properties.pageSetUpPr is None:
                ws.sheet_properties.pageSetUpPr = PageSetupProperties()
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            
            # Print area: T√ºm ≈üablon alanƒ± (≈üablondaki max satƒ±r/s√ºtun)
            from openpyxl.utils import get_column_letter
            max_row = ws.max_row or 79
            max_col = ws.max_column or 20
            print_area = f"A1:{get_column_letter(max_col)}{max_row}"
            ws.print_area = print_area
            
            # Kenar bo≈üluklarƒ± - minimum (sayfayƒ± tam kullan)
            ws.page_margins.left = 0.1    # ~0.25cm
            ws.page_margins.right = 0.1
            ws.page_margins.top = 0.1
            ws.page_margins.bottom = 0.1
            ws.page_margins.header = 0
            ws.page_margins.footer = 0
            
            # ≈ûablonun satƒ±r y√ºksekliklerini KORU - deƒüi≈ütirme
            # V4 ≈üablonu zaten A4'e uygun hazƒ±rlandƒ±
            
            logger.info(f"Sayfa ayarlarƒ±: A4 Portrait, FitToPage=1x1, PrintArea={print_area}, MaxRow={max_row}")
        except Exception as e:
            logger.warning(f"Sayfa ayarlarƒ± yapƒ±lamadƒ±: {e}")
        
        # Ge√ßici Excel dosyasƒ± olu≈ütur
        temp_dir = os.path.join(backend_dir, "temp")
        os.makedirs(temp_dir, exist_ok=True)
        
        temp_xlsx = os.path.join(temp_dir, f"case_{case_id}_{uuid_module.uuid4().hex[:8]}.xlsx")
        wb.save(temp_xlsx)
        
        # LibreOffice ile PDF'e d√∂n√º≈üt√ºr (A4 tek sayfa)
        try:
            # √ñnce ODS'ye √ßevir, sonra PDF'e (sayfa ayarlarƒ± daha iyi korunuyor)
            # ODS d√∂n√º≈ü√ºm√º
            result1 = subprocess.run([
                'libreoffice', '--headless', '--convert-to', 'ods',
                '--outdir', temp_dir, temp_xlsx
            ], capture_output=True, timeout=60, text=True)
            
            temp_ods = temp_xlsx.replace('.xlsx', '.ods')
            
            # ODS dosyasƒ±nƒ± a√ß ve sayfa ayarlarƒ±nƒ± d√ºzelt
            if os.path.exists(temp_ods):
                try:
                    # ODS dosyasƒ±nƒ± zipfile olarak a√ß ve styles.xml'i d√ºzenle
                    import zipfile
                    import xml.etree.ElementTree as ET
                    
                    # ODS'yi a√ß
                    with zipfile.ZipFile(temp_ods, 'r') as zf:
                        content = {name: zf.read(name) for name in zf.namelist()}
                    
                    # styles.xml'i d√ºzenle - A4 Dikey (Portrait) ve tek sayfaya sƒ±ƒüdƒ±r
                    if 'styles.xml' in content:
                        styles_xml = content['styles.xml'].decode('utf-8')
                        
                        # Sayfa ayarlarƒ±nƒ± deƒüi≈ütir
                        # fo:page-width="21cm" fo:page-height="29.7cm" (A4 portrait)
                        # style:scale-to-pages="1"
                        import re
                        
                        # A4 portrait boyutlarƒ±
                        styles_xml = re.sub(r'fo:page-width="[^"]*"', 'fo:page-width="21cm"', styles_xml)
                        styles_xml = re.sub(r'fo:page-height="[^"]*"', 'fo:page-height="29.7cm"', styles_xml)
                        
                        # Orientation'ƒ± portrait yap
                        styles_xml = re.sub(r'style:print-orientation="[^"]*"', 'style:print-orientation="portrait"', styles_xml)
                        
                        # Kenar bo≈üluklarƒ±nƒ± minimize et (A4'√º doldurmak i√ßin)
                        styles_xml = re.sub(r'fo:margin-top="[^"]*"', 'fo:margin-top="0.25cm"', styles_xml)
                        styles_xml = re.sub(r'fo:margin-bottom="[^"]*"', 'fo:margin-bottom="0.25cm"', styles_xml)
                        styles_xml = re.sub(r'fo:margin-left="[^"]*"', 'fo:margin-left="0.25cm"', styles_xml)
                        styles_xml = re.sub(r'fo:margin-right="[^"]*"', 'fo:margin-right="0.25cm"', styles_xml)
                        
                        # Tek sayfaya sƒ±ƒüdƒ±r (1x1 sayfa)
                        styles_xml = re.sub(r'style:scale-to="[^"]*"', '', styles_xml)
                        styles_xml = re.sub(r'style:scale-to-pages="[^"]*"', '', styles_xml)
                        
                        # FitToPage: 1 sayfa geni≈ülik x 1 sayfa y√ºkseklik
                        if 'style:scale-to-X' not in styles_xml:
                            styles_xml = styles_xml.replace(
                                'style:print-orientation="portrait"',
                                'style:print-orientation="portrait" style:scale-to-X="1" style:scale-to-Y="1"'
                            )
                        
                        content['styles.xml'] = styles_xml.encode('utf-8')
                    
                    # Yeni ODS dosyasƒ± olu≈ütur
                    temp_ods_fixed = temp_ods.replace('.ods', '_fixed.ods')
                    with zipfile.ZipFile(temp_ods_fixed, 'w', zipfile.ZIP_DEFLATED) as zf:
                        for name, data in content.items():
                            zf.writestr(name, data)
                    
                    # Eski ODS'yi sil, yenisini kullan
                    os.remove(temp_ods)
                    os.rename(temp_ods_fixed, temp_ods)
                    
                except Exception as e:
                    logger.warning(f"ODS d√ºzenleme hatasƒ±: {e}")
            
            # Kaynak dosya (ODS veya XLSX)
            source_file = temp_ods if os.path.exists(temp_ods) else temp_xlsx
            
            # PDF'e d√∂n√º≈üt√ºr
            result2 = subprocess.run([
                'libreoffice', '--headless', '--convert-to', 'pdf',
                '--outdir', temp_dir, source_file
            ], capture_output=True, timeout=60, text=True)
            
            pdf_path = source_file.replace('.ods', '.pdf').replace('.xlsx', '.pdf')
            
            if os.path.exists(pdf_path):
                # PDF'i oku
                with open(pdf_path, 'rb') as f:
                    pdf_content = f.read()
                
                # Temizlik
                for f_path in [temp_xlsx, temp_ods, pdf_path]:
                    try:
                        if os.path.exists(f_path):
                            os.remove(f_path)
                    except:
                        pass
                
                case_number = case_doc.get("case_number", case_id[:8])
                date_str = get_turkey_time().strftime("%Y-%m-%d")
                filename = f"VAKA_FORMU_{case_number}_{date_str}.pdf"
                
                from io import BytesIO
                return StreamingResponse(
                    BytesIO(pdf_content),
                    media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'}
                )
            else:
                error_msg = result2.stderr if result2.stderr else result2.stdout
                logger.error(f"PDF olu≈üturulamadƒ±. LibreOffice √ßƒ±ktƒ±sƒ±: {error_msg}")
                for f_path in [temp_xlsx, temp_ods]:
                    try:
                        if os.path.exists(f_path):
                            os.remove(f_path)
                    except:
                        pass
                raise HTTPException(status_code=500, detail=f"PDF d√∂n√º≈üt√ºrme ba≈üarƒ±sƒ±z: {error_msg[:200]}")
                
        except subprocess.TimeoutExpired:
            os.remove(temp_xlsx)
            raise HTTPException(status_code=500, detail="PDF d√∂n√º≈üt√ºrme zaman a≈üƒ±mƒ±")
        except FileNotFoundError:
            # LibreOffice yok - Excel d√∂nd√ºr
            logger.warning("LibreOffice bulunamadƒ±, Excel d√∂nd√ºr√ºl√ºyor")
            with open(temp_xlsx, 'rb') as f:
                excel_content = f.read()
            os.remove(temp_xlsx)
            
            case_number = case_doc.get("case_number", case_id[:8])
            date_str = get_turkey_time().strftime("%Y-%m-%d")
            filename = f"VAKA_FORMU_{case_number}_{date_str}.xlsx"
            
            from io import BytesIO
            return StreamingResponse(
                BytesIO(excel_content),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'}
            )
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        logger.error(f"PDF export hatasƒ±: {str(e)}\nTraceback:\n{error_trace}")
        # Hata detaylarƒ±nƒ± response'a ekle (debug i√ßin)
        error_detail = {
            "error": str(e),
            "type": type(e).__name__,
            "trace": error_trace[-500:] if len(error_trace) > 500 else error_trace
        }
        raise HTTPException(status_code=500, detail=f"PDF olu≈üturma hatasƒ±: {type(e).__name__}: {str(e)[:300]}")


# ============================================================================
# VAKA Sƒ∞LME - OPERASYON M√úD√úR√ú YETKƒ∞Sƒ∞
# ============================================================================

@router.delete("/{case_id}")
async def delete_case(case_id: str, request: Request):
    """
    Vakayƒ± sistemden sil
    Sadece Operasyon M√ºd√ºr√º ve Merkez Ofis yetkilidir
    
    Bu endpoint eski vakalarƒ± temizlemek i√ßin kullanƒ±lƒ±r
    """
    # Sadece operasyon_muduru ve merkez_ofis silebilir
    user = await require_roles(["operasyon_muduru", "merkez_ofis"])(request)
    
    # Vakayƒ± bul
    case_doc = await cases_collection.find_one({"_id": case_id})
    if not case_doc:
        raise HTTPException(status_code=404, detail="Vaka bulunamadƒ±")
    
    case_number = case_doc.get("case_number", case_id[:8])
    
    # Silme i≈ülemi
    result = await cases_collection.delete_one({"_id": case_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=500, detail="Silme i≈ülemi ba≈üarƒ±sƒ±z oldu")
    
    logger.info(f"Vaka silindi: {case_number} (ID: {case_id}) - Silen: {user.name} ({user.role})")
    
    return {
        "message": f"Vaka '{case_number}' ba≈üarƒ±yla silindi",
        "deleted_case_id": case_id,
        "deleted_case_number": case_number,
        "deleted_by": user.id,
        "deleted_by_name": user.name
    }