from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import FileResponse, StreamingResponse
from datetime import datetime
import os
import logging

from services.excel_to_pdf_with_data import excel_form_to_pdf_with_data
from services.template_pdf_generator import generate_pdf_from_template
from services.libreoffice_pdf import generate_case_pdf_with_libreoffice, check_libreoffice_installed
from services.vaka_form_mapping import get_cell_mapping_for_display, VAKA_FORM_CELL_MAPPING, CHECKBOX_MAPPINGS
from auth_utils import get_current_user
from database import cases_collection, pdf_templates_collection, db

router = APIRouter(prefix="/pdf", tags=["pdf"])
logger = logging.getLogger(__name__)

# LibreOffice kurulu mu kontrol et
LIBREOFFICE_AVAILABLE = check_libreoffice_installed()
logger.info(f"LibreOffice available: {LIBREOFFICE_AVAILABLE}")


@router.get("/case/{case_id}")
async def generate_case_pdf(case_id: str, request: Request):
    """Generate PDF for a specific case using LibreOffice (if available) or fallback"""
    
    # Authenticate user
    user = await get_current_user(request)
    
    # Fetch case data from database
    case = await cases_collection.find_one({"_id": case_id})
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    
    # Get path to templates directory
    backend_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    excel_template = os.path.join(backend_dir, "templates", "VAKA_FORMU_TEMPLATE.xlsx")
    
    if not os.path.exists(excel_template):
        logger.error(f"Excel template not found at: {excel_template}")
        raise HTTPException(status_code=500, detail=f"Excel template not found")
    
    # Prepare form data from case
    form_data = case.get('form_data', {})
    
    # Generate filename
    filename = f"Ambulans_Vaka_Formu_{case.get('case_number', case_id)}.pdf"
    
    try:
        # Use LibreOffice if available (better quality)
        if LIBREOFFICE_AVAILABLE:
            logger.info(f"Using LibreOffice for case {case_id}")
            result_path = generate_case_pdf_with_libreoffice(
                template_path=excel_template,
                case_data=case,
                form_data=form_data,
                output_filename=f"case_{case_id}.pdf"
            )
        else:
            # Fallback to ReportLab method
            logger.info(f"LibreOffice not available, using fallback for case {case_id}")
            temp_dir = os.path.join(backend_dir, "temp")
            os.makedirs(temp_dir, exist_ok=True)
            pdf_path = os.path.join(temp_dir, f"case_{case_id}.pdf")
            
            result_path = excel_form_to_pdf_with_data(
                excel_path=excel_template,
                pdf_path=pdf_path,
                case_data=case,
                form_data=form_data
            )
        
        # Return PDF as download
        return FileResponse(
            result_path,
            media_type="application/pdf",
            filename=filename,
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
        
    except Exception as e:
        logger.error(f"Error generating PDF for case {case_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")


@router.post("/case/{case_id}/with-form-data")
async def generate_case_pdf_with_form_data(
    case_id: str, 
    form_data: dict,
    request: Request
):
    """Generate PDF for a specific case with additional form data from frontend"""
    
    # Authenticate user
    user = await get_current_user(request)
    
    # Fetch case data from database
    case = await cases_collection.find_one({"_id": case_id})
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    
    # Get path to templates directory
    backend_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    excel_template = os.path.join(backend_dir, "templates", "VAKA_FORMU_TEMPLATE.xlsx")
    
    if not os.path.exists(excel_template):
        logger.error(f"Excel template not found at: {excel_template}")
        raise HTTPException(status_code=500, detail=f"Excel template not found")
    
    # Generate filename
    filename = f"Ambulans_Vaka_Formu_{case.get('case_number', case_id)}.pdf"
    
    try:
        # Use LibreOffice if available (better quality)
        if LIBREOFFICE_AVAILABLE:
            logger.info(f"Using LibreOffice for case {case_id} with form data")
            result_path = generate_case_pdf_with_libreoffice(
                template_path=excel_template,
                case_data=case,
                form_data=form_data,
                output_filename=f"case_{case_id}.pdf"
            )
        else:
            # Fallback to ReportLab method
            logger.info(f"LibreOffice not available, using fallback for case {case_id}")
            temp_dir = os.path.join(backend_dir, "temp")
            os.makedirs(temp_dir, exist_ok=True)
            pdf_path = os.path.join(temp_dir, f"case_{case_id}.pdf")
            
            result_path = excel_form_to_pdf_with_data(
                excel_path=excel_template,
                pdf_path=pdf_path,
                case_data=case,
                form_data=form_data
            )
        
        # Return PDF as download
        return FileResponse(
            result_path,
            media_type="application/pdf",
            filename=filename,
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
        
    except Exception as e:
        logger.error(f"Error generating PDF for case {case_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")


# ============== VAKA FORM MAPPING ENDPOINTS ==============

@router.get("/vaka-form-template-cells")
async def get_vaka_form_template_cells(request: Request):
    """VAKA_FORMU_TEMPLATE.xlsx şablonundaki hücre değerlerini döndürür"""
    user = await get_current_user(request)
    
    template_path = os.path.join(os.path.dirname(__file__), '..', 'templates', 'VAKA_FORMU_TEMPLATE.xlsx')
    
    if not os.path.exists(template_path):
        return {"cells": {}, "error": "Şablon dosyası bulunamadı"}
    
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        
        wb = load_workbook(template_path, data_only=True)
        ws = wb.active
        
        cells = {}
        for row in range(1, min(ws.max_row + 1, 101)):  # Max 100 satır
            for col in range(1, min(ws.max_column + 1, 36)):  # Max 35 sütun
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    address = f"{get_column_letter(col)}{row}"
                    cells[address] = str(cell.value)
        
        return {"cells": cells, "max_row": ws.max_row, "max_column": ws.max_column}
    except Exception as e:
        logger.error(f"Şablon okuma hatası: {e}")
        return {"cells": {}, "error": str(e)}


@router.get("/vaka-form-mapping")
async def get_vaka_form_cell_mapping(request: Request):
    """Vaka formu Excel hücre eşlemelerini döndürür - FormTemplates sayfası için"""
    user = await get_current_user(request)
    
    # Önce veritabanından özelleştirilmiş mapping'i kontrol et
    custom_mapping = await db.vaka_form_mappings.find_one({"_id": "default"})
    
    if custom_mapping:
        return {
            "cell_mappings": custom_mapping.get("cell_mappings", {}),
            "checkbox_mappings": custom_mapping.get("checkbox_mappings", {}),
            "flat_mappings": custom_mapping.get("flat_mappings", {}),
            "logo": custom_mapping.get("logo", {}),
            "total_cells": custom_mapping.get("total_cells", 0),
            "total_checkboxes": custom_mapping.get("total_checkboxes", 0),
            "is_custom": True
        }
    
    # Varsayılan mapping'i döndür
    result = get_cell_mapping_for_display()
    result["is_custom"] = False
    result["flat_mappings"] = {}
    result["logo"] = {}
    return result


@router.put("/vaka-form-mapping/{cell}")
async def update_vaka_form_cell_mapping(cell: str, data: dict, request: Request):
    """Belirli bir hücrenin eşlemesini günceller"""
    user = await get_current_user(request)
    
    # Mevcut mapping'i al veya oluştur
    existing = await db.vaka_form_mappings.find_one({"_id": "default"})
    
    if not existing:
        # Varsayılan mapping'i kopyala
        default_mapping = get_cell_mapping_for_display()
        existing = {
            "_id": "default",
            "cell_mappings": default_mapping["cell_mappings"],
            "checkbox_mappings": default_mapping["checkbox_mappings"],
            "total_cells": default_mapping["total_cells"],
            "total_checkboxes": default_mapping["total_checkboxes"]
        }
    
    # Hücreyi bul ve güncelle
    updated = False
    for section, cells in existing["cell_mappings"].items():
        for i, cell_info in enumerate(cells):
            if cell_info["cell"] == cell:
                existing["cell_mappings"][section][i].update(data)
                updated = True
                break
        if updated:
            break
    
    if not updated:
        raise HTTPException(status_code=404, detail=f"Hücre bulunamadı: {cell}")
    
    # Veritabanına kaydet
    await db.vaka_form_mappings.replace_one(
        {"_id": "default"},
        existing,
        upsert=True
    )
    
    return {"message": "Hücre eşlemesi güncellendi", "cell": cell}


@router.post("/vaka-form-mapping/reset")
async def reset_vaka_form_mapping(request: Request):
    """Vaka form mapping'ini varsayılana sıfırla"""
    user = await get_current_user(request)
    
    await db.vaka_form_mappings.delete_one({"_id": "default"})
    
    return {"message": "Eşlemeler varsayılana sıfırlandı"}


@router.put("/vaka-form-mapping/bulk")
async def bulk_update_vaka_form_mapping(request: Request):
    """
    Tüm hücre eşlemelerini toplu güncelle (Görsel Editör için)
    Body: { mappings: { "A1": "fieldKey", ... }, logo: { url: "...", cell: "A1" } }
    """
    user = await get_current_user(request)
    data = await request.json()
    
    mappings = data.get("mappings", {})
    logo = data.get("logo", {})
    
    existing = await db.vaka_form_mappings.find_one({"_id": "default"})
    
    if not existing:
        existing = {
            "_id": "default",
            "cell_mappings": {},
            "checkbox_mappings": {},
            "total_cells": 0,
            "total_checkboxes": 0
        }
    
    existing["flat_mappings"] = mappings
    existing["logo"] = logo
    existing["total_cells"] = len(mappings)
    existing["updated_at"] = datetime.now()
    existing["updated_by"] = user.id
    
    await db.vaka_form_mappings.update_one(
        {"_id": "default"},
        {"$set": existing},
        upsert=True
    )
    
    return {
        "message": "Mapping kaydedildi",
        "total_mappings": len(mappings),
        "has_logo": bool(logo.get("url"))
    }