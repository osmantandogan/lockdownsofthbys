#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Enhanced Excel to PDF Generator with Case Data Population
Reads Excel template and generates PDF with populated case information
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime
import os


def excel_form_to_pdf_with_data(
    excel_path: str,
    pdf_path: str,
    case_data: dict = None,
    form_data: dict = None,
    sheet_name: str = "Sayfa1"
):
    """
    Generate PDF from Excel template with optional case data population
    
    Args:
        excel_path: Path to Excel template file (.xlsx)
        pdf_path: Output PDF path
        case_data: Dictionary with case information from backend
        form_data: Dictionary with form fields from frontend
        sheet_name: Excel sheet name (default: "Sayfa1")
    
    Returns:
        str: Path to generated PDF file
    """
    
    # Try to register Turkish character font
    font_name = "DejaVu"
    try:
        # Try to find DejaVuSans.ttf in current directory or system
        font_paths = [
            "DejaVuSans.ttf",
            "C:/Windows/Fonts/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "C:/Users/Serdar/Desktop/better-python-xlsx-shit/DejaVuSans.ttf"
        ]
        
        font_found = False
        for font_path in font_paths:
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont(font_name, font_path))
                font_found = True
                break
        
        if not font_found:
            print("Warning: DejaVuSans.ttf not found, using default font")
            font_name = "Helvetica"
    except Exception as e:
        print(f"Warning: Could not load DejaVu font: {e}, using Helvetica")
        font_name = "Helvetica"

    # Load Excel workbook
    wb = load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sayfa bulunamadı: {sheet_name}")
    ws = wb[sheet_name]

    # Populate case data into specific cells before generating PDF
    if case_data or form_data:
        case_data = case_data or {}
        form_data = form_data or {}
        
        # Helper to safely get nested dictionary values
        def safe_get(d, *keys, default=''):
            for key in keys:
                if isinstance(d, dict):
                    d = d.get(key, {})
                else:
                    return default
            return d if d else default
        
        # Format date
        date_str = form_data.get('date', '')
        if not date_str and case_data.get('created_at'):
            try:
                date_obj = datetime.fromisoformat(case_data['created_at'].replace('Z', '+00:00'))
                date_str = date_obj.strftime('%d.%m.%Y')
            except:
                date_str = ''
        
        # Format times
        call_time = form_data.get('callTime', '')
        if not call_time and case_data.get('created_at'):
            try:
                date_obj = datetime.fromisoformat(case_data['created_at'].replace('Z', '+00:00'))
                call_time = date_obj.strftime('%H:%M')
            except:
                call_time = ''
        
        # Build patient name
        patient_name = form_data.get('patientName', '')
        if not patient_name:
            first = safe_get(case_data, 'patient', 'name')
            last = safe_get(case_data, 'patient', 'surname')
            patient_name = f"{first} {last}".strip()
        
        # Map case data to cell locations (based on actual Excel template analysis)
        # Corrected to use the actual INPUT cells, not the label cells
        cell_mapping = {
            # Header section - ALL INPUTS ARE BELOW THE LABELS
            # ATN NO (label at T3, INPUT BELOW at T5)
            'T5': case_data.get('case_number', ''),  # ATN NO input
            
            # BAŞLANGIÇ KM (label at W3, INPUT BELOW at W5)
            'W5': form_data.get('startKm', ''),  # BAŞLANGIÇ KM input
            
            # BİTİŞ KM (label at Z3, INPUT BELOW at Z5)
            'Z5': form_data.get('endKm', ''),  # BİTİŞ KM input
            
            # İSTASYON section (labels at C9, C11, C13 - inputs are at E9, E11, E13)
            'E9': form_data.get('healmedyProtocol') or case_data.get('case_number', ''),  # PROTOKOL NO input
            'E11': date_str,  # TARİH input
            'E13': form_data.get('vehicleType') or safe_get(case_data, 'assigned_team', 'vehicle_id'),  # PLAKA input
            
            # SAATLER section (labels at G9-G13, inputs at J9-J13)
            'J9': call_time,  # ÇAĞRI SAATİ input
            'J10': form_data.get('arrivalTime', ''),  # OLAY YERİNE VARIŞ input
            'J11': form_data.get('arrivalTime', ''),  # HASTAYA VARIŞ input
            'J12': form_data.get('departureTime', ''),  # OLAY YERİNDEN AYRILIŞ input
            'J13': form_data.get('hospitalArrivalTime', ''),  # HASTANEYE VARIŞ input
            
            # HASTA BİLGİLERİ section (labels at L9, L10, L14 - inputs at N9, N10, N14)
            'N9': patient_name,  # ADI SOYADI input
            'N10': form_data.get('address') or safe_get(case_data, 'location', 'address'),  # ADRESİ input
            'N14': form_data.get('phone') or safe_get(case_data, 'patient', 'phone') or safe_get(case_data, 'caller', 'phone'),  # TELEFON input
            
            # CİNSİYET / YAŞ section (label at T13, input at U13 - this was correct!)
            'U13': form_data.get('age') or str(safe_get(case_data, 'patient', 'age')),  # YAŞ input
            
            # T.C. KİMLİK NO (label at R15, input at V15)
            'V15': form_data.get('tcNo') or safe_get(case_data, 'patient', 'tc_no'),  # T.C. KİMLİK NO input
            
            # KRONİK HASTALIKLAR (label at Y8, INPUT BELOW at Y9)
            'Y9': form_data.get('chronicDiseases', ''),  # KRONİK HASTALIKLAR input
            
            # HASTANIN ŞİKAYETİ (label at Y11, INPUT BELOW at Y12)
            'Y12': form_data.get('complaint') or safe_get(case_data, 'patient', 'complaint') or case_data.get('description', ''),  # ŞİKAYET input
        }
        
        # Write values to cells (handle merged cells by finding top-left cell)
        merged_cells_map = {}
        for mr in ws.merged_cells.ranges:
            # For each cell in the merge range, map it to the top-left cell
            for row in range(mr.min_row, mr.max_row + 1):
                for col in range(mr.min_col, mr.max_col + 1):
                    cell_coord = ws.cell(row=row, column=col).coordinate
                    top_left = ws.cell(row=mr.min_row, column=mr.min_col).coordinate
                    merged_cells_map[cell_coord] = top_left
        
        for cell_ref, value in cell_mapping.items():
            if value:
                try:
                    # If this is a merged cell, write to the top-left cell instead
                    target_cell = merged_cells_map.get(cell_ref, cell_ref)
                    ws[target_cell] = str(value)
                except Exception as e:
                    print(f"Warning: Could not write to cell {cell_ref}: {e}")

    # Find used range (values + merges)
    min_row_val = None
    max_row_val = 0
    min_col_val = None
    max_col_val = 0

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v not in (None, ""):
                if min_row_val is None or r < min_row_val:
                    min_row_val = r
                if min_col_val is None or c < min_col_val:
                    min_col_val = c
                if r > max_row_val:
                    max_row_val = r
                if c > max_col_val:
                    max_col_val = c

    # Include merged cells in range
    min_row_merge = None
    min_col_merge = None
    max_row_merge = 0
    max_col_merge = 0
    for mr in ws.merged_cells.ranges:
        if min_row_merge is None or mr.min_row < min_row_merge:
            min_row_merge = mr.min_row
        if min_col_merge is None or mr.min_col < min_col_merge:
            min_col_merge = mr.min_col
        if mr.max_row > max_row_merge:
            max_row_merge = mr.max_row
        if mr.max_col > max_col_merge:
            max_col_merge = mr.max_col

    if min_row_val is None:
        min_row_val = min_row_merge or 1
    if min_col_val is None:
        min_col_val = min_col_merge or 1

    used_min_row = min(filter(None, [min_row_val, min_row_merge]))
    used_min_col = min(filter(None, [min_col_val, min_col_merge]))
    used_max_row = max(max_row_val, max_row_merge)
    used_max_col = max(max_col_val, max_col_merge)

    # Get Excel column widths and scale to PDF
    excel_widths = []
    for c in range(used_min_col, used_max_col + 1):
        col_letter = get_column_letter(c)
        w = ws.column_dimensions[col_letter].width
        if w is None:
            w = 8.43
        excel_widths.append(w)

    total_excel_width = sum(excel_widths)

    page_width, page_height = landscape(A4)
    usable_width = page_width - 10  # margin: 5 + 5
    scale = usable_width / total_excel_width

    col_widths = [w * scale for w in excel_widths]

    # Style / font
    styles = getSampleStyleSheet()
    base_style = styles["Normal"]
    base_style.fontName = font_name
    base_style.fontSize = 5.5
    base_style.leading = 6.3

    # Read cells (only used range)
    data = []
    for r in range(used_min_row, used_max_row + 1):
        row_cells = []
        for c in range(used_min_col, used_max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                row_cells.append("")
            else:
                text = str(v)
                para = Paragraph(text.replace("\n", "<br/>"), base_style)
                row_cells.append(para)
        data.append(row_cells)

    table = Table(data, colWidths=col_widths)

    style = TableStyle(
        [
            ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ]
    )

    # Apply merges with offset
    for mr in ws.merged_cells.ranges:
        if mr.max_row < used_min_row or mr.min_row > used_max_row:
            continue
        if mr.max_col < used_min_col or mr.min_col > used_max_col:
            continue

        min_row = mr.min_row - used_min_row
        max_row_m = mr.max_row - used_min_row
        min_col = mr.min_col - used_min_col
        max_col_m = mr.max_col - used_min_col

        style.add("SPAN", (min_col, min_row), (max_col_m, max_row_m))

    table.setStyle(style)

    # Generate PDF
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=landscape(A4),
        leftMargin=5,
        rightMargin=5,
        topMargin=5,
        bottomMargin=5,
    )

    doc.build([table])
    
    return pdf_path


# Test function
if __name__ == "__main__":
    excel_path = "ambulans_vaka_formu.xlsx"
    pdf_path = "ambulans_vaka_formu_test.pdf"
    
    # Example case data
    test_case_data = {
        'case_number': 'CASE-2024-001',
        'created_at': '2024-12-09T14:30:00Z',
        'patient': {
            'name': 'Ahmet',
            'surname': 'Yılmaz',
            'age': 45,
            'phone': '0532 123 4567',
            'tc_no': '12345678901',
            'complaint': 'Göğüs ağrısı'
        },
        'location': {
            'address': 'Atatürk Cad. No:15, Ankara'
        },
        'assigned_team': {
            'vehicle_id': 'AMB-01'
        }
    }
    
    test_form_data = {
        'healmedyProtocol': 'PROT-123',
        'arrivalTime': '14:45',
        'departureTime': '15:30',
        'hospitalArrivalTime': '15:50'
    }
    
    try:
        result = excel_form_to_pdf_with_data(
            excel_path, 
            pdf_path, 
            case_data=test_case_data,
            form_data=test_form_data
        )
        print(f"PDF generated successfully: {result}")
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

